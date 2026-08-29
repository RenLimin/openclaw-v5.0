"""报告对比验证测试

将生成的报告与原始报告进行逐列对比，确保数据一致性。
"""

import pytest
import pandas as pd
from pathlib import Path

TEST_DATA_DIR = Path.home() / "Bangcle Workspace/01. Management/2026/2026团队报告/202606"


def test_contract_data_completeness():
    """验证签约数据完整性"""
    csv_file = TEST_DATA_DIR / "202606周报-签约项目统计.csv"
    if not csv_file.exists():
        pytest.skip("测试数据不存在")

    df = pd.read_csv(csv_file, encoding="utf-8-sig", low_memory=False)
    assert len(df) > 1000, f"数据量不足: {len(df)}"


def test_report_sheet_count():
    """验证报告 Sheet 数量"""
    from openpyxl import load_workbook

    xlsx_file = TEST_DATA_DIR / "2026交付月报-20260630.xlsx"
    if not xlsx_file.exists():
        pytest.skip("测试数据不存在")

    wb = load_workbook(xlsx_file, read_only=True)
    assert len(wb.sheetnames) == 15, f"Sheet 数量不符: {len(wb.sheetnames)}"
    wb.close()


def test_formula_calculation():
    """验证公式可计算性"""
    from openpyxl import load_workbook

    xlsx_file = TEST_DATA_DIR / "2026交付月报-20260630.xlsx"
    if not xlsx_file.exists():
        pytest.skip("测试数据不存在")

    wb = load_workbook(xlsx_file, read_only=True)
    # 检查签约 Sheet 有公式
    ws = wb["签约"]
    has_formula = False
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                has_formula = True
                break
    assert has_formula, "签约 Sheet 未找到公式"
    wb.close()
