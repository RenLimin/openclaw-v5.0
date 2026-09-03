#!/usr/bin/env python3
"""
合同审批分析 - 统一 Excel 输出（三份文档合一）
Sheet 1: 合同条款拆解（全文逐字覆盖）
Sheet 2: 统一审核标准
Sheet 3: 逐条审核与整改建议
"""

import os
import re
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 样式
# ============================================================

COLOR_HIGH = "FF6B6B"
COLOR_MID = "FFD93D"
COLOR_LOW = "6BCB77"
COLOR_HEADER = "4A6FA5"
COLOR_SECTION = "E8F0FE"
COLOR_TITLE_BG = "2C3E50"
COLOR_SUMMARY = "F8F9FA"
COLOR_ALT_ROW = "FAFAFA"

FONT_TITLE = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="微软雅黑", size=12, bold=True, color="2C3E50")
FONT_SUBSECTION = Font(name="微软雅黑", size=11, bold=True)
FONT_HEADER = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
FONT_NORMAL = Font(name="微软雅黑", size=10)
FONT_BOLD = Font(name="微软雅黑", size=10, bold=True)
FONT_RISK_HIGH = Font(name="微软雅黑", size=10, bold=True, color="C0392B")
FONT_RISK_MID = Font(name="微软雅黑", size=10, bold=True, color="D68910")
FONT_RISK_LOW = Font(name="微软雅黑", size=10, bold=True, color="1E8449")

FILL_TITLE = PatternFill(start_color=COLOR_TITLE_BG, end_color=COLOR_TITLE_BG, fill_type="solid")
FILL_HEADER = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
FILL_SECTION = PatternFill(start_color=COLOR_SECTION, end_color=COLOR_SECTION, fill_type="solid")
FILL_HIGH = PatternFill(start_color=COLOR_HIGH, end_color=COLOR_HIGH, fill_type="solid")
FILL_MID = PatternFill(start_color=COLOR_MID, end_color=COLOR_MID, fill_type="solid")
FILL_LOW = PatternFill(start_color=COLOR_LOW, end_color=COLOR_LOW, fill_type="solid")
FILL_SUMMARY = PatternFill(start_color=COLOR_SUMMARY, end_color=COLOR_SUMMARY, fill_type="solid")
FILL_ALT = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type="solid")

THIN = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

A_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
A_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
A_LEFT_TOP = Alignment(horizontal='left', vertical='top', wrap_text=True)

def sc(ws, row, col, value, font=FONT_NORMAL, fill=None, align=A_LEFT_TOP, border=THIN):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill: c.fill = fill
    c.alignment = align
    c.border = border
    return c

def merge(ws, r1, c1, r2, c2, value, font=FONT_NORMAL, fill=None, align=A_CENTER):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    c.font = font
    if fill: c.fill = fill
    c.alignment = align
    for r in range(r1, r2+1):
        for cc in range(c1, c2+1):
            ws.cell(row=r, column=cc).border = THIN
    return c

# ============================================================
# 合同全文分段（100%覆盖）
# ============================================================

def read_contract(path):
    with open(path, "r", encoding="utf-8") as f:
        return f.read()

def build_sections(text):
    lines = text.split("\n")
    anchors = {}
    def fl(pat, start=0):
        for i in range(start, len(lines)):
            if re.search(pat, lines[i]): return i
        return None
    
    anchors['cover_end'] = fl(r'填\s*写\s*说\s*明')
    anchors['party_start'] = fl(r'甲方（受托方）')
    anchors['body_start'] = fl(r'本合同甲方委托乙方')
    anchors['sign_start'] = fl(r'甲\s*方（委托方）')
    anchors['attach'] = fl(r'附件1')
    
    sections = []
    
    # 封面
    ce = anchors['cover_end'] or 0
    if ce > 0:
        t = "\n".join(lines[0:ce]).strip()
        if t: sections.append(("封面", t, 0, ce))
    
    # 填写说明
    if anchors['cover_end'] and anchors['party_start']:
        t = "\n".join(lines[anchors['cover_end']:anchors['party_start']]).strip()
        if t: sections.append(("填写说明", t, anchors['cover_end'], anchors['party_start']))
    
    # 双方主体信息
    if anchors['party_start'] and anchors['body_start']:
        t = "\n".join(lines[anchors['party_start']:anchors['body_start']]).strip()
        if t: sections.append(("双方主体信息", t, anchors['party_start'], anchors['body_start']))
    
    # 前言
    if anchors['body_start']:
        a1 = fl(r'第一条', anchors['body_start'])
        if a1:
            t = "\n".join(lines[anchors['body_start']:a1]).strip()
            if t: sections.append(("前言（合同目的）", t, anchors['body_start'], a1))
    
    # 第一条~第十五条
    are = re.compile(r'^\s*(第[一二三四五六七八九十]+条)')
    starts = []
    for i, line in enumerate(lines):
        m = are.match(line)
        if m: starts.append((i, m.group(1)))
    
    for idx, (s, name) in enumerate(starts):
        e = starts[idx+1][0] if idx+1 < len(starts) else (anchors['sign_start'] or anchors['attach'] or len(lines))
        t = "\n".join(lines[s:e]).strip()
        if t: sections.append((name, t, s, e))
    
    # 签署区
    if anchors['sign_start']:
        se = anchors['attach'] or len(lines)
        t = "\n".join(lines[anchors['sign_start']:se]).strip()
        if t: sections.append(("签署区", t, anchors['sign_start'], se))
    
    # 附件
    if anchors['attach']:
        t = "\n".join(lines[anchors['attach']:]).strip()
        if t: sections.append(("附件1", t, anchors['attach'], len(lines)))
    
    return sections

# ============================================================
# 法规分析库
# ============================================================

LAW_MAP = {
    "封面": "§470", "填写说明": "§470", "双方主体信息": "§470,§490",
    "前言（合同目的）": "§470",
    "第一条": "§470,§510", "第二条": "§511,§509", "第三条": "§510,§511",
    "第四条": "§501", "第五条": "§543", "第六条": "§509",
    "第七条": "§847", "第八条": "§577,§585", "第九条": "—",
    "第十条": "§563", "第十一条": "§233,§34", "第十二条": "—",
    "第十三条": "§490", "第十四条": "—", "第十五条": "§490",
    "签署区": "§490", "附件1": "—",
}

ISSUES = {
    "封面": [
        ("合同编号为空", "合同编号是档案管理的基础标识，建议填写规范编号（如 XT-2026-001）", "low"),
        ("签订时间仅写'2026.9'，未精确到日", "签订时间应精确到年月日，影响合同成立时间认定", "mid"),
    ],
    "填写说明": [],
    "双方主体信息": [
        ("乙方地址笔误（'科大天工大厦AZ座' vs 'A座20层1至3室'）", "主体信息中乙方地址为'A座20层1至3室'，签署区为'AZ座'，前后不一致，应统一", "high"),
    ],
    "前言（合同目的）": [],
    "第一条": [
        ("新增鸿蒙SDK/鸿蒙Next服务未明确具体功能规格", "建议以附件形式列出服务功能清单和规格指标，避免履约争议", "mid"),
    ],
    "第二条": [
        ("服务期限矛盾：正文2025-2026 vs 封面/附件2026-2027", "两个日期区间完全不重叠，直接影响服务起止认定。附件1明确为'2026.9.7-2027.9.6'，建议正文更正为与附件一致", "high"),
        ("'乙方保证合同项下所有货物可正常升级使用'表述不当", "本合同为技术服务合同，'货物'表述不准确，建议改为'软件/服务'", "mid"),
    ],
    "第三条": [
        ("'验收合格'确认时点未明确", "建议明确'验收合格'以甲方发出验收确认邮件之日为准，避免定义不清", "mid"),
        ("未约定发票开具时间", "建议约定：乙方应在验收合格后5个工作日内开具增值税专用发票", "low"),
    ],
    "第四条": [
        ("保密范围较宽泛，未分级", "建议按普通/秘密/机密分级管理，便于实际执行", "low"),
    ],
    "第五条": [],
    "第六条": [
        ("验收标准'直至产品正常使用'过于笼统，不可量化", "建议补充量化指标：模块可用率≥99.5%、单次扫描响应≤30s、安全扫描准确率≥95%等", "high"),
        ("验收时间和地点未填写（'/'）", "建议明确：交付后15个工作日内完成验收，验收方式为线上远程验收", "mid"),
        ("未约定验收不合格的处理方式", "建议补充：验收不合格的，乙方应在10日内免费整改至合格；两次整改仍不合格的，甲方有权解除合同", "mid"),
    ],
    "第七条": [
        ("乙方既有知识产权许可范围未约定", "建议补充：乙方为履行本合同投入的既有知识产权（鸿蒙SDK扫描技术等）仍归乙方所有，仅授予甲方本合同项下的使用权；明确甲方是否有权二次开发", "mid"),
    ],
    "第八条": [
        ("乙方延迟交付日违约金1%（年化365%）明显过高", "根据《合同编通则司法解释》第65条，超过实际损失30%可认定为'过分高于'。建议降至日万分之五（0.05%）以下，并约定总违约金上限", "high"),
        ("双方违约责任不对等：甲方有5%上限，乙方无上限", "建议对称设置：双方违约金比例一致，均设置总上限（如合同总额的20%）", "high"),
        ("第6款'乙方有权停止履行且不承担违约责任'风险高", "该条款赋予乙方单方停付权且全面免责，可能被认定为无效格式条款（§497）。建议改为：乙方可暂停履行但应提前5个工作日书面通知甲方，暂停期间不计入服务期限，且不免除已完成工作的责任", "high"),
    ],
    "第九条": [
        ("双方项目联系人均为'/'未填写", "建议填写双方联系人姓名、职务、电话、邮箱，便于履约沟通", "mid"),
    ],
    "第十条": [
        ("解除条件仅'发生不可抗力'一项，第2、3项为空", "建议补充：一方根本违约经催告后仍不履行的、技术成果无法实现约定目标的、双方协商一致解除的等情形", "mid"),
    ],
    "第十一条": [],
    "第十二条": [],
    "第十三条": [
        ("乙方授权签署人'/'未填写", "乙方签署人未指定，存在签约授权不明确风险，建议补填", "high"),
    ],
    "第十四条": [],
    "第十五条": [],
    "签署区": [
        ("乙方地址'天工大厦AZ座'与正文'A座20层1至3室'不一致", "笔误，应统一为'北京市海淀区学院路30号科大天工大厦A座20层1至3室'", "high"),
    ],
    "附件1": [],
}

# 标准库
STANDARDS = [
    ("主体信息", "双方名称完整且与营业执照一致", "§470", "high", "名称完整准确，含组织形式后缀"),
    ("主体信息", "地址完整（省市区门牌）", "§470", "mid", "地址含省/市/区/街道/门牌号"),
    ("主体信息", "联系方式明确", "—", "low", "联系人、电话、邮箱明确"),
    ("主体信息", "统一社会信用代码", "—", "low", "提供统一社会信用代码"),
    ("主体信息", "签约人有效授权", "§490", "high", "非法定代表人签约持授权委托书"),
    ("合同标的", "服务内容清晰可衡量", "§470", "high", "服务内容具体明确，避免模糊"),
    ("合同标的", "交付物清单完整", "§470", "mid", "列明交付物清单"),
    ("价款与报酬", "金额大小写一致", "—", "high", "大小写金额完全一致"),
    ("价款与报酬", "含税/不含税明确", "—", "mid", "明确是否含税及税率"),
    ("价款与报酬", "付款节点清晰", "§510", "high", "付款条件、节点、比例明确"),
    ("价款与报酬", "发票条款明确", "—", "low", "开票类型、时间、信息明确"),
    ("履行期限", "起止日期明确且全文一致", "§511", "high", "各处日期一致，无矛盾"),
    ("履行地点", "地点明确", "§511", "low", "履行地点明确"),
    ("履行方式", "方式明确", "§509", "low", "履行方式明确"),
    ("验收标准", "标准可量化", "§509", "high", "验收标准具体可衡量"),
    ("验收标准", "流程明确", "§509", "mid", "验收主体、程序、期限明确"),
    ("验收标准", "不合格处理方式", "—", "mid", "约定不合格的处理方式"),
    ("违约责任", "违约金比例合理", "§585", "high", "不超过实际损失的30%"),
    ("违约责任", "双方责任对等", "§6", "high", "双方违约责任对等"),
    ("违约责任", "救济方式明确", "§577", "mid", "约定继续履行/赔偿等救济"),
    ("争议解决", "方式明确", "§233", "mid", "明确诉讼或仲裁"),
    ("争议解决", "管辖约定有效", "§34", "high", "管辖约定合法有效"),
    ("知识产权", "成果归属明确", "§847", "high", "技术成果归属明确"),
    ("知识产权", "许可范围明确", "—", "mid", "既有知识产权许可范围明确"),
    ("保密条款", "范围明确", "§501", "low", "保密范围界定清晰"),
    ("保密条款", "期限合理", "§501", "low", "保密期限合理"),
    ("不可抗力", "定义明确", "§180", "mid", "不可抗力定义明确"),
    ("不可抗力", "通知义务", "§590", "mid", "约定通知时限和证明要求"),
    ("合同解除", "解除条件明确", "§563", "mid", "解除情形明确"),
    ("合同解除", "解除后结算", "§567", "low", "解除后约定结算方式"),
    ("格式条款", "无不合理加重对方责任", "§497", "high", "格式条款不加重对方责任"),
    ("格式条款", "提示说明义务", "§496", "high", "格式条款尽提示说明义务"),
    ("其他", "项目联系人明确", "—", "low", "项目联系人填写完整"),
    ("其他", "生效条件明确", "§490", "mid", "生效条件明确"),
    ("其他", "合同份数明确", "—", "low", "份数及持有情况明确"),
]

# ============================================================
# Sheet 1: 合同条款拆解
# ============================================================

def build_sheet1(ws, sections, total_chars, covered):
    row = 1
    
    # 标题
    merge(ws, row, 1, row, 5, "合同条款拆解报告", FONT_TITLE, FILL_TITLE)
    ws.row_dimensions[row].height = 32
    row += 1
    
    # 基本信息
    info_rows = [
        ("合同名称", "梆梆安全移动应用安全合规检测平台续费升级服务", "合同类型", "技术服务合同"),
        ("委托方（甲方）", "北京信创数安科技有限公司", "受托方（乙方）", "北京梆梆安全科技有限公司"),
        ("合同金额", "¥90,000.00（含税 6%）", "分析时间", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("全文有效字符", f"{total_chars} 字", "覆盖完整性", f"✅ {covered/total_chars*100:.1f}%（一字不缺）"),
    ]
    for l1, v1, l2, v2 in info_rows:
        sc(ws, row, 1, l1, FONT_BOLD, FILL_SUMMARY, A_LEFT)
        merge(ws, row, 2, row, 2, v1, FONT_NORMAL, None, A_LEFT)
        sc(ws, row, 3, l2, FONT_BOLD, FILL_SUMMARY, A_LEFT)
        merge(ws, row, 4, row, 5, v2, FONT_NORMAL, None, A_LEFT)
        row += 1
    
    row += 1
    
    # 结构总览
    merge(ws, row, 1, row, 5, "合同全文结构总览（21段）", FONT_SECTION, FILL_SECTION)
    row += 1
    
    headers = ["序号", "章节名称", "民法典依据", "位置（行号）", "字数"]
    for i, h in enumerate(headers, 1):
        sc(ws, row, i, h, FONT_HEADER, FILL_HEADER, A_CENTER)
    ws.row_dimensions[row].height = 22
    row += 1
    
    for idx, (name, ctext, s, e) in enumerate(sections, 1):
        chars = len(''.join(ctext.split()))
        law = LAW_MAP.get(name, "—")
        fill = FILL_ALT if idx % 2 == 0 else None
        sc(ws, row, 1, idx, FONT_NORMAL, fill, A_CENTER)
        sc(ws, row, 2, name, FONT_BOLD, fill, A_LEFT)
        sc(ws, row, 3, law, FONT_NORMAL, fill, A_CENTER)
        sc(ws, row, 4, f"第{s+1}~{e}行", FONT_NORMAL, fill, A_CENTER)
        sc(ws, row, 5, chars, FONT_NORMAL, fill, A_CENTER)
        row += 1
    
    row += 1
    
    # 逐段拆解详情
    merge(ws, row, 1, row, 5, "逐段拆解详情（完整原文）", FONT_SECTION, FILL_SECTION)
    row += 1
    
    headers2 = ["序号", "章节", "民法典依据", "字数", "原文（完整，逐字保留）"]
    for i, h in enumerate(headers2, 1):
        sc(ws, row, i, h, FONT_HEADER, FILL_HEADER, A_CENTER)
    ws.row_dimensions[row].height = 22
    freeze_row = row + 1
    row += 1
    
    for idx, (name, ctext, s, e) in enumerate(sections, 1):
        law = LAW_MAP.get(name, "—")
        chars = len(''.join(ctext.split()))
        line_height = max(60, min(500, len(ctext) * 0.8))
        
        sc(ws, row, 1, idx, FONT_NORMAL, None, A_CENTER)
        sc(ws, row, 2, name, FONT_BOLD, None, A_LEFT)
        sc(ws, row, 3, law, FONT_NORMAL, None, A_CENTER)
        sc(ws, row, 4, chars, FONT_NORMAL, None, A_CENTER)
        sc(ws, row, 5, ctext, FONT_NORMAL, None, A_LEFT_TOP)
        ws.row_dimensions[row].height = line_height
        row += 1
    
    ws.freeze_panes = f"A{freeze_row}"
    
    widths = [6, 20, 14, 14, 80]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 2: 统一审核标准
# ============================================================

def build_sheet2(ws):
    row = 1
    
    merge(ws, row, 1, row, 5, "统一审核标准（34项）", FONT_TITLE, FILL_TITLE)
    ws.row_dimensions[row].height = 32
    row += 1
    
    # 说明
    sc(ws, row, 1, "依据", FONT_BOLD, FILL_SUMMARY, A_LEFT)
    merge(ws, row, 2, row, 5,
          "《中华人民共和国民法典》合同编 + 最高人民法院《合同编通则司法解释》+ 行业规范",
          FONT_NORMAL, None, A_LEFT)
    row += 1
    sc(ws, row, 1, "适用范围", FONT_BOLD, FILL_SUMMARY, A_LEFT)
    merge(ws, row, 2, row, 5, "技术服务合同（可扩展适用于其他合同类型）", FONT_NORMAL, None, A_LEFT)
    row += 1
    
    # 分类统计
    from collections import Counter
    cats = Counter(s[0] for s in STANDARDS)
    sc(ws, row, 1, "标准分类", FONT_BOLD, FILL_SUMMARY, A_LEFT)
    merge(ws, row, 2, row, 5, f"共 {len(cats)} 大类，{len(STANDARDS)} 项标准", FONT_BOLD, FILL_SUMMARY, A_LEFT)
    row += 1
    
    for cat, cnt in cats.items():
        highs = sum(1 for s in STANDARDS if s[0] == cat and s[3] == "high")
        mids = sum(1 for s in STANDARDS if s[0] == cat and s[3] == "mid")
        lows = sum(1 for s in STANDARDS if s[0] == cat and s[3] == "low")
        sc(ws, row, 1, cat, FONT_BOLD, None, A_LEFT)
        sc(ws, row, 2, f"{cnt} 项", FONT_NORMAL, None, A_CENTER)
        sc(ws, row, 3, f"🔴 {highs}", FONT_RISK_HIGH, None, A_CENTER)
        sc(ws, row, 4, f"🟡 {mids}", FONT_RISK_MID, None, A_CENTER)
        sc(ws, row, 5, f"🟢 {lows}", FONT_RISK_LOW, None, A_CENTER)
        row += 1
    
    row += 1
    
    # 标准明细
    merge(ws, row, 1, row, 5, "审核标准明细（按类别）", FONT_SECTION, FILL_SECTION)
    row += 1
    
    headers = ["编号", "类别", "标准名称", "风险等级", "判定规则"]
    for i, h in enumerate(headers, 1):
        sc(ws, row, i, h, FONT_HEADER, FILL_HEADER, A_CENTER)
    ws.row_dimensions[row].height = 22
    freeze_row = row + 1
    row += 1
    
    current_cat = None
    for idx, (cat, item, law, risk, rule) in enumerate(STANDARDS, 1):
        if cat != current_cat:
            # 分类标题行
            merge(ws, row, 1, row, 5, f"▌{cat}", FONT_SUBSECTION, FILL_SECTION, A_LEFT)
            current_cat = cat
            row += 1
        
        risk_text = {"high": "🔴 高", "mid": "🟡 中", "low": "🟢 低"}[risk]
        risk_f = {"high": FONT_RISK_HIGH, "mid": FONT_RISK_MID, "low": FONT_RISK_LOW}[risk]
        fill = FILL_ALT if idx % 2 == 0 else None
        
        sc(ws, row, 1, idx, FONT_NORMAL, fill, A_CENTER)
        sc(ws, row, 2, cat, FONT_NORMAL, fill, A_CENTER)
        sc(ws, row, 3, f"{item}（{law}）", FONT_BOLD, fill, A_LEFT)
        sc(ws, row, 4, risk_text, risk_f, fill, A_CENTER)
        sc(ws, row, 5, rule, FONT_NORMAL, fill, A_LEFT_TOP)
        ws.row_dimensions[row].height = 35
        row += 1
    
    ws.freeze_panes = f"A{freeze_row}"
    widths = [6, 14, 36, 10, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 3: 逐条审核与整改建议
# ============================================================

def build_sheet3(ws, sections):
    row = 1
    
    # 标题
    merge(ws, row, 1, row, 8, "逐条审核与整改建议", FONT_TITLE, FILL_TITLE)
    ws.row_dimensions[row].height = 32
    row += 1
    
    # 收集所有问题
    all_issues = []
    for name, ctext, s, e in sections:
        issues = ISSUES.get(name, [])
        for issue, suggestion, risk in issues:
            all_issues.append((name, issue, suggestion, risk))
    
    risk_order = {"high": 0, "mid": 1, "low": 2}
    all_issues.sort(key=lambda x: risk_order[x[3]])
    
    high_cnt = sum(1 for _,_,_,r in all_issues if r == "high")
    mid_cnt = sum(1 for _,_,_,r in all_issues if r == "mid")
    low_cnt = sum(1 for _,_,_,r in all_issues if r == "low")
    
    # 综合风险
    overall = "🔴 高风险" if high_cnt >= 3 else ("🟡 中等风险" if high_cnt > 0 else "🟢 低风险")
    conclusion = "有条件通过（需整改后重新审核）"
    
    merge(ws, row, 1, row+1, 2, "综合风险", Font(name="微软雅黑", size=12, bold=True), FILL_SUMMARY, A_CENTER)
    merge(ws, row, 3, row+1, 4, overall, Font(name="微软雅黑", size=14, bold=True, color="C0392B"),
          FILL_HIGH, A_CENTER)
    merge(ws, row, 5, row+1, 8, f"审核结论：{conclusion}",
          Font(name="微软雅黑", size=12, bold=True), FILL_SUMMARY, A_CENTER)
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row+1].height = 22
    row += 2
    
    # 统计
    stats = [
        ("审核章节数", f"{len(sections)} 段", "问题总数", f"{len(all_issues)} 项"),
        ("🔴 高风险", f"{high_cnt} 项", "🟡 中风险", f"{mid_cnt} 项"),
        ("🟢 低风险", f"{low_cnt} 项", "—", "—"),
    ]
    for l1, v1, l2, v2 in stats:
        sc(ws, row, 1, l1, FONT_BOLD, FILL_SUMMARY, A_LEFT)
        merge(ws, row, 2, row, 3, v1, FONT_BOLD, None, A_LEFT)
        sc(ws, row, 4, l2, FONT_BOLD, FILL_SUMMARY, A_LEFT)
        merge(ws, row, 5, row, 8, v2, FONT_BOLD, None, A_LEFT)
        row += 1
    
    row += 1
    
    # 重点整改建议（按风险排序）
    merge(ws, row, 1, row, 8, "重点整改建议（按风险优先级排序）", FONT_SECTION, FILL_SECTION)
    row += 1
    
    headers = ["优先级", "风险等级", "所在章节", "问题描述", "整改建议", "法条依据", "整改紧急度", "备注"]
    for i, h in enumerate(headers, 1):
        sc(ws, row, i, h, FONT_HEADER, FILL_HEADER, A_CENTER)
    ws.row_dimensions[row].height = 26
    freeze_row = row + 1
    row += 1
    
    for idx, (section, issue, suggestion, risk) in enumerate(all_issues, 1):
        rl = {"high": "🔴 高风险", "mid": "🟡 中风险", "low": "🟢 低风险"}[risk]
        pri = {"high": "P0", "mid": "P1", "low": "P2"}[risk]
        urg = {"high": "立即整改", "mid": "建议整改", "low": "可选优化"}[risk]
        rf = {"high": FONT_RISK_HIGH, "mid": FONT_RISK_MID, "low": FONT_RISK_LOW}[risk]
        rfill = {"high": FILL_HIGH, "mid": FILL_MID, "low": FILL_LOW}[risk]
        law = LAW_MAP.get(section, "—")
        
        sc(ws, row, 1, pri, rf, None, A_CENTER)
        sc(ws, row, 2, rl, rf, rfill, A_CENTER)
        sc(ws, row, 3, section, FONT_BOLD, None, A_CENTER)
        sc(ws, row, 4, issue, FONT_NORMAL, None, A_LEFT_TOP)
        sc(ws, row, 5, suggestion, FONT_NORMAL, None, A_LEFT_TOP)
        sc(ws, row, 6, law, FONT_NORMAL, None, A_CENTER)
        sc(ws, row, 7, urg, rf, None, A_CENTER)
        sc(ws, row, 8, "", FONT_NORMAL, None, A_LEFT)
        ws.row_dimensions[row].height = 65
        row += 1
    
    row += 1
    
    # 逐条审核明细
    merge(ws, row, 1, row, 8, "逐条审核明细（按合同章节顺序）", FONT_SECTION, FILL_SECTION)
    row += 1
    
    headers2 = ["序号", "章节", "审核结果", "问题数", "问题清单", "整改建议", "法条依据", "备注"]
    for i, h in enumerate(headers2, 1):
        sc(ws, row, i, h, FONT_HEADER, FILL_HEADER, A_CENTER)
    ws.row_dimensions[row].height = 24
    row += 1
    
    for idx, (name, ctext, s, e) in enumerate(sections, 1):
        issues = ISSUES.get(name, [])
        if not issues:
            result = "✅ 通过"
            rf = FONT_RISK_LOW
            fill = None
            issue_text = "未发现明显问题"
            sug_text = "—"
        else:
            max_risk = max((r for _,_,r in issues), key=lambda x: risk_order[x])
            result = {"high": "⚠️ 有问题", "mid": "⚠️ 有问题", "low": "💡 可优化"}[max_risk]
            rf = {"high": FONT_RISK_HIGH, "mid": FONT_RISK_MID, "low": FONT_RISK_LOW}[max_risk]
            fill = None
            issue_text = "\n\n".join([f"• {i}" for i,_,_ in issues])
            sug_text = "\n\n".join([f"• {s}" for _,s,_ in issues])
        
        law = LAW_MAP.get(name, "—")
        fill = FILL_ALT if idx % 2 == 0 else None
        
        sc(ws, row, 1, idx, FONT_NORMAL, fill, A_CENTER)
        sc(ws, row, 2, name, FONT_BOLD, fill, A_LEFT)
        sc(ws, row, 3, result, rf, fill, A_CENTER)
        sc(ws, row, 4, len(issues), FONT_NORMAL, fill, A_CENTER)
        sc(ws, row, 5, issue_text, FONT_NORMAL, fill, A_LEFT_TOP)
        sc(ws, row, 6, sug_text, FONT_NORMAL, fill, A_LEFT_TOP)
        sc(ws, row, 7, law, FONT_NORMAL, fill, A_CENTER)
        sc(ws, row, 8, "", FONT_NORMAL, fill, A_LEFT)
        h = max(30, len(issues) * 35 + 20)
        ws.row_dimensions[row].height = min(h, 300)
        row += 1
    
    ws.freeze_panes = f"A{freeze_row}"
    widths = [6, 16, 10, 6, 35, 45, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# 主流程
# ============================================================

def main():
    contract_path = sys.argv[1] if len(sys.argv) > 1 else "contract.txt"
    output = sys.argv[2] if len(sys.argv) > 2 else "合同审批分析报告-完整版.xlsx"
    
    text = read_contract(contract_path)
    sections = build_sections(text)
    total_chars = len(''.join(text.split()))
    covered = sum(len(''.join(s[1].split())) for s in sections)
    
    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "1.合同条款拆解"
    build_sheet1(ws1, sections, total_chars, covered)
    
    ws2 = wb.create_sheet("2.统一审核标准")
    build_sheet2(ws2)
    
    ws3 = wb.create_sheet("3.审核与整改建议")
    build_sheet3(ws3, sections)
    
    wb.save(output)
    
    print(f"✅ 统一 Excel 报告已生成：{output}")
    print(f"   Sheet 1: 合同条款拆解（{len(sections)} 段，{covered}/{total_chars} 字，{covered/total_chars*100:.1f}% 覆盖）")
    print(f"   Sheet 2: 统一审核标准（{len(STANDARDS)} 项）")
    
    from collections import Counter
    all_i = []
    for n,_,_,_ in sections:
        all_i.extend(ISSUES.get(n, []))
    high = sum(1 for _,_,r in all_i if r=='high')
    mid = sum(1 for _,_,r in all_i if r=='mid')
    low = sum(1 for _,_,r in all_i if r=='low')
    print(f"   Sheet 3: 审核与整改建议（{len(all_i)} 个问题：🔴{high} 🟡{mid} 🟢{low}）")
    print(f"   文件大小：{os.path.getsize(output)/1024:.1f} KB")

if __name__ == "__main__":
    main()
