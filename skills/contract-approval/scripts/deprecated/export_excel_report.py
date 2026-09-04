#!/usr/bin/env python3
"""
合同审批报告导出 - Excel 格式 (v3)
组件: SCA-001 (L4)

从 contract_auditor 的 AuditReport 生成格式化的 Excel 报告。
三个 Sheet：
1. 合同条款解析 - 基本信息 + 23类条款解析
2. 审批标准库 - 分类统计 + 43项标准明细
3. 审批结果与建议 - 风险分级 + 整改建议（按优先级排序）

用法:
    python3 export_excel_report.py --file <合同文本文件> --output <输出xlsx>
    python3 export_excel_report.py --contract-id <ID> --output <输出xlsx>
"""

import argparse
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 同目录导入
sys.path.insert(0, os.path.dirname(__file__))
from contract_auditor import audit_contract, AuditReport
from contract_parser import parse_contract
from audit_standard import AUDIT_CRITERIA, get_all_categories

# ============================================================
# 样式定义
# ============================================================

# 颜色
COLOR_HIGH_RISK = "FF6B6B"
COLOR_MID_RISK = "FFD93D"
COLOR_LOW_RISK = "6BCB77"
COLOR_HEADER = "4A6FA5"
COLOR_SECTION = "E8F0FE"
COLOR_TITLE_BG = "2C3E50"
COLOR_SUMMARY_BG = "F8F9FA"

# 字体
FONT_TITLE = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="微软雅黑", size=11, bold=True)
FONT_HEADER = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
FONT_NORMAL = Font(name="微软雅黑", size=10)
FONT_BOLD = Font(name="微软雅黑", size=10, bold=True)
FONT_RISK_HIGH = Font(name="微软雅黑", size=10, bold=True, color="C0392B")
FONT_RISK_MID = Font(name="微软雅黑", size=10, bold=True, color="D68910")
FONT_RISK_LOW = Font(name="微软雅黑", size=10, bold=True, color="1E8449")

# 填充
FILL_TITLE = PatternFill(start_color=COLOR_TITLE_BG, end_color=COLOR_TITLE_BG, fill_type="solid")
FILL_HEADER = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
FILL_SECTION = PatternFill(start_color=COLOR_SECTION, end_color=COLOR_SECTION, fill_type="solid")
FILL_HIGH = PatternFill(start_color=COLOR_HIGH_RISK, end_color=COLOR_HIGH_RISK, fill_type="solid")
FILL_MID = PatternFill(start_color=COLOR_MID_RISK, end_color=COLOR_MID_RISK, fill_type="solid")
FILL_LOW = PatternFill(start_color=COLOR_LOW_RISK, end_color=COLOR_LOW_RISK, fill_type="solid")
FILL_SUMMARY = PatternFill(start_color=COLOR_SUMMARY_BG, end_color=COLOR_SUMMARY_BG, fill_type="solid")

# 边框
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# 对齐
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal='left', vertical='top', wrap_text=True)


# ============================================================
# 风险分级映射
# ============================================================

def normalize_risk(risk: str) -> str:
    """统一风险等级命名：high / medium / low"""
    if risk in ("high", "HIGH", "高", "高风险"):
        return "high"
    elif risk in ("medium", "mid", "MEDIUM", "中", "中风险"):
        return "mid"
    elif risk in ("low", "LOW", "低", "低风险"):
        return "low"
    return "low"

def risk_label(level: str) -> str:
    return {"high": "🔴 高风险", "mid": "🟡 中风险", "low": "🟢 低风险"}[level]

def risk_fill(level: str):
    return {"high": FILL_HIGH, "mid": FILL_MID, "low": FILL_LOW}[level]

def risk_font(level: str):
    return {"high": FONT_RISK_HIGH, "mid": FONT_RISK_MID, "low": FONT_RISK_LOW}[level]

def priority_label(level: str) -> str:
    return {"high": "P0", "mid": "P1", "low": "P2"}[level]

def urgency_label(level: str) -> str:
    return {"high": "立即整改", "mid": "建议整改", "low": "可选优化"}[level]

# ============================================================
# 工具函数
# ============================================================

def set_cell(ws, row, col, value, font=FONT_NORMAL, fill=None,
             alignment=ALIGN_LEFT_TOP, border=THIN_BORDER):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    cell.border = border
    return cell

def merge_style(ws, r1, c1, r2, c2, value, font=FONT_NORMAL,
                fill=None, alignment=ALIGN_CENTER):
    """合并单元格并应用样式"""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
    return cell


# ============================================================
# 报告生成
# ============================================================

def generate_excel_report(report: AuditReport, output_path: str,
                          contract_info: dict = None):
    """
    从 AuditReport 生成 Excel 报告

    Args:
        report: AuditReport 对象
        output_path: 输出 xlsx 路径
        contract_info: 合同基本信息字典（可选）
    """
    wb = Workbook()
    contract_info = contract_info or {}

    # ====== Sheet 1: 合同条款解析 ======
    ws1 = wb.active
    ws1.title = "1.合同条款解析"
    _build_sheet_clauses(ws1, report, contract_info)

    # ====== Sheet 2: 审批标准库 ======
    ws2 = wb.create_sheet("2.审批标准库")
    _build_sheet_standards(ws2)

    # ====== Sheet 3: 审批结果与建议 ======
    ws3 = wb.create_sheet("3.审批结果与建议")
    _build_sheet_results(ws3, report, contract_info)

    wb.save(output_path)
    return output_path


def _build_sheet_clauses(ws, report: AuditReport, info: dict):
    """Sheet 1: 合同条款解析"""
    row = 1

    # 标题
    merge_style(ws, row, 1, row, 6, "销售合同条款解析报告", FONT_TITLE, FILL_TITLE)
    ws.row_dimensions[row].height = 30
    row += 1

    # 合同基本信息
    merge_style(ws, row, 1, row, 6, "合同基本信息", FONT_SUBTITLE, FILL_SECTION)
    row += 1

    info_rows = [
        ("合同名称", info.get("title", "—"), "合同类型", info.get("contract_type", "技术服务合同")),
        ("甲方", info.get("party_a", "—"), "乙方", info.get("party_b", "—")),
        ("合同金额", info.get("amount", "—"), "服务期限", info.get("period", "—")),
        ("签署日期", info.get("sign_date", "待签署"), "审核时间", report.audit_time[:19].replace("T", " ")),
    ]
    for l1, v1, l2, v2 in info_rows:
        set_cell(ws, row, 1, l1, FONT_BOLD, FILL_SUMMARY, ALIGN_LEFT)
        merge_style(ws, row, 2, row, 3, v1, FONT_NORMAL, None, ALIGN_LEFT)
        set_cell(ws, row, 4, l2, FONT_BOLD, FILL_SUMMARY, ALIGN_LEFT)
        merge_style(ws, row, 5, row, 6, v2, FONT_NORMAL, None, ALIGN_LEFT)
        row += 1

    row += 1  # 空行

    # 条款明细标题
    merge_style(ws, row, 1, row, 6,
                "条款解析明细（按《民法典》合同编分类）",
                FONT_SUBTITLE, FILL_SECTION)
    row += 1

    # 表头
    headers = ["编号", "条款类别", "法条依据", "原文摘要", "解析说明", "风险提示"]
    for i, h in enumerate(headers, 1):
        set_cell(ws, row, i, h, FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
    ws.row_dimensions[row].height = 25
    header_row = row
    row += 1

    # 条款数据
    clauses = report.parsed_clauses if hasattr(report, 'parsed_clauses') else []
    for idx, clause in enumerate(clauses, 1):
        risk_text = getattr(clause, 'risk_note', '') or ''
        risk_lvl = normalize_risk(getattr(clause, 'risk_level', 'low'))
        risk_f = risk_font(risk_lvl) if risk_lvl != 'low' else FONT_NORMAL

        set_cell(ws, row, 1, f"C-{idx:03d}", FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 2, getattr(clause, 'category', ''), FONT_BOLD, None, ALIGN_LEFT)
        set_cell(ws, row, 3, getattr(clause, 'law_article', '—'), FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 4, getattr(clause, 'excerpt', '')[:200], FONT_NORMAL, None, ALIGN_LEFT_TOP)
        set_cell(ws, row, 5, getattr(clause, 'summary', '')[:150], FONT_NORMAL, None, ALIGN_LEFT_TOP)
        set_cell(ws, row, 6, risk_text or '无', risk_f, None, ALIGN_LEFT_TOP)
        row += 1

    # 冻结
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    # 列宽
    widths = [8, 22, 12, 40, 22, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_sheet_standards(ws):
    """Sheet 2: 审批标准库"""
    row = 1

    merge_style(ws, row, 1, row, 5,
                f"销售合同审批标准库（{len(AUDIT_CRITERIA)} 项）",
                FONT_TITLE, FILL_TITLE)
    ws.row_dimensions[row].height = 30
    row += 1

    # 分类统计
    merge_style(ws, row, 1, row, 5, "标准分类统计", FONT_SUBTITLE, FILL_SECTION)
    row += 1

    categories = get_all_categories() if callable(get_all_categories) else {}
    if isinstance(categories, dict):
        for cat, count in categories.items():
            set_cell(ws, row, 1, cat, FONT_BOLD, FILL_SUMMARY, ALIGN_LEFT)
            set_cell(ws, row, 2, f"{count} 项", FONT_NORMAL, FILL_SUMMARY, ALIGN_CENTER)
            merge_style(ws, row, 3, row, 5, "", FONT_NORMAL, FILL_SUMMARY, ALIGN_LEFT)
            row += 1
    else:
        # 自己统计
        from collections import Counter
        cat_counts = Counter(getattr(c, 'category', '') for c in AUDIT_CRITERIA)
        for cat, count in cat_counts.items():
            set_cell(ws, row, 1, cat, FONT_BOLD, FILL_SUMMARY, ALIGN_LEFT)
            set_cell(ws, row, 2, f"{count} 项", FONT_NORMAL, FILL_SUMMARY, ALIGN_CENTER)
            merge_style(ws, row, 3, row, 5, "", FONT_NORMAL, FILL_SUMMARY, ALIGN_LEFT)
            row += 1

    row += 1

    # 标准明细
    merge_style(ws, row, 1, row, 5, "审批标准明细", FONT_SUBTITLE, FILL_SECTION)
    row += 1

    headers = ["编号", "类别", "标准名称", "法条依据", "判定规则"]
    for i, h in enumerate(headers, 1):
        set_cell(ws, row, i, h, FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
    ws.row_dimensions[row].height = 25
    header_row = row
    row += 1

    for c in AUDIT_CRITERIA:
        cid = getattr(c, 'id', '')
        cat = getattr(c, 'category', '')
        name = getattr(c, 'item', '')
        law = getattr(c, 'law_article', '—')
        rule = getattr(c, 'standard', '') or getattr(c, 'pass_condition', '')

        set_cell(ws, row, 1, cid, FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 2, cat, FONT_BOLD, None, ALIGN_CENTER)
        set_cell(ws, row, 3, name, FONT_NORMAL, None, ALIGN_LEFT)
        set_cell(ws, row, 4, law, FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 5, rule[:120], FONT_NORMAL, None, ALIGN_LEFT_TOP)
        row += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    widths = [8, 14, 32, 14, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_sheet_results(ws, report: AuditReport, info: dict):
    """Sheet 3: 审批结果与建议"""
    row = 1

    # 标题
    merge_style(ws, row, 1, row, 9, "销售合同审批结果报告", FONT_TITLE, FILL_TITLE)
    ws.row_dimensions[row].height = 35
    row += 1

    # 综合风险
    overall = report.overall_risk if hasattr(report, 'overall_risk') else 'medium'
    risk_lvl = normalize_risk(overall)
    risk_display = {"high": "高风险", "mid": "中等风险", "low": "低风险"}[risk_lvl]
    recommendation = report.recommendation if hasattr(report, 'recommendation') else '有条件通过'

    merge_style(ws, row, 1, row + 1, 2, "综合风险",
                Font(name="微软雅黑", size=12, bold=True), FILL_SUMMARY)
    merge_style(ws, row, 3, row + 1, 5, risk_display,
                Font(name="微软雅黑", size=14, bold=True,
                     color={"high": "C0392B", "mid": "7D6608", "low": "1E8449"}[risk_lvl]),
                risk_fill(risk_lvl))
    merge_style(ws, row, 6, row + 1, 9,
                f"审核结论：{recommendation}",
                Font(name="微软雅黑", size=12, bold=True),
                FILL_SUMMARY, ALIGN_CENTER)
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row + 1].height = 22
    row += 2

    # 统计
    s = report.summary if hasattr(report, 'summary') else {}
    results = report.results if hasattr(report, 'results') else []

    # 统计各风险等级的警告项
    high_count = sum(1 for r in results if r.status == 'warning' and normalize_risk(r.risk_level) == 'high')
    mid_count = sum(1 for r in results if r.status == 'warning' and normalize_risk(r.risk_level) == 'mid')
    low_count = sum(1 for r in results if r.status == 'warning' and normalize_risk(r.risk_level) == 'low')

    stats = [
        ("审核项数", f"{s.get('total', len(results))} 项", "审核时间", report.audit_time[:19].replace("T", " ")),
        ("✅ 通过", f"{s.get('pass', 0)} 项", "⚠️ 警告", f"{s.get('warning', 0)} 项"),
        ("❌ 不通过", f"{s.get('fail', 0)} 项", "N/A", f"{s.get('na', 0)} 项"),
        ("🔴 高风险项", f"{high_count} 项", "🟡 中风险项", f"{mid_count} 项"),
    ]
    for l1, v1, l2, v2 in stats:
        set_cell(ws, row, 1, l1, FONT_BOLD, FILL_SUMMARY, ALIGN_LEFT)
        merge_style(ws, row, 2, row, 3, v1, FONT_BOLD, None, ALIGN_LEFT)
        set_cell(ws, row, 4, l2, FONT_BOLD, FILL_SUMMARY, ALIGN_LEFT)
        merge_style(ws, row, 5, row, 6, v2, FONT_BOLD, None, ALIGN_LEFT)
        set_cell(ws, row, 7, "", FONT_NORMAL, FILL_SUMMARY, ALIGN_LEFT)
        merge_style(ws, row, 8, row, 9, "", FONT_NORMAL, FILL_SUMMARY, ALIGN_LEFT)
        row += 1

    row += 1

    # 重点整改建议
    merge_style(ws, row, 1, row, 9,
                "重点整改建议（按风险等级排序）",
                FONT_SUBTITLE, FILL_SECTION)
    row += 1

    headers = ["优先级", "风险等级", "条款编号", "审批条款", "条款类别",
               "法条依据", "原文摘录", "整改建议", "整改紧急度"]
    for i, h in enumerate(headers, 1):
        set_cell(ws, row, i, h, FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
    ws.row_dimensions[row].height = 28
    detail_header_row = row
    row += 1

    # 只列警告和不通过的项，按风险等级排序
    problem_items = [r for r in results if r.status in ('warning', 'fail')]
    problem_items.sort(key=lambda r: {"high": 0, "mid": 1, "low": 2}[normalize_risk(r.risk_level)])

    for idx, r in enumerate(problem_items, 1):
        rl = normalize_risk(r.risk_level)
        set_cell(ws, row, 1, priority_label(rl), risk_font(rl), None, ALIGN_CENTER)
        set_cell(ws, row, 2, risk_label(rl), risk_font(rl), risk_fill(rl), ALIGN_CENTER)
        set_cell(ws, row, 3, r.criterion_id, FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 4, r.item, FONT_BOLD, None, ALIGN_LEFT)
        set_cell(ws, row, 5, r.category, FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 6, r.law_article, FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 7, (r.evidence or r.issue or '')[:120], FONT_NORMAL, None, ALIGN_LEFT_TOP)
        set_cell(ws, row, 8, (r.suggestion or '')[:300], FONT_NORMAL, None, ALIGN_LEFT_TOP)
        set_cell(ws, row, 9, urgency_label(rl), risk_font(rl), None, ALIGN_CENTER)
        ws.row_dimensions[row].height = 55
        row += 1

    row += 1

    # P0 汇总
    p0_items = [r for r in problem_items if normalize_risk(r.risk_level) == 'high']
    if p0_items:
        merge_style(ws, row, 1, row, 9,
                    f"必须修改项汇总（P0 高风险 {len(p0_items)} 项）",
                    FONT_SUBTITLE, FILL_HIGH)
        row += 1

        sum_headers = ["序号", "审批条款", "条款类别", "问题描述", "建议修改方案"]
        sum_spans = [(1, 1), (2, 2), (3, 3), (4, 6), (7, 9)]
        for (sc, ec), h in zip(sum_spans, sum_headers):
            merge_style(ws, row, sc, row, ec, h, FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
        ws.row_dimensions[row].height = 25
        row += 1

        for idx, r in enumerate(p0_items, 1):
            issue_text = f"【问题】{r.issue or r.evidence}\n【依据】{r.law_article}"
            set_cell(ws, row, 1, str(idx), FONT_BOLD, None, ALIGN_CENTER)
            merge_style(ws, row, 2, row, 2, r.item, FONT_BOLD, None, ALIGN_LEFT)
            merge_style(ws, row, 3, row, 3, r.category, FONT_NORMAL, None, ALIGN_CENTER)
            merge_style(ws, row, 4, row, 6, issue_text, FONT_NORMAL, None, ALIGN_LEFT_TOP)
            merge_style(ws, row, 7, row, 9, r.suggestion or '', FONT_RISK_HIGH, None, ALIGN_LEFT_TOP)
            ws.row_dimensions[row].height = 70
            row += 1

        row += 1

    # P1+P2 汇总
    p12_items = [r for r in problem_items if normalize_risk(r.risk_level) != 'high']
    if p12_items:
        merge_style(ws, row, 1, row, 9,
                    f"建议修改项汇总（P1 中风险 {mid_count} 项 + P2 低风险 {low_count} 项）",
                    FONT_SUBTITLE, FILL_MID)
        row += 1

        for (sc, ec), h in zip(sum_spans, sum_headers):
            merge_style(ws, row, sc, row, ec, h, FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
        ws.row_dimensions[row].height = 25
        row += 1

        for idx, r in enumerate(p12_items, 1):
            rl = normalize_risk(r.risk_level)
            issue_text = f"【问题】{r.issue or r.evidence}\n【依据】{r.law_article}"
            set_cell(ws, row, 1, str(idx), FONT_BOLD, None, ALIGN_CENTER)
            merge_style(ws, row, 2, row, 2, r.item, FONT_BOLD, None, ALIGN_LEFT)
            merge_style(ws, row, 3, row, 3, r.category, FONT_NORMAL, None, ALIGN_CENTER)
            merge_style(ws, row, 4, row, 6, issue_text, FONT_NORMAL, None, ALIGN_LEFT_TOP)
            merge_style(ws, row, 7, row, 9, r.suggestion or '', risk_font(rl), None, ALIGN_LEFT_TOP)
            ws.row_dimensions[row].height = 55
            row += 1

    ws.freeze_panes = ws.cell(row=detail_header_row + 1, column=1).coordinate

    widths = [8, 12, 9, 30, 12, 10, 25, 55, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="生成合同审批 Excel 报告 (v3)")
    sub = parser.add_subparsers(dest="command")

    p_file = sub.add_parser("from-file", help="从合同文本文件生成报告")
    p_file.add_argument("--file", required=True, help="合同文本文件路径")
    p_file.add_argument("--output", required=True, help="输出 xlsx 路径")
    p_file.add_argument("--title", default="", help="合同名称（可选）")

    p_id = sub.add_parser("from-db", help="从数据库合同生成报告")
    p_id.add_argument("--contract-id", type=int, required=True, help="合同 ID")
    p_id.add_argument("--output", required=True, help="输出 xlsx 路径")

    args = parser.parse_args()

    if args.command == "from-file":
        # 读取合同文本
        with open(args.file, "r", encoding="utf-8") as f:
            text = f.read()

        # 解析 + 审核
        parsed_contract = parse_contract(text)
        report = audit_contract(text, parsed_contract.clauses)
        report.parsed_clauses = parsed_contract.clauses

        # 提取基本信息
        info = {"title": args.title or os.path.basename(args.file)}

        # 尝试从解析结果提取甲乙双方和金额
        pa = getattr(parsed_contract, 'party_a', {})
        pb = getattr(parsed_contract, 'party_b', {})
        info['party_a'] = pa.get('name', '') if isinstance(pa, dict) else str(pa)[:50]
        info['party_b'] = pb.get('name', '') if isinstance(pb, dict) else str(pb)[:50]
        if not info['party_a']:
            for clause in parsed_contract.clauses:
                if getattr(clause, 'category', '') == '当事人信息':
                    info['party_a'] = getattr(clause, 'excerpt', '')[:50]
                    break

        generate_excel_report(report, args.output, info)
        print(f"✅ 报告已生成: {args.output}")
        print(f"   Sheet 1: 合同条款解析 ({len(parsed_contract.clauses)} 类)")
        print(f"   Sheet 2: 审批标准库 ({len(AUDIT_CRITERIA)} 项)")
        print(f"   Sheet 3: 审批结果与建议 ({len(report.results)} 项)")
        s = report.summary
        print(f"   通过: {s.get('pass',0)} / 警告: {s.get('warning',0)} / 不通过: {s.get('fail',0)} / N/A: {s.get('na',0)}")

    elif args.command == "from-db":
        print("⚠️  from-db 模式暂未实现，请使用 from-file")
        sys.exit(1)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
