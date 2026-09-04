#!/usr/bin/env python3
"""
合同审批报告导出 - Excel 格式 (v2)
组件: SCA-001 (L4)

在 v1 基础上新增:
- Sheet 4「签署要素审计」: 签名/印章检测结果 + 截图嵌入
- Sheet 1「合同基本信息」新增签署完整性摘要

用法:
    python3 export_excel_report_v2.py --file <合同文本文件> --output <输出xlsx> [--ocr-result <ocr_json>]
    python3 export_excel_report_v2.py --contract-id <ID> --output <输出xlsx>

    --ocr-result: OCR v5 检测结果 JSON（含签名/印章截图路径），可选
"""

import argparse
import json
import os
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 同目录导入
sys.path.insert(0, os.path.dirname(__file__))
from export_excel_report import (  # 复用 v1 的样式和基础构建
    COLOR_HIGH_RISK, COLOR_MID_RISK, COLOR_LOW_RISK,
    COLOR_HEADER, COLOR_SECTION, COLOR_TITLE_BG, COLOR_SUMMARY_BG,
    FONT_TITLE, FONT_SUBTITLE, FONT_HEADER, FONT_NORMAL, FONT_BOLD,
    FONT_RISK_HIGH, FONT_RISK_MID, FONT_RISK_LOW,
    FILL_TITLE, FILL_HEADER, FILL_SECTION, FILL_HIGH, FILL_MID, FILL_LOW, FILL_SUMMARY,
    THIN_BORDER, ALIGN_CENTER, ALIGN_LEFT, ALIGN_LEFT_TOP,
    set_cell, merge_style,
    generate_excel_report as generate_v1,
    _build_sheet_clauses, _build_sheet_standards, _build_sheet_results,
)

# 新增颜色
COLOR_OK = "2ECC71"
COLOR_WARN = "F39C12"
COLOR_MISS = "E74C3C"
COLOR_SEAL_BG = "FDEDEC"
COLOR_SIG_BG = "EAF2F8"

FILL_OK = PatternFill(start_color=COLOR_OK, end_color=COLOR_OK, fill_type="solid")
FILL_WARN = PatternFill(start_color=COLOR_WARN, end_color=COLOR_WARN, fill_type="solid")
FILL_MISS = PatternFill(start_color=COLOR_MISS, end_color=COLOR_MISS, fill_type="solid")
FILL_SEAL = PatternFill(start_color=COLOR_SEAL_BG, end_color=COLOR_SEAL_BG, fill_type="solid")
FILL_SIG = PatternFill(start_color=COLOR_SIG_BG, end_color=COLOR_SIG_BG, fill_type="solid")


# ============================================================
# 签署要素数据结构
# ============================================================

SIGNING_REQUIREMENTS = [
    {
        "party": "甲方",
        "element": "公章/合同章",
        "type": "seal",
        "required": True,
        "note": "甲方（需方）公章",
    },
    {
        "party": "乙方",
        "element": "公章/合同章",
        "type": "seal",
        "required": True,
        "note": "乙方（供方）公章",
    },
    {
        "party": "甲方",
        "element": "法定代表人签字",
        "type": "signature",
        "required": True,
        "note": "法定代表人或授权代表签字",
    },
    {
        "party": "乙方",
        "element": "法定代表人签字",
        "type": "signature",
        "required": True,
        "note": "法定代表人或授权代表签字",
    },
    {
        "party": "双方",
        "element": "签署日期",
        "type": "date",
        "required": True,
        "note": "合同落款日期",
    },
    {
        "party": "双方",
        "element": "页码完整性",
        "type": "pagination",
        "required": True,
        "note": "所有页完整，无缺页",
    },
]


def _find_matching_region(regions, party_keywords, region_type):
    """在检测结果中查找匹配甲方/乙方的指定类型区域"""
    best = None
    for reg in regions:
        if reg.get("type") != region_type:
            continue
        label = reg.get("label", "")
        # 标签匹配：标签里含"甲"或"乙"
        if any(kw in label for kw in party_keywords):
            # 优先选置信度高的
            if best is None or reg.get("confidence", 0) > best.get("confidence", 0):
                best = reg
    return best


def build_signing_summary(ocr_data: dict) -> dict:
    """从 OCR v5 结果构建签署要素摘要
    
    Args:
        ocr_data: OCR 结果 dict，含 seals / signatures 列表
                  (每项: {type, page, label, confidence, image_path})
    
    Returns:
        {
            "all_checked": bool,       # 是否全部满足
            "total": int,              # 应检项
            "passed": int,             # 通过项
            "items": [ {element, party, status, page, confidence, image_path}, ... ]
        }
    """
    seals = ocr_data.get("seals", [])
    signatures = ocr_data.get("signatures", [])
    all_regions = seals + signatures

    # 找甲方/乙方印章
    a_seal = _find_matching_region(all_regions, ["甲", "买", "需"], "seal")
    b_seal = _find_matching_region(all_regions, ["乙", "卖", "供"], "seal")

    # 找甲方/乙方签字
    a_sig = _find_matching_region(all_regions, ["甲", "买", "需"], "signature")
    b_sig = _find_matching_region(all_regions, ["乙", "卖", "供"], "signature")

    # 签署日期：从全文找日期模式
    full_text = ocr_data.get("text", "")
    import re
    date_found = None
    # 匹配 2026年X月X日 / 2026-XX-XX / 2026.XX.XX
    for m in re.finditer(r"(20\d{2}[年./-]\d{1,2}[月./-]\d{1,2}日?)", full_text):
        date_found = m.group(0)
        break
    date_status = "✅" if date_found else "❌"
    date_conf = 0.9 if date_found else 0.0
    date_page = "全文"

    # 页数
    total_pages = ocr_data.get("meta", {}).get("pages", 0)
    page_ok = total_pages >= 1
    page_status = "✅" if page_ok else "❌"

    items = [
        {"element": "公章/合同章", "party": "甲方", "status": "✅" if a_seal else "❌",
         "page": a_seal.get("page", "-") if a_seal else "-",
         "confidence": a_seal.get("confidence", 0) if a_seal else 0,
         "image_path": a_seal.get("image_path", "") if a_seal else ""},
        {"element": "公章/合同章", "party": "乙方", "status": "✅" if b_seal else "❌",
         "page": b_seal.get("page", "-") if b_seal else "-",
         "confidence": b_seal.get("confidence", 0) if b_seal else 0,
         "image_path": b_seal.get("image_path", "") if b_seal else ""},
        {"element": "法定代表人签字", "party": "甲方", "status": "✅" if a_sig else "❌",
         "page": a_sig.get("page", "-") if a_sig else "-",
         "confidence": a_sig.get("confidence", 0) if a_sig else 0,
         "image_path": a_sig.get("image_path", "") if a_sig else ""},
        {"element": "法定代表人签字", "party": "乙方", "status": "✅" if b_sig else "❌",
         "page": b_sig.get("page", "-") if b_sig else "-",
         "confidence": b_sig.get("confidence", 0) if b_sig else 0,
         "image_path": b_sig.get("image_path", "") if b_sig else ""},
        {"element": "签署日期", "party": "双方", "status": date_status,
         "page": date_page, "confidence": date_conf,
         "image_path": "", "detail": date_found or "未识别到日期"},
        {"element": "页码完整性", "party": "双方", "status": page_status,
         "page": f"共{total_pages}页" if total_pages else "-",
         "confidence": 1.0 if page_ok else 0.0,
         "image_path": "", "detail": f"{total_pages}页完整" if page_ok else "页数未知"},
    ]

    passed = sum(1 for it in items if it["status"] == "✅")
    return {
        "all_checked": passed == len(items),
        "total": len(items),
        "passed": passed,
        "items": items,
    }


# ============================================================
# Sheet 4: 签署要素审计
# ============================================================

def _build_sheet_signing(ws, summary: dict):
    """Sheet 4: 签署要素审计"""
    row = 1
    merge_style(ws, row, 1, row, 7, "签署要素审计", FONT_TITLE, FILL_TITLE)
    ws.row_dimensions[row].height = 30
    row += 1

    # 摘要
    status_color = COLOR_OK if summary["all_checked"] else COLOR_WARN
    status_fill = FILL_OK if summary["all_checked"] else FILL_WARN
    merge_style(ws, row, 1, row, 7,
                f"签署完整性: {summary['passed']}/{summary['total']} 项满足"
                f"  {'（全部满足，可以签署）' if summary['all_checked'] else '（存在缺失，需补充）'}",
                Font(name="微软雅黑", size=12, bold=True, color="FFFFFF"),
                status_fill, ALIGN_CENTER)
    ws.row_dimensions[row].height = 28
    row += 2

    # 表头
    headers = ["签署要素", "归属方", "状态", "位置(页码)", "置信度", "截图", "说明"]
    for i, h in enumerate(headers, 1):
        set_cell(ws, row, i, h, FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
    ws.row_dimensions[row].height = 25
    header_row = row
    row += 1

    # 数据行
    image_start_col = 6
    for item in summary["items"]:
        status = item["status"]
        fill = FILL_OK if status == "✅" else FILL_MISS
        set_cell(ws, row, 1, item["element"], FONT_NORMAL, None, ALIGN_LEFT)
        set_cell(ws, row, 2, item["party"], FONT_BOLD, None, ALIGN_CENTER)
        set_cell(ws, row, 3, status, FONT_BOLD, fill, ALIGN_CENTER)
        set_cell(ws, row, 4, item["page"], FONT_NORMAL, None, ALIGN_CENTER)
        conf = item.get("confidence", 0)
        conf_str = f"{conf:.0%}" if conf > 0 else "—"
        set_cell(ws, row, 5, conf_str, FONT_NORMAL, None, ALIGN_CENTER)

        # 截图单元格（如果有）
        img_path = item.get("image_path", "")
        if img_path and os.path.exists(img_path):
            try:
                img = XLImage(img_path)
                # 控制缩略图尺寸，不超过 60x60
                ratio = min(1.0, 60 / max(img.width, img.height))
                img.width = int(img.width * ratio)
                img.height = int(img.height * ratio)
                anchor = f"{get_column_letter(image_start_col)}{row}"
                ws.add_image(img, anchor)
                set_cell(ws, row, image_start_col, "见右侧截图", FONT_NORMAL, FILL_SEAL if item["element"].find("章") >= 0 else FILL_SIG, ALIGN_CENTER)
            except Exception as e:
                set_cell(ws, row, image_start_col, f"加载失败", FONT_NORMAL, FILL_WARN, ALIGN_CENTER)
        else:
            detail = item.get("detail", "—")
            set_cell(ws, row, image_start_col, detail, FONT_NORMAL, FILL_SUMMARY, ALIGN_LEFT)

        set_cell(ws, row, 7, item.get("note", "—"), FONT_NORMAL, None, ALIGN_LEFT)
        ws.row_dimensions[row].height = 50  # 给截图留空间
        row += 1

    row += 1

    # 底部说明
    merge_style(ws, row, 1, row, 7,
                "说明：公章/签字区域由 OCR v5 自动检测截图，置信度>60% 视为有效。"
                "签署日期从合同全文自动提取。",
                Font(name="微软雅黑", size=9, italic=True, color="7F8C8D"),
                None, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22
    row += 1
    merge_style(ws, row, 1, row, 7,
                "⚠️ 截图仅供人工复核，正式签署前请核对原件真伪。",
                Font(name="微软雅黑", size=9, bold=True, color="C0392B"),
                FILL_WARN, ALIGN_LEFT)

    # 列宽
    widths = [18, 10, 10, 14, 12, 22, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# 主入口
# ============================================================

def generate_excel_report_v2(report, output_path: str, contract_info: dict = None,
                             ocr_data: dict = None):
    """生成 v2 Excel 报告（含签署要素审计）

    Args:
        report: AuditReport 对象
        output_path: 输出 xlsx 路径
        contract_info: 合同基本信息字典
        ocr_data: OCR v5 检测结果 dict（含 seals/signatures/text/meta）
    """
    wb = Workbook()
    contract_info = contract_info or {}

    # ====== Sheet 1: 合同条款解析 ======
    ws1 = wb.active
    ws1.title = "1.合同条款解析"
    _build_sheet_clauses(ws1, report, contract_info)

    # ====== Sheet 2: 审批标准库 ======
    ws2 = wb.create_sheet("2.审批标准库")
    _build_sheet_standards(ws2)

    # ====== Sheet 3: 审批结果与建议 ======
    ws3 = wb.create_sheet("3.审批结果与建议")
    _build_sheet_results(ws3, report, contract_info)

    # ====== Sheet 4: 签署要素审计 ======
    if ocr_data:
        summary = build_signing_summary(ocr_data)
        ws4 = wb.create_sheet("4.签署要素审计")
        _build_sheet_signing(ws4, summary)
        # 把摘要塞回 contract_info，供 Sheet1 使用
        contract_info["signing_summary"] = summary

    wb.save(output_path)
    return output_path


def load_ocr_json(path: str) -> dict:
    """加载 OCR v5 结果 JSON"""
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def main():
    parser = argparse.ArgumentParser(description="合同审批报告导出 v2（含签署要素审计）")
    parser.add_argument("--file", help="合同文本文件")
    parser.add_argument("--contract-id", help="合同 ID（从数据库读取）")
    parser.add_argument("--output", default="contract_audit_report.xlsx", help="输出 xlsx 路径")
    parser.add_argument("--ocr-result", help="OCR v5 结果 JSON 路径")
    args = parser.parse_args()

    if not args.file and not args.contract_id:
        parser.error("必须指定 --file 或 --contract-id")

    # 复用 v1 的审核逻辑
    from contract_auditor import audit_contract
    from contract_parser import parse_contract

    if args.file:
        text = open(args.file, encoding="utf-8").read()
        parsed = parse_contract(text)
        report = audit_contract(text, parsed.clauses)
        report.parsed_clauses = parsed.clauses
        contract_info = {
            "title": getattr(parsed, "title", "") or os.path.basename(args.file),
            "contract_type": "技术合同",
            "party_a": getattr(parsed, "party_a", ""),
            "party_b": getattr(parsed, "party_b", ""),
            "amount": getattr(parsed, "amount", ""),
        }
    else:
        # 从数据库读取
        import sqlite3
        db = os.path.join(os.path.dirname(__file__), "..", "..", "..", "contracts.db")
        conn = sqlite3.connect(db)
        cur = conn.execute(
            "SELECT title, contract_type, party_a, party_b, amount FROM contracts WHERE id=?",
            (args.contract_id,)
        )
        row = cur.fetchone()
        conn.close()
        if not row:
            print(f"❌ 合同 {args.contract_id} 不存在")
            return 1
        text = row[0] or ""
        parsed = parse_contract(text)
        report = audit_contract(text, parsed.clauses)
        report.parsed_clauses = parsed.clauses
        contract_info = {
            "title": row[0] or f"合同{args.contract_id}",
            "contract_type": row[1] or "技术合同",
            "party_a": row[2] or "",
            "party_b": row[3] or "",
            "amount": row[4] or "",
        }

    ocr_data = None
    if args.ocr_result:
        ocr_data = load_ocr_json(args.ocr_result)
        print(f"   OCR 结果: {len(ocr_data.get('seals', []))} 印章, "
              f"{len(ocr_data.get('signatures', []))} 签名")

    generate_excel_report_v2(report, args.output, contract_info, ocr_data)
    print(f"✅ Excel 报告已生成: {args.output}")
    print(f"   Sheet 1: 合同条款解析")
    print(f"   Sheet 2: 审批标准库")
    print(f"   Sheet 3: 审批结果与建议")
    if ocr_data:
        print(f"   Sheet 4: 签署要素审计 (新增)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
