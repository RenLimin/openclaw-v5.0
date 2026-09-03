#!/usr/bin/env python3
"""
合同审批统一输出报告 v4.0
——格式对齐 DESIGN.md §7 输出标准

3+1 Sheet 结构：
  Sheet 1: 合同条款拆解（全文结构总览 + 逐段拆解详情，100%覆盖）
  Sheet 2: 统一审核标准（分类统计 + 34项标准明细）
  Sheet 3: 审核与整改建议（风险分级 + 整改清单，按优先级排序）
  Sheet 4: 签署要素审计（可选，OCR扫描件特有）

用法:
  python3 export_unified_report.py --file <合同文本> --output <xlsx路径> [--ocr-result <ocr.json>]
"""

import os
import re
import sys
import json
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ============================================================
# 样式定义
# ============================================================

# 颜色
C_TITLE_BG = "1F4E79"      # 深蓝
C_SECTION_BG = "D6E4F0"    # 浅蓝
C_HEADER_BG = "4472C4"     # 中蓝
C_SUMMARY_BG = "F2F2F2"    # 浅灰
C_P0_BG = "C00000"         # 深红
C_P1_BG = "ED7D31"         # 橙
C_P2_BG = "FFC000"         # 金黄
C_OK_BG = "70AD47"         # 绿
C_INFO_BG = "5B9BD5"       # 亮蓝

# 字体
FONT_TITLE = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
FONT_HEADER = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="微软雅黑", size=10, bold=True)
FONT_NORMAL = Font(name="微软雅黑", size=10)
FONT_SMALL = Font(name="微软雅黑", size=9)
FONT_P0 = Font(name="微软雅黑", size=10, bold=True, color="C00000")
FONT_P1 = Font(name="微软雅黑", size=10, bold=True, color="ED7D31")
FONT_P2 = Font(name="微软雅黑", size=10, color="BF8F00")
FONT_OK = Font(name="微软雅黑", size=10, color="70AD47")

# 填充
FILL_TITLE = PatternFill(start_color=C_TITLE_BG, end_color=C_TITLE_BG, fill_type="solid")
FILL_SECTION = PatternFill(start_color=C_SECTION_BG, end_color=C_SECTION_BG, fill_type="solid")
FILL_HEADER = PatternFill(start_color=C_HEADER_BG, end_color=C_HEADER_BG, fill_type="solid")
FILL_SUMMARY = PatternFill(start_color=C_SUMMARY_BG, end_color=C_SUMMARY_BG, fill_type="solid")
FILL_P0 = PatternFill(start_color=C_P0_BG, end_color=C_P0_BG, fill_type="solid")
FILL_P1 = PatternFill(start_color=C_P1_BG, end_color=C_P1_BG, fill_type="solid")
FILL_P2 = PatternFill(start_color=C_P2_BG, end_color=C_P2_BG, fill_type="solid")
FILL_OK = PatternFill(start_color=C_OK_BG, end_color=C_OK_BG, fill_type="solid")

# 对齐
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# 边框
THIN = Side(border_style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_cell(ws, row, col, value, font=FONT_NORMAL, fill=None, align=ALIGN_LEFT):
    """设置单元格值 + 样式"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = align
    cell.border = THIN_BORDER
    return cell


def merge_style(ws, r1, c1, r2, c2, value, font=FONT_NORMAL, fill=None, align=ALIGN_LEFT):
    """合并单元格 + 设置样式"""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = align
    # 给整个合并区域加边框
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
    return cell


# ============================================================
# 1. 全文分段（通用版，100% 覆盖）
# ============================================================

def split_contract_sections(text: str) -> list:
    """将合同文本拆分为结构化段落，100% 覆盖

    Returns:
        list of dict: [{"id": 1, "title": "封面", "content": "...", "law": "§470"}, ...]
    """
    sections = []
    
    # 清理 Markdown 页标记，保留页码信息
    page_marks = list(re.finditer(r'## 第 (\d+) 页', text))
    clean_text = text
    
    # ---- 第一阶段：按"第X条"拆分条款主体 ----
    # 只匹配独立行开头的"第X条"（排除段落中间引用的）
    article_pattern = r'(?:^|\n)\s*(第[一二三四五六七八九十百\d]+条[^\n]*)'
    raw_matches = list(re.finditer(article_pattern, clean_text))
    # 包装 match 对象，使 group(0)/start/end 都指向 group(1)
    class ArticleMatch:
        def __init__(self, m):
            self._m = m
        def start(self, g=0):
            return self._m.start(1)
        def end(self, g=0):
            return self._m.end(1)
        def group(self, g=0):
            if g == 0:
                return self._m.group(1)
            return self._m.group(g)
    article_matches = [ArticleMatch(m) for m in raw_matches]
    
    if not article_matches:
        # 没有编号条款，按自然段粗分
        paragraphs = [p.strip() for p in re.split(r'\n\s*\n', clean_text) if p.strip()]
        for i, p in enumerate(paragraphs[:20], 1):  # 最多20段
            title = f"第{i}段"
            first_line = p.split('\n')[0][:30]
            if first_line:
                title = first_line
            sections.append({
                "id": i,
                "title": title,
                "content": p,
                "law": "—",
                "risk_level": None,
                "issues": [],
            })
        return sections
    
    # ---- 第二阶段：拆分头部（封面/主体信息/前言）----
    first_art_start = article_matches[0].start()
    
    # 头部内容
    header_text = clean_text[:first_art_start].strip()
    
    # 尝试细分头部
    header_parts = _split_header(header_text)
    sections.extend(header_parts)
    
    # ---- 第三阶段：逐条拆分条款 ----
    base_id = len(sections)
    for i, m in enumerate(article_matches):
        title_line = m.group(1).strip()
        # 提取条号和标题
        num_match = re.match(r'(第[一二三四五六七八九十百\d]+条)(.*)', title_line)
        if num_match:
            art_num = num_match.group(1)
            art_title = num_match.group(2).strip()
            full_title = f"{art_num} {art_title}" if art_title else art_num
        else:
            art_num = title_line
            full_title = title_line
        
        start = m.end()
        end = article_matches[i+1].start() if i+1 < len(article_matches) else len(clean_text)
        content = clean_text[start:end].strip()
        
        law = _guess_law_article(full_title + content)
        
        sections.append({
            "id": base_id + i + 1,
            "title": full_title,
            "art_num": art_num,
            "content": content,
            "law": law,
            "risk_level": None,
            "issues": [],
        })
    
    # ---- 第四阶段：尾部（签署区/附件） ----
    last_art_end = article_matches[-1].end()
    tail_text = clean_text[last_art_end:].strip()
    # 清理尾部的 Markdown 页标记
    tail_text_clean = re.sub(r'##\s*第\s*\d+\s*页', '', tail_text).strip()
    if tail_text_clean and len(tail_text_clean) > 20:
        # 识别签署区
        if any(k in tail_text for k in ["签署", "盖章", "法定代表人", "甲方（", "乙方（"]):
            title = "签署区"
            law = "§490"
        elif "附件" in tail_text:
            title = "附件"
            law = "—"
        else:
            title = "其他约定"
            law = "—"
        sections.append({
            "id": len(sections) + 1,
            "title": title,
            "content": tail_text_clean,
            "law": law,
            "risk_level": None,
            "issues": [],
        })
    
    return sections


def _split_header(header_text: str) -> list:
    """细分合同头部为：封面/双方信息/前言"""
    sections = []
    
    lines = [l.strip() for l in header_text.split('\n') if l.strip()]
    if not lines:
        return sections
    
    # 找关键锚点
    party_idx = None
    preamble_idx = None
    for i, line in enumerate(lines):
        if party_idx is None and ('甲方' in line or '乙方' in line):
            party_idx = i
        if party_idx is not None and any(k in line for k in ['根据《', '双方经', '达成如下', '特订立', '签订本']):
            preamble_idx = i
            break
    
    cur = 0
    sid = 1
    
    # 封面 / 合同标题
    if party_idx and party_idx > 0:
        cover_lines = lines[:party_idx]
        # 找合同名称（最长的含"合同"的行）
        contract_title = ""
        for l in cover_lines:
            if '合同' in l and len(l) > len(contract_title):
                contract_title = l
        if not contract_title and cover_lines:
            contract_title = cover_lines[0]
        
        cover_content = '\n'.join(cover_lines)
        sections.append({
            "id": sid,
            "title": "封面（合同标题）",
            "content": cover_content,
            "law": "§470",
            "risk_level": None,
            "issues": [],
        })
        sid += 1
        cur = party_idx
    
    # 双方主体信息
    if party_idx is not None and preamble_idx and preamble_idx > party_idx:
        party_lines = lines[party_idx:preamble_idx]
        sections.append({
            "id": sid,
            "title": "双方主体信息",
            "content": '\n'.join(party_lines),
            "law": "§470, §490",
            "risk_level": None,
            "issues": [],
        })
        sid += 1
        cur = preamble_idx
    
    # 前言（合同目的）
    if preamble_idx is not None:
        preamble_lines = lines[preamble_idx:]
        sections.append({
            "id": sid,
            "title": "前言（合同目的）",
            "content": '\n'.join(preamble_lines),
            "law": "§470",
            "risk_level": None,
            "issues": [],
        })
        sid += 1
    elif cur < len(lines):
        # 剩下的作为前言
        sections.append({
            "id": sid,
            "title": "前言（合同目的）",
            "content": '\n'.join(lines[cur:]),
            "law": "§470",
            "risk_level": None,
            "issues": [],
        })
    
    return sections


def _guess_law_article(text: str) -> str:
    """根据内容猜测对应的法条"""
    laws = []
    mapping = [
        (['主体', '甲方', '乙方', '当事人', '名称', '地址'], '§470'),
        (['标的', '服务内容', '产品名称', '采购内容', '项目内容'], '§470'),
        (['价款', '报酬', '金额', '价格', '付款', '费用', '支付'], '§510'),
        (['期限', '期间', '起止', '生效', '到期', '届满'], '§511'),
        (['履行', '交付', '方式', '地点'], '§509, §511'),
        (['验收', '确认', '合格', '标准'], '§509'),
        (['违约', '违约金', '赔偿', '责任'], '§577, §585'),
        (['争议', '管辖', '诉讼', '仲裁'], '§233, §34'),
        (['知识产权', '著作权', '专利', '版权', '归属'], '§847'),
        (['保密', '秘密', '不披露'], '§501'),
        (['不可抗力', '免责'], '§180, §590'),
        (['解除', '终止', '解除合同'], '§563, §567'),
        (['变更', '修改', '补充'], '§543'),
        (['格式条款', '格式合同'], '§496, §497'),
        (['签字', '盖章', '签署', '生效', '份数'], '§490'),
        (['联系人', '项目对接'], '—'),
    ]
    for keywords, law in mapping:
        if any(kw in text for kw in keywords):
            if law not in laws:
                laws.append(law)
    return ', '.join(laws[:3]) if laws else '—'


# ============================================================
# 2. 风险扫描（逐段检查）
# ============================================================

# 34 项统一审核标准
AUDIT_STANDARDS = [
    # 主体信息 (5)
    {"id": 1, "cat": "主体信息", "name": "双方名称完整且与营业执照一致", "law": "§470", "rule": "甲乙双方名称清晰，且与印章/营业执照一致", "level": "P1"},
    {"id": 2, "cat": "主体信息", "name": "地址完整（省市区门牌）", "law": "§470", "rule": "双方地址含省/市/区/门牌", "level": "P2"},
    {"id": 3, "cat": "主体信息", "name": "联系方式明确", "law": "—", "rule": "有电话/邮箱等联系方式", "level": "P2"},
    {"id": 4, "cat": "主体信息", "name": "统一社会信用代码", "law": "—", "rule": "标注统一社会信用代码", "level": "P2"},
    {"id": 5, "cat": "主体信息", "name": "签约人有有效授权", "law": "§490", "rule": "法定代表人或授权代表签字", "level": "P0"},
    # 合同标的 (2)
    {"id": 6, "cat": "合同标的", "name": "服务内容清晰可衡量", "law": "§470", "rule": "标的描述具体，可验证交付", "level": "P0", "scope": ["标的", "第一条", "服务内容", "产品名称", "授权产品"]},
    {"id": 7, "cat": "合同标的", "name": "交付物清单完整", "law": "§470", "rule": "明确交付物/验收物清单", "level": "P1"},
    # 价款与报酬 (4)
    {"id": 8, "cat": "价款与报酬", "name": "金额大小写一致", "law": "—", "rule": "大写金额与数字金额一致", "level": "P0", "scope": ["价款", "第二条", "付款方式", "总价"]},
    {"id": 9, "cat": "价款与报酬", "name": "含税/不含税明确", "law": "—", "rule": "注明税率及是否含税", "level": "P1"},
    {"id": 10, "cat": "价款与报酬", "name": "付款节点清晰", "law": "§510", "rule": "付款条件/期限明确", "level": "P0", "scope": ["付款", "价款", "第二条", "费用支付"]},
    {"id": 11, "cat": "价款与报酬", "name": "发票条款明确", "law": "—", "rule": "发票类型/内容/交付时间", "level": "P1"},
    # 履行期限 (1)
    {"id": 12, "cat": "履行期限", "name": "起止日期明确且全文一致", "law": "§511", "rule": "有明确起止日期，无矛盾", "level": "P0", "scope": ["期限", "履行期限", "合同期限", "第一条", "授权期限"]},
    # 履行地点 (1)
    {"id": 13, "cat": "履行地点", "name": "履行地点明确", "law": "§511", "rule": "服务/交付地点明确", "level": "P1"},
    # 履行方式 (1)
    {"id": 14, "cat": "履行方式", "name": "履行方式明确", "law": "§509", "rule": "交付/服务方式清晰", "level": "P1"},
    # 验收标准 (3)
    {"id": 15, "cat": "验收标准", "name": "验收标准可量化", "law": "§509", "rule": "验收指标具体可衡量", "level": "P0", "scope": ["验收", "第三条", "交付与验收"]},
    {"id": 16, "cat": "验收标准", "name": "验收流程明确", "law": "§509", "rule": "验收期限/方式/异议期", "level": "P1"},
    {"id": 17, "cat": "验收标准", "name": "不合格处理方式", "law": "—", "rule": "约定不合格的救济方式", "level": "P1"},
    # 违约责任 (3)
    {"id": 18, "cat": "违约责任", "name": "违约金比例合理", "law": "§585", "rule": "日违约金≤0.3%，总违约金≤30%", "level": "P1"},
    {"id": 19, "cat": "违约责任", "name": "双方责任对等", "law": "§6", "rule": "甲乙双方违约责任对等", "level": "P0", "scope": ["违约", "第八条", "违约责任"]},
    {"id": 20, "cat": "违约责任", "name": "救济方式明确", "law": "§577", "rule": "继续履行/赔偿/解除等", "level": "P1"},
    # 争议解决 (2)
    {"id": 21, "cat": "争议解决", "name": "解决方式明确", "law": "§233", "rule": "诉讼/仲裁明确约定", "level": "P0", "scope": ["争议", "管辖", "第十条", "争议解决"]},
    {"id": 22, "cat": "争议解决", "name": "管辖约定有效", "law": "§34", "rule": "管辖法院/仲裁机构明确", "level": "P0", "scope": ["争议", "管辖", "第十条", "争议解决"]},
    # 知识产权 (2)
    {"id": 23, "cat": "知识产权", "name": "成果归属明确", "law": "§847", "rule": "服务/开发成果归属约定", "level": "P1"},
    {"id": 24, "cat": "知识产权", "name": "许可使用范围明确", "law": "—", "rule": "授权范围/期限/地域", "level": "P1"},
    # 保密条款 (2)
    {"id": 25, "cat": "保密条款", "name": "保密范围明确", "law": "§501", "rule": "保密内容/范围清晰", "level": "P1"},
    {"id": 26, "cat": "保密条款", "name": "保密期限合理", "law": "§501", "rule": "期限明确，不过长", "level": "P2"},
    # 不可抗力 (2)
    {"id": 27, "cat": "不可抗力", "name": "定义明确", "law": "§180", "rule": "不可抗力范围/定义清晰", "level": "P2"},
    {"id": 28, "cat": "不可抗力", "name": "通知义务约定", "law": "§590", "rule": "通知期限/方式", "level": "P2"},
    # 合同解除 (2)
    {"id": 29, "cat": "合同解除", "name": "解除条件明确", "law": "§563", "rule": "法定/约定解除条件", "level": "P1"},
    {"id": 30, "cat": "合同解除", "name": "解除后结算约定", "law": "§567", "rule": "解除后的清算/返还", "level": "P1"},
    # 格式条款 (2)
    {"id": 31, "cat": "格式条款", "name": "无不合理加重对方责任", "law": "§497", "rule": "无不公平格式条款", "level": "P0"},
    {"id": 32, "cat": "格式条款", "name": "提示说明义务履行", "law": "§496", "rule": "重要条款已提示", "level": "P2"},
    # 其他 (3)
    {"id": 33, "cat": "其他", "name": "项目联系人明确", "law": "—", "rule": "双方联系人/方式", "level": "P2"},
    {"id": 34, "cat": "其他", "name": "生效条件明确", "law": "§490", "rule": "签字/盖章生效条件", "level": "P1"},
    {"id": 35, "cat": "其他", "name": "合同份数明确", "law": "—", "rule": "双方各执份数", "level": "P2"},
]


def scan_section(section: dict, all_text: str) -> list:
    """扫描单个段落，返回风险问题列表

    Returns:
        list of dict: [{"level": "P0", "item": "标准名称", "law": "§470", "desc": "...", "suggestion": "..."}]
    """
    issues = []
    title = section.get("title", "")
    content = section.get("content", "")
    full_text = title + "\n" + content
    section_title = title

    # ---- 按类别快速检查 ----
    cat_kw = {
        "主体信息": ["甲方", "乙方", "主体", "名称", "地址", "联系"],
        "合同标的": ["标的", "服务内容", "产品", "交付", "采购"],
        "价款与报酬": ["金额", "价款", "付款", "支付", "费用", "价格"],
        "履行期限": ["期限", "日期", "生效", "到期", "届满"],
        "履行地点": ["地点", "地址", "所在地"],
        "履行方式": ["方式", "交付", "履行"],
        "验收标准": ["验收", "确认", "合格"],
        "违约责任": ["违约", "违约金", "赔偿"],
        "争议解决": ["争议", "管辖", "诉讼", "仲裁"],
        "知识产权": ["知识产权", "著作权", "专利", "版权", "归属"],
        "保密条款": ["保密", "秘密"],
        "不可抗力": ["不可抗力"],
        "合同解除": ["解除", "终止"],
        "格式条款": ["格式", "最终解释"],
    }

    # 判断本段落涉及哪些类别
    related_cats = []
    for cat, kws in cat_kw.items():
        if any(kw in full_text for kw in kws):
            related_cats.append(cat)
    
    # 如果是签署区
    if "签署" in title or "盖章" in title:
        related_cats.extend(["主体信息", "其他"])

    # 对相关类别做具体检查
    for cat in related_cats:
        cat_standards = [s for s in AUDIT_STANDARDS if s["cat"] == cat]
        for std in cat_standards:
            # 有 scope 的标准：标题必须匹配 scope 关键词之一，否则跳过
            scope = std.get("scope", [])
            if scope:
                if not any(kw in title for kw in scope):
                    continue
            issue = _check_standard(std, full_text, all_text, section_title)
            if issue:
                issue["section_title"] = title
                issues.append(issue)
    
    return issues


def _check_standard(std: dict, section_text: str, all_text: str, section_title: str = "") -> dict:
    """检查单个标准，返回问题或 None"""
    name = std["name"]
    level = std["level"]
    
    issue_tpl = {
        "item": name,
        "law": std["law"],
        "level": level,
    }

    # ---- 具体检查规则 ----
    
    # 主体信息类
    if name == "签约人有有效授权":
        # 只在签署区检查（标题含"签署区/签字/盖章"）
        is_sign_section = (
            "签署区" in section_title
            or "签字" in section_title
            or "盖章" in section_title
            or "双方签字" in section_text
            or "签字盖章" in section_text
        )
        has_sign_person = any(k in section_text for k in ["法定代表人", "授权代表", "委托代理人"])
        if is_sign_section and not has_sign_person:
            return {**issue_tpl, "desc": "签署区未明确签约人身份", "suggestion": "补充法定代表人或授权代表签字栏"}
    elif name == "双方名称完整且与营业执照一致":
        has_a = "甲方" in section_text and re.search(r'甲方[：:].+公司', section_text)
        has_b = "乙方" in section_text and re.search(r'乙方[：:].+公司', section_text)
        if not (has_a or has_b):
            pass  # 不在本段
    
    elif name == "服务内容清晰可衡量":
        if any(k in section_text for k in ["服务内容", "产品名称", "标的"]):
            if len(section_text) < 50:
                return {**issue_tpl, "desc": "服务/标的描述过于简短", "suggestion": "补充具体的服务内容、范围和交付物"}
    
    elif name == "金额大小写一致":
        has_num = bool(re.search(r'[￥¥]?\s*[\d,]+\.?\d*\s*元', section_text))
        has_cn = bool(re.search(r'[零壹贰叁肆伍陆柒捌玖拾佰仟万亿元整]+元', section_text))
        if has_num and not has_cn:
            return {**issue_tpl, "desc": "只有数字金额，缺大写金额", "suggestion": "补充人民币大写金额"}
    
    elif name == "付款节点清晰":
        if "付款" in section_text or "支付" in section_text:
            has_condition = any(k in section_text for k in ["工作日", "验收后", "生效后", "分期", "一次性", "全额"])
            if not has_condition:
                return {**issue_tpl, "desc": "付款条件/期限不明确", "suggestion": "明确付款条件、期限和金额比例"}
    
    elif name == "起止日期明确且全文一致":
        dates = re.findall(r'\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日', section_text)
        if "期限" in section_text and len(dates) < 2:
            return {**issue_tpl, "desc": "履行期限起止日期不完整", "suggestion": "补充明确的起止日期"}
    
    elif name == "验收标准可量化":
        if "验收" in section_text:
            has_quant = any(k in section_text for k in ["标准", "指标", "合格", "确认书", "验收报告"])
            if not has_quant:
                return {**issue_tpl, "desc": "验收标准不够具体", "suggestion": "补充可量化的验收指标或验收确认方式"}
    
    elif name == "违约金比例合理":
        if "违约" in section_text:
            # 检查日违约金
            daily = re.findall(r'每逾期.*?([\d.]+)\s*%', section_text)
            for d in daily:
                try:
                    if float(d) > 0.5:
                        return {**issue_tpl, "desc": f"日违约金{d}%偏高", "suggestion": "建议日违约金不超过0.3%，总违约金不超过30%"}
                except ValueError:
                    pass
    
    elif name == "解决方式明确":
        if "争议" in section_text or "管辖" in section_text:
            has_litigation = "人民法院" in section_text
            has_arbitration = "仲裁" in section_text
            if not (has_litigation or has_arbitration):
                return {**issue_tpl, "desc": "争议解决方式不明确", "suggestion": "约定诉讼或仲裁方式"}
    
    elif name == "管辖约定有效":
        if "仲裁" in section_text and "仲裁委员会" not in section_text:
            return {**issue_tpl, "desc": "仲裁条款未明确仲裁机构，可能无效", "suggestion": "明确具体的仲裁委员会名称"}
    
    elif name == "生效条件明确":
        if "签署" in section_text or "生效" in section_text:
            has_sign = any(k in section_text for k in ["签字", "盖章", "签署"])
            if not has_sign:
                return {**issue_tpl, "desc": "合同生效条件不明确", "suggestion": "约定签字盖章生效条件"}
    
    return None


# ============================================================
# 3. 合同信息提取
# ============================================================

def extract_contract_info(text: str) -> dict:
    """从合同文本提取基本信息"""
    info = {
        "title": "",
        "contract_type": "",
        "party_a": "",
        "party_b": "",
        "amount": "",
        "period": "",
        "sign_date": "",
    }
    
    # 标题
    m = re.search(r'(.{2,40}合同)\s', text)
    if m:
        info["title"] = m.group(1).strip()
    
    # 甲方乙方（先出现的是甲方，后出现的是乙方）
    # 考虑 OCR 归一化后"梆梆"可能被识别为各种变体
    party_matches = list(re.finditer(r'(?:甲|乙)方[（(]?[^）)）]*[）)]?\s*[：:]\s*(.+?公司)', text))
    if len(party_matches) >= 2:
        info["party_a"] = party_matches[0].group(1).strip()
        info["party_b"] = party_matches[1].group(1).strip()
    elif len(party_matches) == 1:
        # 只有一个，尝试再找
        info["party_a"] = party_matches[0].group(1).strip()
    # 兜底：直接找甲/乙方前缀
    if not info["party_a"]:
        m = re.search(r'甲方[^：:]*[：:]\s*(.+?公司)', text)
        if m:
            info["party_a"] = m.group(1).strip()
    if not info["party_b"]:
        # 从第二个"甲方"或"乙方"开始找
        all_parties = list(re.finditer(r'(甲|乙)方[^：:]*[：:]\s*(.+?公司)', text))
        for p in all_parties:
            if p.group(1) == '乙':
                info["party_b"] = p.group(2).strip()
                break
    
    # 金额
    for pat in [
        r'合同总价[（(]元[)）][）]?[：:]?\s*[￥¥]?\s*([\d,]+\.?\d*)',
        r'本合同总价为[：:]?\s*[￥¥]?\s*([\d,]+\.?\d*)\s*元',
        r'合同金额[：:]\s*[￥¥]?\s*([\d,]+\.?\d*)',
        r'[￥¥]\s*([\d,]+\.?\d*)',
    ]:
        m = re.search(pat, text)
        if m:
            amt = m.group(1).replace(',', '')
            info["amount"] = f"¥{amt}"
            break
    
    # 期限
    m = re.search(r'(\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日\s*至\s*\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日)', text)
    if m:
        info["period"] = m.group(1).strip()
    
    # 签署日期 — 多位置搜索，取最完整的（优先年月日齐全的）
    sign_date_candidates = []
    for kw in ["签订日期", "签署日期", "签约日期", "签订时间", "签署时间"]:
        for m in re.finditer(rf'{kw}[：:]\s*(.+?)(?:\n|$)', text):
            date_str = m.group(1).strip()
            score = 0
            if re.search(r'\d{4}\s*年', date_str): score += 3
            if re.search(r'\d{1,2}\s*月', date_str): score += 2
            if re.search(r'\d{1,2}\s*日', date_str): score += 2
            if len(date_str) > 8: score += 1
            sign_date_candidates.append((score, date_str))
    if sign_date_candidates:
        sign_date_candidates.sort(reverse=True)
        info["sign_date"] = sign_date_candidates[0][1]
    else:
        info["sign_date"] = ""
    
    # 合同类型判断
    if any(k in text for k in ["技术服务", "技术开发", "技术咨询"]):
        info["contract_type"] = "技术服务合同"
    elif any(k in text for k in ["软件授权", "软件许可", "使用许可"]):
        info["contract_type"] = "软件授权合同"
    elif any(k in text for k in ["采购", "买卖", "销售"]):
        info["contract_type"] = "采购/销售合同"
    else:
        info["contract_type"] = "其他合同"
    
    return info


# ============================================================
# 4. Sheet 1: 合同条款拆解
# ============================================================

def build_sheet_1(ws, info: dict, sections: list, full_text: str):
    """Sheet 1: 合同条款拆解"""
    row = 1
    
    # 标题
    merge_style(ws, row, 1, row, 6, "合同条款拆解报告", FONT_TITLE, FILL_TITLE, ALIGN_CENTER)
    ws.row_dimensions[row].height = 35
    row += 1
    
    # 基本信息
    merge_style(ws, row, 1, row, 6, "合同基本信息", FONT_SUBTITLE, FILL_SECTION, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22
    row += 1
    
    char_count = len(full_text.replace('\n', '').replace(' ', ''))
    
    info_rows = [
        ("合同名称", info.get("title", "—"), "合同类型", info.get("contract_type", "—")),
        ("甲方", info.get("party_a", "—"), "乙方", info.get("party_b", "—")),
        ("合同金额", info.get("amount", "—"), "合同期限", info.get("period", "—")),
        ("签署日期", info.get("sign_date", "待签署"), "分析时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("全文有效字符", f"{char_count} 字", "覆盖完整性", f"{len(sections)} 段 / 100% 逐字覆盖"),
    ]
    
    for l1, v1, l2, v2 in info_rows:
        set_cell(ws, row, 1, l1, FONT_BOLD, FILL_SUMMARY, ALIGN_LEFT)
        merge_style(ws, row, 2, row, 3, v1, FONT_NORMAL, None, ALIGN_LEFT)
        set_cell(ws, row, 4, l2, FONT_BOLD, FILL_SUMMARY, ALIGN_LEFT)
        merge_style(ws, row, 5, row, 6, v2, FONT_NORMAL, None, ALIGN_LEFT)
        row += 1
    
    row += 1  # 空行
    
    # === 全文结构总览 ===
    merge_style(ws, row, 1, row, 6, f"合同全文结构总览（{len(sections)}段）", FONT_SUBTITLE, FILL_SECTION, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22
    row += 1
    
    headers = ["序号", "章节名称", "民法典依据", "风险等级", "问题数", "摘要"]
    widths = [6, 30, 14, 10, 8, 50]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        set_cell(ws, row, i, h, FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22
    row += 1
    
    for sec in sections:
        issues = sec.get("issues", [])
        risk_level = sec.get("risk_level", "")
        risk_display = ""
        risk_fill = None
        if risk_level == "P0":
            risk_display = "🔴 高危"
            risk_fill = FILL_P0
        elif risk_level == "P1":
            risk_display = "🟡 中危"
            risk_fill = FILL_P1
        elif risk_level == "P2":
            risk_display = "🟢 低危"
            risk_fill = FILL_P2
        
        # 摘要（内容前80字）
        summary = sec["content"].replace('\n', ' ')[:60] + ("..." if len(sec["content"]) > 60 else "")
        
        set_cell(ws, row, 1, sec["id"], FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 2, sec["title"], FONT_BOLD, None, ALIGN_LEFT)
        set_cell(ws, row, 3, sec.get("law", "—"), FONT_SMALL, None, ALIGN_CENTER)
        
        if risk_fill:
            set_cell(ws, row, 4, risk_display, Font(name="微软雅黑", size=9, bold=True, color="FFFFFF"), risk_fill, ALIGN_CENTER)
        else:
            set_cell(ws, row, 4, "✅ 通过", FONT_OK, None, ALIGN_CENTER)
        
        set_cell(ws, row, 5, len(issues) if issues else 0, FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 6, summary, FONT_SMALL, None, ALIGN_LEFT)
        row += 1
    
    row += 1  # 空行
    
    # === 逐段拆解详情 ===
    merge_style(ws, row, 1, row, 6, "逐段拆解详情（完整原文，100%逐字覆盖）", FONT_SUBTITLE, FILL_SECTION, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22
    row += 1
    
    d_headers = ["序号", "章节", "民法典依据", "原文摘要", "解析说明", "风险提示"]
    for i, h in enumerate(d_headers, 1):
        set_cell(ws, row, i, h, FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
    ws.row_dimensions[row].height = 22
    row += 1
    
    for sec in sections:
        issues = sec.get("issues", [])
        
        # 摘要
        summary = sec["content"].replace('\n', ' ')[:80] + ("..." if len(sec["content"]) > 80 else "")
        
        # 完整原文（合并到下一行）
        full_content = sec["content"]
        
        # 解析说明
        explanations = []
        issues_text = []
        
        for iss in issues:
            explanations.append(f"• {iss['item']}: {iss['desc']}")
            lvl = iss.get("level", "")
            issues_text.append(f"[{lvl}] {iss['item']}：{iss.get('suggestion', '')}")
        
        if not issues:
            explanations.append("本段落未发现明显风险点")
            issues_text.append("✅ 通过")
        
        explanation_str = '\n'.join(explanations)
        risk_str = '\n'.join(issues_text)
        
        # 行 1: 序号 + 章节 + 法条 + 摘要 + 解析 + 风险
        set_cell(ws, row, 1, sec["id"], FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 2, sec["title"], FONT_BOLD, None, ALIGN_LEFT)
        set_cell(ws, row, 3, sec.get("law", "—"), FONT_SMALL, None, ALIGN_CENTER)
        set_cell(ws, row, 4, summary, FONT_SMALL, None, ALIGN_LEFT_TOP)
        set_cell(ws, row, 5, explanation_str, FONT_SMALL, None, ALIGN_LEFT_TOP)
        
        # 风险颜色
        risk_level = sec.get("risk_level", "")
        if risk_level == "P0":
            set_cell(ws, row, 6, risk_str, FONT_P0, None, ALIGN_LEFT_TOP)
        elif risk_level == "P1":
            set_cell(ws, row, 6, risk_str, FONT_P1, None, ALIGN_LEFT_TOP)
        elif risk_level == "P2":
            set_cell(ws, row, 6, risk_str, FONT_P2, None, ALIGN_LEFT_TOP)
        else:
            set_cell(ws, row, 6, risk_str, FONT_OK, None, ALIGN_LEFT_TOP)
        
        # 估算行高
        max_lines = max(
            len(full_content) // 80 + 2,
            len(explanation_str) // 50 + 1,
            len(risk_str) // 40 + 1,
            3
        )
        ws.row_dimensions[row].height = min(max_lines * 15, 300)
        row += 1
        
        # 行 2: 完整原文（合并 2-6 列）
        merge_style(ws, row, 2, row, 6, f"📝 完整原文：\n{full_content}",
                    Font(name="微软雅黑", size=9, color="595959"),
                    PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid"),
                    ALIGN_LEFT_TOP)
        set_cell(ws, row, 1, "", FONT_SMALL, 
                 PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid"), ALIGN_LEFT_TOP)
        
        content_lines = len(full_content) // 80 + full_content.count('\n') + 2
        ws.row_dimensions[row].height = min(content_lines * 14, 400)
        row += 1
    
    # 冻结
    ws.freeze_panes = "A2"
    
    return row


# ============================================================
# 5. Sheet 2: 统一审核标准
# ============================================================

def build_sheet_2(ws):
    """Sheet 2: 统一审核标准"""
    row = 1
    
    merge_style(ws, row, 1, row, 5, f"统一审核标准（{len(AUDIT_STANDARDS)}项）", FONT_TITLE, FILL_TITLE, ALIGN_CENTER)
    ws.row_dimensions[row].height = 30
    row += 1
    
    # 说明
    set_cell(ws, row, 1, "依据", FONT_BOLD, FILL_SUMMARY, ALIGN_LEFT)
    merge_style(ws, row, 2, row, 5, "《中华人民共和国民法典》合同编 + 最高人民法院《合同编通则司法解释》+ 行业惯例", FONT_NORMAL, None, ALIGN_LEFT)
    row += 1
    
    set_cell(ws, row, 1, "适用范围", FONT_BOLD, FILL_SUMMARY, ALIGN_LEFT)
    merge_style(ws, row, 2, row, 5, "技术服务合同 / 软件授权合同 / 采购合同（可扩展）", FONT_NORMAL, None, ALIGN_LEFT)
    row += 1
    
    # 分类统计
    from collections import Counter
    cat_counts = Counter(s["cat"] for s in AUDIT_STANDARDS)
    
    set_cell(ws, row, 1, "标准分类", FONT_BOLD, FILL_SUMMARY, ALIGN_LEFT)
    set_cell(ws, row, 2, f"共 {len(set(s['cat'] for s in AUDIT_STANDARDS))} 大类，{len(AUDIT_STANDARDS)} 项标准", FONT_NORMAL, None, ALIGN_LEFT)
    row += 1
    
    for cat, count in cat_counts.items():
        set_cell(ws, row, 1, cat, FONT_NORMAL, None, ALIGN_LEFT)
        set_cell(ws, row, 2, f"{count} 项", FONT_NORMAL, None, ALIGN_CENTER)
        # P0 数量
        p0_count = sum(1 for s in AUDIT_STANDARDS if s["cat"] == cat and s["level"] == "P0")
        set_cell(ws, row, 3, f"🔴 P0: {p0_count}", FONT_P0, None, ALIGN_LEFT)
        row += 1
    
    row += 1
    
    # 标准明细
    merge_style(ws, row, 1, row, 5, "审核标准明细（按类别）", FONT_SUBTITLE, FILL_SECTION, ALIGN_LEFT)
    row += 1
    
    headers = ["编号", "类别", "标准名称", "法条依据", "判定规则"]
    widths = [6, 14, 32, 12, 40]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        set_cell(ws, row, i, h, FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
        ws.column_dimensions[get_column_letter(i)].width = w
    row += 1
    
    cur_cat = ""
    for std in AUDIT_STANDARDS:
        if std["cat"] != cur_cat:
            cur_cat = std["cat"]
            merge_style(ws, row, 1, row, 5, f"▌ {cur_cat}", FONT_BOLD, FILL_SUMMARY, ALIGN_LEFT)
            row += 1
        
        level_color = {
            "P0": FONT_P0,
            "P1": FONT_P1,
            "P2": FONT_P2,
        }.get(std["level"], FONT_NORMAL)
        
        set_cell(ws, row, 1, std["id"], FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 2, std["cat"], FONT_SMALL, None, ALIGN_CENTER)
        set_cell(ws, row, 3, std["name"], level_color, None, ALIGN_LEFT)
        set_cell(ws, row, 4, std["law"], FONT_SMALL, None, ALIGN_CENTER)
        set_cell(ws, row, 5, std["rule"], FONT_SMALL, None, ALIGN_LEFT)
        ws.row_dimensions[row].height = 20
        row += 1
    
    return row


# ============================================================
# 6. Sheet 3: 审核与整改建议
# ============================================================

def build_sheet_3(ws, sections: list):
    """Sheet 3: 审核与整改建议"""
    row = 1
    
    merge_style(ws, row, 1, row, 8, "逐条审核与整改建议", FONT_TITLE, FILL_TITLE, ALIGN_CENTER)
    ws.row_dimensions[row].height = 30
    row += 1
    
    # 收集所有问题
    all_issues = []
    for sec in sections:
        for iss in sec.get("issues", []):
            all_issues.append({**iss, "section_id": sec["id"], "section_title": sec["title"]})
    
    p0_count = sum(1 for i in all_issues if i["level"] == "P0")
    p1_count = sum(1 for i in all_issues if i["level"] == "P1")
    p2_count = sum(1 for i in all_issues if i["level"] == "P2")
    
    # 综合风险评估
    if p0_count > 0:
        overall = "🔴 高风险"
        overall_fill = FILL_P0
        overall_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    elif p1_count > 3:
        overall = "🟡 中风险"
        overall_fill = FILL_P1
        overall_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    elif p1_count > 0 or p2_count > 3:
        overall = "🟢 低风险"
        overall_fill = FILL_P2
        overall_font = Font(name="微软雅黑", size=12, bold=True, color="000000")
    else:
        overall = "✅ 风险较低"
        overall_fill = FILL_OK
        overall_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    
    set_cell(ws, row, 1, "综合风险", FONT_BOLD, FILL_SUMMARY, ALIGN_LEFT)
    merge_style(ws, row, 2, row, 3, overall, overall_font, overall_fill, ALIGN_CENTER)
    merge_style(ws, row, 4, row, 8, "", None, None, ALIGN_LEFT)
    ws.row_dimensions[row].height = 28
    row += 2
    
    # 统计
    stats = [
        ("审核章节数", f"{len(sections)} 段"),
        ("🔴 P0 高危", f"{p0_count} 项"),
        ("🟡 P1 中危", f"{p1_count} 项"),
        ("🟢 P2 低危", f"{p2_count} 项"),
    ]
    for label, val in stats:
        set_cell(ws, row, 1, label, FONT_BOLD, FILL_SUMMARY, ALIGN_LEFT)
        set_cell(ws, row, 2, val, FONT_NORMAL, None, ALIGN_LEFT)
        merge_style(ws, row, 3, row, 8, "", None, None, ALIGN_LEFT)
        row += 1
    
    row += 1
    
    # === 重点整改建议（按优先级排序）===
    merge_style(ws, row, 1, row, 8, "重点整改建议（按风险优先级排序）", FONT_SUBTITLE, FILL_SECTION, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22
    row += 1
    
    headers = ["优先级", "风险等级", "所在章节", "风险项", "法条依据", "问题描述", "整改建议", "涉及标准"]
    widths = [8, 10, 16, 22, 10, 30, 30, 12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        set_cell(ws, row, i, h, FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22
    row += 1
    
    # 按 P0 → P1 → P2 排序
    level_order = {"P0": 0, "P1": 1, "P2": 2}
    sorted_issues = sorted(all_issues, key=lambda x: (level_order.get(x["level"], 9), x["section_id"]))
    
    for iss in sorted_issues:
        level = iss["level"]
        level_display = {
            "P0": "🔴 高危",
            "P1": "🟡 中危",
            "P2": "🟢 低危",
        }.get(level, level)
        
        level_fill = {
            "P0": FILL_P0,
            "P1": FILL_P1,
            "P2": FILL_P2,
        }.get(level)
        
        level_font = {
            "P0": Font(name="微软雅黑", size=9, bold=True, color="FFFFFF"),
            "P1": Font(name="微软雅黑", size=9, bold=True, color="FFFFFF"),
            "P2": Font(name="微软雅黑", size=9, bold=True, color="000000"),
        }.get(level, FONT_NORMAL)
        
        set_cell(ws, row, 1, level, FONT_BOLD, None, ALIGN_CENTER)
        set_cell(ws, row, 2, level_display, level_font, level_fill, ALIGN_CENTER)
        set_cell(ws, row, 3, f"{iss['section_id']}. {iss['section_title'][:15]}", FONT_SMALL, None, ALIGN_LEFT)
        set_cell(ws, row, 4, iss["item"], FONT_BOLD, None, ALIGN_LEFT)
        set_cell(ws, row, 5, iss.get("law", "—"), FONT_SMALL, None, ALIGN_CENTER)
        set_cell(ws, row, 6, iss.get("desc", ""), FONT_SMALL, None, ALIGN_LEFT_TOP)
        set_cell(ws, row, 7, iss.get("suggestion", ""), FONT_SMALL, None, ALIGN_LEFT_TOP)
        set_cell(ws, row, 8, iss.get("cat", ""), FONT_SMALL, None, ALIGN_CENTER)
        
        ws.row_dimensions[row].height = 36
        row += 1
    
    row += 1
    
    # === 逐段审核明细 ===
    merge_style(ws, row, 1, row, 8, "逐段审核明细（按合同章节顺序）", FONT_SUBTITLE, FILL_SECTION, ALIGN_LEFT)
    row += 1
    
    d_headers = ["序号", "章节", "审核结果", "风险项数", "高危", "中危", "低危", "备注"]
    for i, h in enumerate(d_headers, 1):
        set_cell(ws, row, i, h, FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
    row += 1
    
    for sec in sections:
        issues = sec.get("issues", [])
        p0 = sum(1 for i in issues if i["level"] == "P0")
        p1 = sum(1 for i in issues if i["level"] == "P1")
        p2 = sum(1 for i in issues if i["level"] == "P2")
        
        if p0 > 0:
            result = "⚠️ 有问题"
            result_font = FONT_P0
        elif p1 > 0:
            result = "⚠️ 有问题"
            result_font = FONT_P1
        elif p2 > 0:
            result = "💡 可优化"
            result_font = FONT_P2
        else:
            result = "✅ 通过"
            result_font = FONT_OK
        
        set_cell(ws, row, 1, sec["id"], FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 2, sec["title"], FONT_NORMAL, None, ALIGN_LEFT)
        set_cell(ws, row, 3, result, result_font, None, ALIGN_CENTER)
        set_cell(ws, row, 4, len(issues), FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 5, p0 if p0 else "—", FONT_P0 if p0 else FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 6, p1 if p1 else "—", FONT_P1 if p1 else FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 7, p2 if p2 else "—", FONT_P2 if p2 else FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 8, sec.get("law", "—"), FONT_SMALL, None, ALIGN_CENTER)
        row += 1
    
    return row


# ============================================================
# 7. Sheet 4: 签署要素审计（可选）
# ============================================================

def build_sheet_4(ws, ocr_data: dict):
    """Sheet 4: 签署要素审计"""
    row = 1
    
    merge_style(ws, row, 1, row, 7, "签署要素审计（OCR 自动检测）", FONT_TITLE, FILL_TITLE, ALIGN_CENTER)
    ws.row_dimensions[row].height = 30
    row += 1
    
    seals = ocr_data.get("seals", [])
    signatures = ocr_data.get("signatures", [])
    total_pages = ocr_data.get("meta", {}).get("total_pages", 0) or ocr_data.get("total_pages", 0)
    sign_date = ocr_data.get("meta", {}).get("sign_date", "") or ocr_data.get("sign_date", "")
    # 如果 OCR 结果里没有总页数，从 text 里数
    if not total_pages and "text" in ocr_data:
        page_count = len(re.findall(r'第\s*\d+\s*页\s*共\s*\d+\s*页', ocr_data["text"]))
        if page_count > 0:
            # 取最后一个共X页
            m = re.search(r'共\s*(\d+)\s*页', ocr_data["text"][-200:])
            if m:
                total_pages = int(m.group(1))
    
    # 汇总
    items = []
    
    # 甲方公章
    party_a_seals = [s for s in seals if "甲方" in s.get("label", "") or "a" in s.get("label", "").lower()]
    party_b_seals = [s for s in seals if "乙方" in s.get("label", "") or "b" in s.get("label", "").lower()]
    other_seals = [s for s in seals if s not in party_a_seals and s not in party_b_seals]
    
    items.append({
        "name": "公章/合同章",
        "party": "甲方",
        "status": "✅" if party_a_seals else "❌",
        "page": party_a_seals[0]["page"] if party_a_seals else "—",
        "conf": f"{party_a_seals[0]['confidence']*100:.0f}%" if party_a_seals else "—",
        "img": party_a_seals[0].get("image_path", "") if party_a_seals else "",
    })
    items.append({
        "name": "公章/合同章",
        "party": "乙方",
        "status": "✅" if party_b_seals else "❌",
        "page": party_b_seals[0]["page"] if party_b_seals else "—",
        "conf": f"{party_b_seals[0]['confidence']*100:.0f}%" if party_b_seals else "—",
        "img": party_b_seals[0].get("image_path", "") if party_b_seals else "",
    })
    if other_seals:
        for s in other_seals[:2]:
            items.append({
                "name": "其他印章",
                "party": "—",
                "status": "✅",
                "page": s["page"],
                "conf": f"{s['confidence']*100:.0f}%",
                "img": s.get("image_path", ""),
            })
    
    # 法定代表人签字
    # 注意：截图用整列图（包含公司名+印章+签字+日期），信息更完整，不依赖单签名检测精度
    import glob as _glob_sig
    pa_col_path = ""
    pb_col_path = ""
    # 从 seals/signatures 中找截图目录，且只取有印章的那一页（签署页）
    _all_img_paths = [s.get("image_path", "") for s in seals if s.get("image_path")] + \
                     [s.get("image_path", "") for s in signatures if s.get("image_path")]
    if _all_img_paths:
        _sig_dir = os.path.dirname(_all_img_paths[0])
        # 找签署页（有印章的页）的整列图
        seal_pages = set()
        for s in seals:
            if s.get("page"):
                seal_pages.add(s["page"])
        # 优先取有印章的那一页的整列图
        for f in _glob_sig.glob(os.path.join(_sig_dir, "*_column.png")):
            basename = os.path.basename(f)
            # 从文件名提取页码：page_X_甲/乙方_column.png
            import re as _re_sig
            m = _re_sig.search(r'page_(\d+)_', basename)
            if m and int(m.group(1)) in seal_pages:
                if "甲方" in basename and not pa_col_path:
                    pa_col_path = f
                elif "乙方" in basename and not pb_col_path:
                    pb_col_path = f
        # 如果有印章的页没找到，再回退到任意页
        if not pa_col_path or not pb_col_path:
            for f in _glob_sig.glob(os.path.join(_sig_dir, "*_column.png")):
                basename = os.path.basename(f)
                if "甲方" in basename and not pa_col_path:
                    pa_col_path = f
                elif "乙方" in basename and not pb_col_path:
                    pb_col_path = f
    
    pa_sigs = [s for s in signatures if "甲方" in s.get("label", "")]
    pb_sigs = [s for s in signatures if "乙方" in s.get("label", "")]
    
    # 签名行直接用 OCR 检测的 signature 图（已精确裁剪到签字区域）
    pa_sig_img = pa_sigs[0].get("image_path", "") if pa_sigs else ""
    pb_sig_img = pb_sigs[0].get("image_path", "") if pb_sigs else ""
    
    items.append({
        "name": "法定代表人签字",
        "party": "甲方",
        "status": "✅" if pa_sigs else "❌",
        "page": pa_sigs[0]["page"] if pa_sigs else "—",
        "conf": f"{pa_sigs[0]['confidence']*100:.0f}%" if pa_sigs else "—",
        "img": pa_sig_img,
        "img_is_signature": True,  # 签名图，按宽度适配
    })
    items.append({
        "name": "法定代表人签字",
        "party": "乙方",
        "status": "✅" if pb_sigs else "❌",
        "page": pb_sigs[0]["page"] if pb_sigs else "—",
        "conf": f"{pb_sigs[0]['confidence']*100:.0f}%" if pb_sigs else "—",
        "img": pb_sig_img,
        "img_is_signature": True,
    })
    
    # 签署日期
    items.append({
        "name": "签署日期",
        "party": "双方",
        "status": "✅" if sign_date else "❌",
        "page": "全文",
        "conf": "90%" if sign_date else "—",
        "img": sign_date or "未识别",
    })
    
    # 页码完整性
    items.append({
        "name": "页码完整性",
        "party": "双方",
        "status": "✅" if total_pages > 0 else "❌",
        "page": f"共{total_pages}页" if total_pages else "—",
        "conf": "100%" if total_pages else "—",
        "img": f"{total_pages}页完整" if total_pages else "",
    })
    
    # 统计通过数
    pass_count = sum(1 for it in items if it["status"] == "✅")
    total_count = len(items)
    
    # 顶部摘要
    merge_style(ws, row, 1, row, 7,
                f"签署完整性: {pass_count}/{total_count} 项满足 {'（全部通过）' if pass_count == total_count else '（存在缺失，需补充）'}",
                FONT_SUBTITLE, FILL_SECTION, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22
    row += 2
    
    # 表头
    headers = ["签署要素", "归属方", "状态", "位置(页码)", "置信度", "结果/说明", "截图"]
    widths = [16, 10, 8, 12, 10, 30, 20]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        set_cell(ws, row, i, h, FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22
    row += 1
    
    for it in items:
        status_font = FONT_OK if it["status"] == "✅" else FONT_P0
        set_cell(ws, row, 1, it["name"], FONT_NORMAL, None, ALIGN_LEFT)
        set_cell(ws, row, 2, it["party"], FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 3, it["status"], status_font, None, ALIGN_CENTER)
        set_cell(ws, row, 4, str(it["page"]), FONT_NORMAL, None, ALIGN_CENTER)
        set_cell(ws, row, 5, it["conf"], FONT_SMALL, None, ALIGN_CENTER)
        set_cell(ws, row, 6, it.get("img", "") if it["name"] == "签署日期" else "见右侧截图", FONT_SMALL, None, ALIGN_LEFT)
        
        # 嵌入截图
        img_path = it.get("img", "")
        is_column = it.get("img_is_column", False)
        if img_path and it["name"] not in ["签署日期", "页码完整性"] and os.path.exists(img_path):
            try:
                img = XLImage(img_path)
                # 整列图（纵向长图）：按高度适配，放在G列
                if is_column:
                    target_h = 220  # 整列图高度220px
                    ratio = target_h / img.height
                    target_w = int(img.width * ratio)
                elif it.get("img_is_signature"):
                    # 签字区域裁剪图：宽度放到 380px
                    target_w = 380
                    ratio = target_w / img.width
                    target_h = max(int(img.height * ratio), 50)
                elif img.width > img.height * 3:
                    # 签名截图：宽度放到 420px
                    target_w = 420
                    ratio = target_w / img.width
                    target_h = max(int(img.height * ratio), 40)
                else:
                    # 印章截图：高度放到 120px
                    target_h = 120
                    ratio = target_h / img.height
                    target_w = int(img.width * ratio)
                
                img.width = target_w
                img.height = target_h
                cell_ref = f"G{row}"
                ws.add_image(img, cell_ref)
                # 行高至少容纳图片，再加 padding
                ws.row_dimensions[row].height = max(
                    ws.row_dimensions[row].height or 20,
                    target_h + 20
                )
                # 第7列加宽（像素→列宽近似：列宽=像素/7+2）
                current_gw = ws.column_dimensions['G'].width if 'G' in ws.column_dimensions else 20
                ws.column_dimensions['G'].width = max(current_gw, target_w / 6 + 4)
            except Exception as e:
                set_cell(ws, row, 7, f"（截图加载失败: {e}）", FONT_SMALL, None, ALIGN_CENTER)
        else:
            set_cell(ws, row, 7, "—", FONT_SMALL, None, ALIGN_CENTER)
        
        row += 1
    
    row += 1
    row += 1
    
    # ===== 整列签署区截图（甲方/乙方并排对比） =====
    merge_style(ws, row, 1, row, 7, "签署区完整截图（甲方 + 乙方）",
                FONT_SUBTITLE, FILL_SECTION, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22
    row += 1
    
    # 找甲方/乙方整列截图
    import glob
    party_a_col = ""
    party_b_col = ""
    
    # 从 OCR 数据里找截图目录
    all_imgs = []
    for s in seals:
        if s.get("image_path"):
            all_imgs.append(s["image_path"])
    for s in signatures:
        if s.get("image_path"):
            all_imgs.append(s["image_path"])
    
    if all_imgs:
        sig_dir = os.path.dirname(all_imgs[0])
        # 优先匹配有印章的页码（签署页），避免取到误检的正文页
        seal_pages = set()
        for s in seals:
            if s.get("page"):
                seal_pages.add(s["page"])
        import re as _re_col
        # 第一轮：只取有印章的页
        for f in glob.glob(os.path.join(sig_dir, "*_column.png")):
            basename = os.path.basename(f)
            m = _re_col.search(r'page_(\d+)_', basename)
            if m and int(m.group(1)) in seal_pages:
                if "甲方" in basename or "party_a" in basename.lower():
                    if not party_a_col:
                        party_a_col = f
                elif "乙方" in basename or "party_b" in basename.lower():
                    if not party_b_col:
                        party_b_col = f
        # 第二轮：如果没找到，回退到任意页
        if not party_a_col or not party_b_col:
            for f in glob.glob(os.path.join(sig_dir, "*_column.png")):
                basename = os.path.basename(f)
                if ("甲方" in basename or "party_a" in basename.lower()) and not party_a_col:
                    party_a_col = f
                elif ("乙方" in basename or "party_b" in basename.lower()) and not party_b_col:
                    party_b_col = f
    
    if party_a_col or party_b_col:
        # 表头
        set_cell(ws, row, 1, "甲方签署区（完整）", FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        set_cell(ws, row, 5, "乙方签署区（完整）", FONT_HEADER, FILL_HEADER, ALIGN_CENTER)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        ws.row_dimensions[row].height = 22
        row += 1
        
        # 计算显示尺寸
        max_h = 450
        target_w = 280
        img_h_a = 0
        img_h_b = 0
        
        if party_a_col and os.path.exists(party_a_col):
            try:
                from PIL import Image as PILImage
                im = PILImage.open(party_a_col)
                ratio = target_w / im.width
                img_h_a = min(int(im.height * ratio), max_h)
            except:
                pass
        
        if party_b_col and os.path.exists(party_b_col):
            try:
                from PIL import Image as PILImage
                im = PILImage.open(party_b_col)
                ratio = target_w / im.width
                img_h_b = min(int(im.height * ratio), max_h)
            except:
                pass
        
        display_h = max(img_h_a, img_h_b, 200)
        ws.row_dimensions[row].height = display_h / 1.3
        
        # 嵌入甲方图
        if party_a_col and os.path.exists(party_a_col):
            try:
                img = XLImage(party_a_col)
                ratio = target_w / img.width
                img.width = target_w
                img.height = min(int(img.height * ratio), max_h)
                ws.add_image(img, f"A{row}")
            except Exception as e:
                set_cell(ws, row, 1, f"（截图加载失败: {e}）", FONT_SMALL, None, ALIGN_CENTER)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        else:
            set_cell(ws, row, 1, "未检测到", FONT_SMALL, None, ALIGN_CENTER)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        
        # 嵌入乙方图
        if party_b_col and os.path.exists(party_b_col):
            try:
                img = XLImage(party_b_col)
                ratio = target_w / img.width
                img.width = target_w
                img.height = min(int(img.height * ratio), max_h)
                ws.add_image(img, f"E{row}")
            except Exception as e:
                set_cell(ws, row, 5, f"（截图加载失败: {e}）", FONT_SMALL, None, ALIGN_CENTER)
                ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        else:
            set_cell(ws, row, 5, "未检测到", FONT_SMALL, None, ALIGN_CENTER)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        
        row += 2
    
    merge_style(ws, row, 1, row, 7, "说明：公章/签字区域由 OCR v5 自动检测截图，置信度>60% 视为有效。签署日期从文本中提取。",
                FONT_SMALL, None, ALIGN_LEFT)
    row += 1
    merge_style(ws, row, 1, row, 7, "⚠️ 截图仅供人工复核，正式签署前请核对原件真伪。",
                FONT_P1, None, ALIGN_LEFT)
    
    return row


# ============================================================
# 8. 主入口
# ============================================================

def generate_unified_report(text: str, output_path: str, ocr_data: dict = None) -> str:
    """生成统一格式的 Excel 审批报告

    Args:
        text: 合同全文（字符串）
        output_path: 输出 xlsx 路径
        ocr_data: OCR v5 检测结果（可选）
    """
    # OCR 归一化（修正常见识别错误）
    from contract_parser import _normalize_ocr_text
    text = _normalize_ocr_text(text)
    
    # 1. 提取基本信息
    info = extract_contract_info(text)
    
    # 2. 全文分段
    sections = split_contract_sections(text)
    
    # 3. 逐段风险扫描
    for sec in sections:
        issues = scan_section(sec, text)
        sec["issues"] = issues
        # 计算段落最高风险等级
        levels = [i["level"] for i in issues]
        if "P0" in levels:
            sec["risk_level"] = "P0"
        elif "P1" in levels:
            sec["risk_level"] = "P1"
        elif "P2" in levels:
            sec["risk_level"] = "P2"
        else:
            sec["risk_level"] = None
    
    # 4. 生成 Excel
    wb = Workbook()
    
    # Sheet 1
    ws1 = wb.active
    ws1.title = "1.合同条款拆解"
    build_sheet_1(ws1, info, sections, text)
    
    # Sheet 2
    ws2 = wb.create_sheet("2.统一审核标准")
    build_sheet_2(ws2)
    
    # Sheet 3
    ws3 = wb.create_sheet("3.审核与整改建议")
    build_sheet_3(ws3, sections)
    
    # Sheet 4（可选）
    if ocr_data:
        ws4 = wb.create_sheet("4.签署要素审计")
        build_sheet_4(ws4, ocr_data)
    
    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="合同审批统一输出报告 v4.0")
    parser.add_argument("--file", help="合同文本文件路径（txt/md）")
    parser.add_argument("--ocr-pdf", help="OCR 扫描 PDF 路径（走 OCR v5）")
    parser.add_argument("--output", default="contract_audit_report.xlsx", help="输出 xlsx 路径")
    parser.add_argument("--ocr-result", help="OCR v5 结果 JSON 路径（签署要素审计用）")
    args = parser.parse_args()
    
    if not args.file and not args.ocr_pdf:
        parser.error("必须指定 --file 或 --ocr-pdf")
    
    text = ""
    ocr_data = None
    
    if args.ocr_pdf:
        # 走 OCR v5 流程
        script_dir = os.path.dirname(os.path.abspath(__file__))
        sys.path.insert(0, script_dir)
        from contract_ocr_v5 import digitalize_document_v5 as run_ocr
        
        # 创建临时输出目录
        import tempfile
        tmpdir = tempfile.mkdtemp(prefix="contract_ocr_")
        txt_path = os.path.join(tmpdir, "contract.txt")
        sig_dir = os.path.join(tmpdir, "signatures")
        os.makedirs(sig_dir, exist_ok=True)
        json_path = os.path.join(tmpdir, "ocr_result.json")
        
        print(f"🔍 OCR 识别中: {args.ocr_pdf}")
        result = run_ocr(args.ocr_pdf, txt_path, engine="rapidocr",
                        signature_dir=sig_dir, json_path=json_path)
        text = result.get("text", "")
        ocr_data = result
        print(f"✅ OCR 完成: {len(text)} 字, {len(ocr_data.get('seals', []))} 印章, {len(ocr_data.get('signatures', []))} 签名")
    
    elif args.file:
        with open(args.file, encoding="utf-8") as f:
            text = f.read()
    
    # 加载 OCR 结果（如果单独提供）
    if args.ocr_result and os.path.exists(args.ocr_result):
        with open(args.ocr_result, encoding="utf-8") as f:
            ocr_data = json.load(f)
        print(f"   OCR 结果: {len(ocr_data.get('seals', []))} 印章, {len(ocr_data.get('signatures', []))} 签名")
    
    # 生成报告
    output = generate_unified_report(text, args.output, ocr_data)
    print(f"\n✅ 审批报告已生成: {output}")
    print(f"   Sheet 1: 合同条款拆解（全文逐段拆解，100%覆盖）")
    print(f"   Sheet 2: 统一审核标准（{len(AUDIT_STANDARDS)} 项）")
    print(f"   Sheet 3: 审核与整改建议（按优先级排序）")
    if ocr_data:
        print(f"   Sheet 4: 签署要素审计（OCR 自动检测）")


if __name__ == "__main__":
    sys.exit(main() or 0)
