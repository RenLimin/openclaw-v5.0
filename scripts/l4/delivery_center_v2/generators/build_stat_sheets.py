#!/usr/bin/env python3
"""
BDMS 交付月报 V2 — 统计 Sheet 生成器
从 ONES CSV 原始数据直接计算并生成透视表格式的 Excel Sheet
复用旧版已经验证的逻辑
"""
import sqlite3
import json
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 精确格式配置（从参考表提取）
import sys
sys.path.insert(0, str(Path(__file__).parent.parent / "config"))
from sheet_formats import SHEET_FORMAT_MAP

# === 全局配置（从主生成器导入常量） ===
ONES_DIR = Path.home() / ".openclaw" / "data" / "ones_exports"
REF_DIR = Path("/Users/bangcle/Bangcle Workspace/01. Management/2026/2026团队报告/202606")
REPORT_DATE = "2026-06-30"
REPORT_MONTH = "202606"
CONFIG_DIR = Path(__file__).parent.parent / "config"

def _write_df_to_sheet(ws, df: pd.DataFrame, sheet_name: str = None, start_row: int = 1, start_col: int = 1):
    """把 DataFrame 写到指定 worksheet
    格式精确匹配参考手工报表（从 SHEET_FORMAT_MAP 读取精确值）
    """
    from openpyxl.utils import get_column_letter
    
    # 获取该 sheet 的精确格式配置
    fmt = SHEET_FORMAT_MAP.get(sheet_name, None) if sheet_name else None
    
    # 列宽：优先用精确配置，否则自动计算
    column_widths = []
    if fmt and fmt.get('widths'):
        for i in range(len(df.columns)):
            if i < len(fmt['widths']):
                column_widths.append(fmt['widths'][i])
            else:
                column_widths.append(10.83203125)
    else:
        for i, col in enumerate(df.columns):
            if len(df) > 0:
                col_str = df.iloc[:, i].astype(str)
                lengths = col_str.map(lambda x: sum(2 if ord(c) > 127 else 1 for c in str(x)) if x is not None and not pd.isna(x) else 3)
                max_len = max(lengths.max(), len(str(col)))
            else:
                max_len = len(str(col))
            width = max_len + 2 if max_len + 2 >= 8 else 8
            column_widths.append(width)
    
    # 写入表头（带样式）
    for col_idx, (col_name, width) in enumerate(zip(df.columns, column_widths), start_col):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # 行高：精确配置优先
    if fmt and fmt.get('row_heights'):
        for row_num, height in fmt['row_heights'].items():
            ws.row_dimensions[row_num].height = height
    else:
        ws.row_dimensions[start_row].height = 20
    
    # 写入数据（带边框和字体）
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, (value, width) in enumerate(zip(row, column_widths), start_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER
    
    # 冻结窗格：精确位置
    if fmt and fmt.get('freeze'):
        ws.freeze_panes = fmt['freeze']
    else:
        ws.freeze_panes = f"A{start_row + 1}"

# 统一样式常量，精确匹配参考表（从手工报表styles.xml提取）
# 表头字体：微软雅黑 10号 加粗 白色
HEADER_FONT = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
# 表头底色：精确 FF2D73BA（蓝色）
HEADER_FILL = PatternFill(start_color="FF2D73BA", end_color="FF2D73BA", fill_type="solid")
# 数据字体：微软雅黑 10号 普通 黑色
DATA_FONT = Font(name="微软雅黑", size=10, color="FF000000")
# 细边框
THIN_BORDER = Border(
    left=Side(style='thin', color="FF000000"),
    right=Side(style='thin', color="FF000000"),
    top=Side(style='thin', color="FF000000"),
    bottom=Side(style='thin', color="FF000000")
)
# 表头对齐：居中
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
# 数据对齐：左对齐
DATA_ALIGN = Alignment(horizontal="left", vertical="center")

# 全局数据缓存
_sign_df = None
_sign_df_full = None  # 全量签约数据（不过滤立项日期）
_poc_df = None
_poc_df_full = None  # 全量 POC 数据（不过滤立项日期）
_abnormal_df = None
_rev_df = None
_acc_df = None
_sign_formula_df = None  # 签约公式列落盘数据
_poc_formula_df = None   # POC公式列落盘数据
_abnormal_formula_df = None  # 异常公式列落盘数据


def _try_read_csv(path, **kwargs):
    """尝试读取CSV，不存在返回None"""
    if path.exists():
        return pd.read_csv(path, low_memory=False, **kwargs)
    return None


def _load_data(force=False):
    """延迟加载源数据"""
    global _sign_df, _sign_df_full, _poc_df, _poc_df_full, _abnormal_df, _rev_df, _acc_df
    if not force and _sign_df is not None and _sign_df_full is not None and _abnormal_df is not None and _poc_df_full is not None:
        return
    
    # 主数据源：ones_exports 目录
    _sign_df = _try_read_csv(ONES_DIR / "签约项目统计.csv")
    _poc_df = _try_read_csv(ONES_DIR / "poc_提前实施.csv")
    _poc_df_full = _poc_df.copy() if _poc_df is not None else None
    
    # 异常数据：优先使用完整版（55列），回退到标准版（40列）
    _abnormal_df = _try_read_csv(ONES_DIR / "202606-签约项目异常处置.csv")
    if _abnormal_df is None or '异常影响情况' not in _abnormal_df.columns:
        _abnormal_df = _try_read_csv(ONES_DIR / "异常处置.csv")
    
    # 全量签约数据（不过滤立项日期，用于产品-授权&维保统计等需要全量数据的场景）
    # 必须在 _sign_df 被过滤前复制
    _sign_df_full = _sign_df.copy() if _sign_df is not None else None
    
    # 按报告日期过滤（与主生成器保持一致）
    if _sign_df is not None and '立项日期' in _sign_df.columns:
        dt = pd.to_datetime(_sign_df['立项日期'].astype(str).str[:10], errors='coerce')
        _sign_df = _sign_df[dt <= pd.to_datetime(REPORT_DATE)].copy()
    
    if _poc_df is not None and '立项日期' in _poc_df.columns:
        dt = pd.to_datetime(_poc_df['立项日期'].astype(str).str[:10], errors='coerce')
        _poc_df = _poc_df[dt <= pd.to_datetime(REPORT_DATE)].copy()
    
    # 如果异常数据缺失或不完整，尝试从 REF_DIR 加载（55列完整版）
    if _abnormal_df is None or '异常报备日期' not in _abnormal_df.columns:
        ref_abn = _try_read_csv(REF_DIR / "202606-签约项目异常处置.csv")
        if ref_abn is not None:
            _abnormal_df = ref_abn


def _compute_status_category(df):
    """计算履约项统计状态（财报-交付/确收状态）"""
    status_map = {
        '实施未开始': '1：正常交付',
        '义务已拆分': '1：正常交付',
        '实施进行中': '1：正常交付',
        '实施已完成': '4：正常验收',
        '交付邮件交接中': '1：正常交付',
        '交付邮件已归档': '4：正常验收',
        '验收文件交接中': '4：正常验收',
        '验收文件已归档': '7：正常服务',
    }
    return df['状态'].map(status_map).fillna('')


def _compute_poc_duration(df, sign_df=None):
    """计算提前实施项目持续周期
    返回: (duration_days, duration_stat, is_linked)
    
    已关联合同的项目：持续周期 = 关联合同归档日期 - 立项日期
    未关联合同的项目：持续周期 = 报告期末 - 立项日期
    
    关联合同归档日期通过销售合同编号从签约数据中查找。
    """
    start = pd.to_datetime(df['立项日期'], errors='coerce')
    
    report_end = pd.Timestamp(f"{REPORT_MONTH[:4]}-{REPORT_MONTH[4:]}-01") + pd.offsets.MonthEnd(0)
    
    # 是否已关联合同（有销售合同编号且非空）
    has_contract = df['销售合同编号'].notna() & (df['销售合同编号'].astype(str).str.strip() != '') & (df['销售合同编号'].astype(str).str.strip() != 'nan')
    
    # 计算关联合同归档日期
    # 优先用 df 自身的合同归档日期，否则从签约数据查找
    link_archive_date = pd.to_datetime(df['合同归档日期'], errors='coerce')
    
    # 如果提供了签约数据，用销售合同编号匹配查找合同归档日期
    if sign_df is not None and '销售合同编号' in sign_df.columns:
        sign_archive = sign_df.dropna(subset=['合同归档日期']).drop_duplicates(
            subset=['销售合同编号'], keep='first'
        ).set_index('销售合同编号')['合同归档日期']
        sign_archive_dt = pd.to_datetime(sign_archive, errors='coerce')
        
        # 用签约数据中找到的归档日期填充空白
        mask_need_fill = has_contract & link_archive_date.isna()
        contract_nos = df.loc[mask_need_fill, '销售合同编号'].astype(str).str.strip()
        matched = contract_nos.map(sign_archive_dt)
        link_archive_date.loc[mask_need_fill] = matched.values
    
    # 计算天数
    duration = pd.Series(np.nan, index=df.index, dtype='float64')
    
    # 已关联且有关联归档日期
    mask_linked = has_contract & link_archive_date.notna() & start.notna()
    duration[mask_linked] = (link_archive_date[mask_linked] - start[mask_linked]).dt.days
    
    # 已关联但无归档日期 - 用报告期末计算
    mask_linked_no_date = has_contract & link_archive_date.isna() & start.notna()
    duration[mask_linked_no_date] = (report_end - start[mask_linked_no_date]).dt.days
    
    # 未关联：用报告期末 - 立项日期
    mask_no_contract = ~has_contract & start.notna()
    duration[mask_no_contract] = (report_end - start[mask_no_contract]).dt.days
    
    # 周期分类
    def classify(days):
        if pd.isna(days):
            return '#N/A'
        days = int(days)
        if days <= 0:
            return '1个月内'
        elif days <= 30:
            return '1个月内'
        elif days <= 90:
            return '3个月内'
        elif days <= 180:
            return '6个月内'
        elif days <= 365:
            return '1年内'
        else:
            return '超过1年'
    
    duration_stat = duration.apply(classify)
    is_linked = has_contract.map({True: '已关联', False: '未关联'})
    
    return duration, duration_stat, is_linked


# ============================================================
# 1. 签约统计
# ============================================================
def build_sign_stats(ws):
    """签约统计：左表=按年份，右表=状态×年份交叉表
    精确匹配参考报表的 15行×15列 透视表格式
    
    参考结构：
    - 行1-3: 筛选器（字段名 + (全部)）
    - 行4: 空
    - 行5: 列名行（行标签 | 计数项:ID | 空×3 | 计数项:ID | 列标签 | 年份... | 总计）
    - 行6-14: 数据行（左：年份计数，右：状态×年份交叉表）
    - 行15-16: 总计行
    """
    _load_data()
    df = _sign_df.copy()
    
    df['立项年份'] = pd.to_datetime(df['立项日期'].astype(str).str[:10], errors='coerce').dt.year
    df['履约项统计状态'] = _compute_status_category(df)
    
    # === 筛选器行（行1-3），左右两份 ===
    filter_labels = ["项目经理所属部门", "统计项目编号", "项目状态"]
    for row in range(1, 4):
        ws.cell(row=row, column=1, value=filter_labels[row-1])
        ws.cell(row=row, column=2, value="(全部)")
        ws.cell(row=row, column=6, value=filter_labels[row-1])
        ws.cell(row=row, column=7, value="(全部)")
    
    # 行5：列名行
    # 参考：列1=行标签, 列2=计数项:ID, 列3-5=空, 列6=计数项:ID, 列7=列标签, 列8+=年份, 最后=总计
    ws.cell(row=5, column=1, value="行标签")
    ws.cell(row=5, column=2, value="计数项:ID")
    # 列3-5留空（3个空列）
    ws.cell(row=5, column=6, value="计数项:ID")
    ws.cell(row=5, column=7, value="列标签")
    
    # === 左表：按立项年份统计 ===
    year_counts = df.groupby('立项年份')['ID'].nunique()
    years = sorted([int(y) for y in year_counts.index if pd.notna(y)])
    
    # 确保包含 2019-2026 所有年份（即使某些年份无数据，也显示为0）
    if years:
        full_years = list(range(2019, max(years) + 1))
        for y in full_years:
            if y not in years:
                years.append(y)
                year_counts[y] = 0
        years = sorted(years)
    
    for i, year in enumerate(years):
        row = 6 + i
        ws.cell(row=row, column=1, value=f"{year}年")
        ws.cell(row=row, column=2, value=int(year_counts.get(year, 0)))
    
    # 左表总计行
    total_row = 6 + len(years)
    ws.cell(row=total_row, column=1, value="总计")
    ws.cell(row=total_row, column=2, value=int(df['ID'].nunique()))
    
    # === 右表：履约项统计状态 × 立项年份 交叉表 ===
    # 参考报表列顺序：2026, 2019, 2020, 2021, ..., 2025, 总计
    current_year = 2026
    ordered_years = [current_year] + [y for y in years if y != current_year]
    year_labels = [f"{y}年" for y in ordered_years]
    
    # 右表表头（行6）
    ws.cell(row=6, column=6, value="行标签")
    for ci, yl in enumerate(year_labels):
        ws.cell(row=6, column=7 + ci, value=yl)
    ws.cell(row=6, column=7 + len(year_labels), value="总计")
    
    status_order = [
        '1：正常交付', '2：应交未交', '3：交付异常', '4：正常验收',
        '5：应验未验', '6：验收异常', '7：正常服务', '8：应结未结', '9：已结项'
    ]
    
    import re
    pivot = df.pivot_table(
        index='履约项统计状态', columns='立项年份',
        values='ID', aggfunc='nunique', fill_value=0
    )
    
    for ri, status in enumerate(status_order):
        row = 7 + ri
        ws.cell(row=row, column=6, value=status)
        row_total = 0
        for ci, year in enumerate(ordered_years):
            val = int(pivot.loc[status, year]) if status in pivot.index and year in pivot.columns else 0
            ws.cell(row=row, column=7 + ci, value=val)
            row_total += val
        ws.cell(row=row, column=7 + len(ordered_years), value=row_total)
    
    # 右表总计行
    total_row_r = 7 + len(status_order)
    ws.cell(row=total_row_r, column=6, value="总计")
    col_totals = []
    for ci, year in enumerate(ordered_years):
        val = int(pivot[year].sum()) if year in pivot.columns else 0
        if val > 0:
            ws.cell(row=total_row_r, column=7 + ci, value=val)
        col_totals.append(val)
    ws.cell(row=total_row_r, column=7 + len(ordered_years), value=sum(col_totals))


# ============================================================
# 2. POC&提前实施统计
# ============================================================
def build_poc_stats(ws):
    """POC&提前实施统计：三表布局
    精确匹配参考报表的 22行×29列 格式
    
    参考结构（22行×29列，1-based）：
    - 行1-4: 筛选器（三组筛选条件）
    - 行5: 列头行
    - 行6-22: 数据行
    
    左表（列1-4）: 履约项立项期间 × 类型(POC/提前实施/总计)
    中表（列8-15）: 提前实施履约项持续周期 × 是否关联(未关联/已关联/各部门)
    右表（列19-29）: POC项目工时合计 × 部门
    """
    _load_data()
    df = _poc_df.copy()
    
    df['立项年份'] = pd.to_datetime(df['立项日期'].astype(str).str[:10], errors='coerce').dt.year
    
    early_df = df[df['项目类型(概览)'] == '提前实施'].copy()
    poc_df = df[df['项目类型(概览)'] == 'POC'].copy()
    
    _, duration_stat, is_linked = _compute_poc_duration(early_df, _sign_df)
    early_df['持续周期-统计'] = duration_stat.values
    early_df['是否关联合同'] = is_linked.values
    
    # === 筛选器区域（行1-4）===
    # 行1: 列7=项目经理所属部门, 列8=(全部) | 列18=项目经理所属部门, 列19=(全部)
    ws.cell(row=1, column=8, value="项目经理所属部门")
    ws.cell(row=1, column=9, value="(全部)")
    ws.cell(row=1, column=19, value="项目经理所属部门")
    ws.cell(row=1, column=20, value="(全部)")
    
    # 行2: 列1=项目经理所属部门, 列2=(全部) | 列7=统计项目编号, 列8=(全部) | 列18=统计项目编号, 列19=(全部)
    ws.cell(row=2, column=1, value="项目经理所属部门")
    ws.cell(row=2, column=2, value="(全部)")
    ws.cell(row=2, column=8, value="统计项目编号")
    ws.cell(row=2, column=9, value="(全部)")
    ws.cell(row=2, column=19, value="统计项目编号")
    ws.cell(row=2, column=19, value="(全部)")
    
    # 行3: 列1=统计项目编号, 列2=(全部) | 列7=项目类型(概览), 列8=提前实施 | 列18=项目类型(概览), 列19=POC
    ws.cell(row=3, column=1, value="统计项目编号")
    ws.cell(row=3, column=2, value="(全部)")
    ws.cell(row=3, column=8, value="项目类型(概览)")
    ws.cell(row=3, column=9, value="提前实施")
    ws.cell(row=3, column=19, value="项目类型(概览)")
    ws.cell(row=3, column=19, value="POC")
    
    # 行5: 字段名行（与REF一致）===
    ws.cell(row=5, column=1, value="履约项立项期间")
    ws.cell(row=5, column=2, value="列标签")
    ws.cell(row=5, column=8, value="提前实施履约项持续周期")
    ws.cell(row=5, column=9, value="列标签")
    ws.cell(row=5, column=19, value="求和项:POC项目工时合计（小时）")
    ws.cell(row=5, column=20, value="列标签")
    
    # === 行6: 列标签行（与REF一致）===
    # 左表（列1-4）
    ws.cell(row=6, column=1, value="行标签")
    ws.cell(row=6, column=2, value="POC")
    ws.cell(row=6, column=3, value="提前实施")
    ws.cell(row=6, column=4, value="总计")
    
    # 中表（列8-15），列11=#N/A
    ws.cell(row=6, column=8, value="行标签")
    ws.cell(row=6, column=9, value="超过1年")
    ws.cell(row=6, column=10, value="1个月内")
    ws.cell(row=6, column=11, value="#N/A")
    ws.cell(row=6, column=12, value="3个月内")
    ws.cell(row=6, column=13, value="6个月内")
    ws.cell(row=6, column=14, value="1年内")
    ws.cell(row=6, column=15, value="总计")
    
    # 右表（列19-29），列26为空
    right_dept_cols = [
        "华中营销部", "西区营销部", "南区营销部", "东区营销部",
        "北区营销部", "北区金融部",
        None,  # 列26 空列
        "车联网行业部", "销售运营管理部",
        "总计"
    ]
    
    ws.cell(row=6, column=19, value="行标签")
    for ci, dept in enumerate(right_dept_cols):
        if dept is not None:
            ws.cell(row=6, column=20 + ci, value=dept)
    
    # === 左表：履约项立项期间 × 类型（数据从行7开始）===
    year_type = df.groupby(['立项年份', '项目类型(概览)'])['ID'].nunique().unstack(fill_value=0)
    years = sorted([int(y) for y in year_type.index if pd.notna(y)])
    
    poc_total = 0
    early_total = 0
    
    for i, year in enumerate(years):
        row = 7 + i
        ws.cell(row=row, column=1, value=f"{year}年")
        poc_count = int(year_type.loc[year, 'POC']) if 'POC' in year_type.columns and year in year_type.index else 0
        early_count = int(year_type.loc[year, '提前实施']) if '提前实施' in year_type.columns and year in year_type.index else 0
        ws.cell(row=row, column=2, value=poc_count)
        ws.cell(row=row, column=3, value=early_count)
        ws.cell(row=row, column=4, value=poc_count + early_count)
        poc_total += poc_count
        early_total += early_count
    
    total_row_left = 7 + len(years)
    ws.cell(row=total_row_left, column=1, value="总计")
    ws.cell(row=total_row_left, column=2, value=poc_total)
    ws.cell(row=total_row_left, column=3, value=early_total)
    ws.cell(row=total_row_left, column=4, value=poc_total + early_total)
    
    # === 中表：提前实施履约项持续周期（数据从行7开始）===
    dur_order = ['超过1年', '1个月内', '#N/A', '3个月内', '6个月内', '1年内']
    dur_col_offset = [9, 10, 11, 12, 13, 14]
    
    def write_mid_row(row_num, label, pivot_series):
        ws.cell(row=row_num, column=8, value=label)
        row_total = 0
        for ci, dur in enumerate(dur_order):
            v = int(pivot_series.get(dur, 0))
            ws.cell(row=row_num, column=dur_col_offset[ci], value=v)
            row_total += v
        ws.cell(row=row_num, column=15, value=row_total)
    
    mid_row = 7
    
    # 未关联汇总
    unlinked = early_df[early_df['是否关联合同'] == '未关联']
    unlinked_piv = unlinked.groupby('持续周期-统计')['ID'].nunique()
    write_mid_row(mid_row, "未关联", unlinked_piv)
    mid_row += 1
    
    # 未关联各部门（按部门名称排序，仅展示总部级部门，排除分公司）
    unlinked_depts = sorted([d for d in unlinked['责任销售所属团队'].dropna().unique() if '分公司' not in str(d)])
    for dept in unlinked_depts:
        dept_data = unlinked[unlinked['责任销售所属团队'] == dept]
        dpiv = dept_data.groupby('持续周期-统计')['ID'].nunique()
        write_mid_row(mid_row, str(dept), dpiv)
        mid_row += 1
    
    # 未关联 #N/A 部门
    unlinked_na = unlinked[unlinked['责任销售所属团队'].isna()]
    if len(unlinked_na) > 0:
        dpiv = unlinked_na.groupby('持续周期-统计')['ID'].nunique()
        write_mid_row(mid_row, "#N/A", dpiv)
        mid_row += 1
    
    # 已关联汇总
    linked = early_df[early_df['是否关联合同'] == '已关联']
    linked_piv = linked.groupby('持续周期-统计')['ID'].nunique()
    write_mid_row(mid_row, "已关联", linked_piv)
    mid_row += 1
    
    # 已关联各部门（按部门名称排序，仅展示总部级部门，排除分公司）
    linked_depts = sorted([d for d in linked['责任销售所属团队'].dropna().unique() if '分公司' not in str(d)])
    for dept in linked_depts:
        dept_data = linked[linked['责任销售所属团队'] == dept]
        dpiv = dept_data.groupby('持续周期-统计')['ID'].nunique()
        write_mid_row(mid_row, str(dept), dpiv)
        mid_row += 1
    
    # 已关联 #N/A 部门
    linked_na = linked[linked['责任销售所属团队'].isna()]
    if len(linked_na) > 0:
        dpiv = linked_na.groupby('持续周期-统计')['ID'].nunique()
        write_mid_row(mid_row, "#N/A", dpiv)
        mid_row += 1
    
    # 总计行
    all_piv = early_df.groupby('持续周期-统计')['ID'].nunique()
    write_mid_row(mid_row, "总计", all_piv)
    mid_row += 1
    
    # === 右表：POC项目工时合计 × 产线（工时=0占位，数据从行7开始）===
    prod_lines = sorted(poc_df['所属产线'].dropna().unique())
    
    right_row = 7
    for pl in prod_lines:
        ws.cell(row=right_row, column=19, value=str(pl))
        for ci in range(len(right_dept_cols)):
            if right_dept_cols[ci] is not None:
                ws.cell(row=right_row, column=20 + ci, value=0)
        right_row += 1
    
    # (空白)行
    ws.cell(row=right_row, column=19, value="(空白)")
    for ci in range(len(right_dept_cols)):
        if right_dept_cols[ci] is not None:
            ws.cell(row=right_row, column=20 + ci, value=0)
    right_row += 1
    
    # 总计行
    ws.cell(row=right_row, column=19, value="总计")
    for ci in range(len(right_dept_cols)):
        if right_dept_cols[ci] is not None:
            ws.cell(row=right_row, column=20 + ci, value=0)
    right_row += 1


# ============================================================
# 3. 异常统计
# ============================================================
def build_abnormal_stats(ws):
    """异常统计：5张交叉表横向排列
    精确匹配参考报表的 28行×54列 格式
    
    5张表布局（列区间，均为1-based）：
    - 表1: 列1-11 (11列) - 异常报备期间-合同归档年度 × 影响情况
    - 表2: 列15-25 (11列) - 异常归档期间-合同归档年度
    - 表3: 列29-35 (7列) - 处理中/异常类别-合同归档年度(2026年)
    - 表4: 列40-48 (10列) - 处理中/异常类别-合同归档年度(全部)
    - 表5: 列53-54 (2列) - 销售事业部-合同归档年度
    
    行结构（1-based）：
    - 行1-4: 筛选器区域
    - 行5: 表1 & 表2 表头; 表3 月筛选
    - 行6-7: 表1 年份汇总; 表2 <2025/3/17 + 2025年
    - 行8-19: 表1&2 月份明细(2025年1-12月); 表3&4 类别区域; 表5 事业部区域
    - 行20-26: 表1&2 2026年月份明细(1-6月)
    - 行27: 表1 & 表2 总计行
    """
    _load_data()
    df = _abnormal_df.copy()
    
    # 确保基础列存在
    if df is None or '合同归档日期' not in df.columns:
        # 用0填充但保持结构
        _fill_abnormal_empty(ws)
        return
    
    df['合同归档年份'] = pd.to_datetime(df['合同归档日期'].astype(str).str[:10], errors='coerce').dt.year
    
    has_report_date = '异常报备日期' in df.columns
    has_archive_date = '异常归档日期' in df.columns
    has_impact = '异常影响情况' in df.columns
    has_category = '异常项目-类别' in df.columns
    has_dept = '事业部（区域）' in df.columns
    if not has_dept:
        has_dept = '责任销售所属团队' in df.columns
    
    if has_report_date:
        df['报备_dt'] = pd.to_datetime(df['异常报备日期'].astype(str).str[:10], errors='coerce')
        df['报备年份'] = df['报备_dt'].dt.year
        df['报备月份'] = df['报备_dt'].dt.month
    else:
        df['报备_dt'] = pd.NaT
        df['报备年份'] = np.nan
        df['报备月份'] = np.nan
    
    if has_archive_date:
        df['归档_dt'] = pd.to_datetime(df['异常归档日期'].astype(str).str[:10], errors='coerce')
        df['归档年份'] = df['归档_dt'].dt.year
        df['归档月份'] = df['归档_dt'].dt.month
    else:
        df['归档_dt'] = pd.NaT
        df['归档年份'] = np.nan
        df['归档月份'] = np.nan
    
    # 过滤"4：不统计"
    if has_impact and '异常影响情况' in df.columns:
        df_stat = df[df['异常影响情况'] != '4：不统计'].copy()
    else:
        df_stat = df.copy()
    
    # =====================================================
    # 定义年份列
    # =====================================================
    # 表1 & 表2 年份列：2016, 2019-2026, 总计 (10个数据列)
    t1_years = [2016, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]
    t1_col_count = len(t1_years) + 1  # +1 for 总计
    
    # 表3 年份列：2022-2026, 总计 (6个数据列)
    t3_years = [2022, 2023, 2024, 2025, 2026]
    t3_col_count = len(t3_years) + 1
    
    # 表4 年份列：2019-2026, 总计 (8个数据列)
    t4_years = [2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]
    t4_col_count = len(t4_years) + 1
    
    # 标准异常类别（8种，不足补0）
    std_categories = [
        '1：甲方不具备交付条件',
        '2：甲方未按照合同约定验收',
        '3：甲方确认终止但无终止协议下单',
        '4：甲方需求/期限变更但无补充协议下单',
        '5：缺少穿透验收单（渠道-最终用户）',
        '6：项目启动延期',
        '7：交付资源不足',
        '8：其他',
    ]
    
    # 标准事业部（10个，取数据中前10个）
    if has_dept:
        all_depts = sorted(df_stat[df_stat['状态'] != '已完成']['责任销售所属团队'].dropna().unique())
        # 优先取总部部门（不含分公司），再补充分公司，凑10个
        hq_depts = [d for d in all_depts if '分公司' not in str(d)]
        branch_depts = [d for d in all_depts if '分公司' in str(d)]
        std_depts = (hq_depts + branch_depts)[:10]
    else:
        std_depts = [f'事业部{i}' for i in range(1, 11)]
    
    # =====================================================
    # 辅助函数：写入一行交叉表数据
    # =====================================================
    def write_pivot_row(row_num, col_start, label, data_series, year_list, show_total=True):
        """写入一行透视表数据
        Args:
            row_num: 行号(1-based)
            col_start: 起始列(1-based)，即行标签列
            label: 行标签文本
            data_series: {year: count} 的字典或Series
            year_list: 年份列表
            show_total: 是否显示总计列
        """
        ws.cell(row=row_num, column=col_start, value=label)
        total = 0
        for ci, year in enumerate(year_list):
            val = int(data_series.get(year, 0)) if isinstance(data_series, dict) else int(data_series.get(year, 0))
            ws.cell(row=row_num, column=col_start + 1 + ci, value=val)
            total += val
        if show_total:
            ws.cell(row=row_num, column=col_start + 1 + len(year_list), value=total)
    
    def get_year_counts(data, year_col='合同归档年份', val_col='ID'):
        """从数据中获取按年份的计数字典"""
        if len(data) == 0:
            return {}
        counts = data.groupby(year_col)[val_col].nunique()
        # 转换为 int 键的字典
        return {int(k): int(v) for k, v in counts.items() if pd.notna(k)}
    
    # =====================================================
    # 筛选器区域（行1-7，1-based）
    # =====================================================
    # 行1: 异常影响情况 + (多项) — 5张表都有
    for col_start in [1, 15, 29, 40, 53]:
        ws.cell(row=1, column=col_start, value="异常影响情况")
        ws.cell(row=1, column=col_start + 1, value="(多项)")
    
    # 行2: 状态 + (全部) — 5张表都有
    for col_start in [1, 15, 29, 40, 53]:
        ws.cell(row=2, column=col_start, value="状态")
        ws.cell(row=2, column=col_start + 1, value="(全部)")
    
    # 行3: 项目经理团队 + (全部) — 5张表都有
    for col_start in [1, 15, 29, 40, 53]:
        ws.cell(row=3, column=col_start, value="项目经理团队")
        ws.cell(row=3, column=col_start + 1, value="(全部)")
    
    # 行4: 年(异常报备日期) — 仅表3/4/5
    ws.cell(row=4, column=29, value="年(异常报备日期)")
    ws.cell(row=4, column=30, value="2026年")
    ws.cell(row=4, column=40, value="年(异常报备日期)")
    ws.cell(row=4, column=41, value="(全部)")
    ws.cell(row=4, column=53, value="年(异常报备日期)")
    ws.cell(row=4, column=54, value="(全部)")
    
    # 行5: 期间/月筛选 — 表1/2是表标题，表3/4/5是月(异常报备日期)
    ws.cell(row=5, column=1, value="异常报备期间-合同归档年度")
    ws.cell(row=5, column=2, value="列标签")
    ws.cell(row=5, column=15, value="异常归档期间-合同归档年度")
    ws.cell(row=5, column=16, value="列标签")
    ws.cell(row=5, column=29, value="月(异常报备日期)")
    ws.cell(row=5, column=30, value="(全部)")
    ws.cell(row=5, column=40, value="月(异常报备日期)")
    ws.cell(row=5, column=41, value="(全部)")
    ws.cell(row=5, column=53, value="月(异常报备日期)")
    ws.cell(row=5, column=54, value="(全部)")
    
    # 行6: 表1/2表头行 + 表3/4/5的年(异常归档日期)筛选器
    ws.cell(row=6, column=29, value="年(异常归档日期)")
    ws.cell(row=6, column=30, value="(全部)")
    ws.cell(row=6, column=40, value="年(异常归档日期)")
    ws.cell(row=6, column=41, value="<2025/3/17")
    ws.cell(row=6, column=53, value="年(异常归档日期)")
    ws.cell(row=6, column=54, value="<2025/3/17")
    
    # 行7: 表1/2数据行(2024年/<2025/3/17) + 表3/4/5的月(异常归档日期)筛选器
    ws.cell(row=7, column=29, value="月(异常归档日期)")
    ws.cell(row=7, column=30, value="(全部)")
    ws.cell(row=7, column=40, value="月(异常归档日期)")
    ws.cell(row=7, column=41, value="(全部)")
    ws.cell(row=7, column=53, value="月(异常归档日期)")
    ws.cell(row=7, column=54, value="(全部)")
    
    # =====================================================
    # 表1：异常报备期间-合同归档年度 × 影响情况（列1-11）
    # =====================================================
    # 行6 (1-based): 表头
    ws.cell(row=6, column=1, value="行标签")
    for ci, y in enumerate(t1_years):
        ws.cell(row=6, column=2 + ci, value=f"{y}年")
    ws.cell(row=6, column=2 + len(t1_years), value="总计")
    
    # 行7-8 (1-based): 年份汇总 (2024, 2025)
    if has_report_date:
        for i, ry in enumerate([2024, 2025]):
            y_data = df_stat[df_stat['报备年份'] == ry]
            y_counts = get_year_counts(y_data)
            write_pivot_row(7 + i, 1, f"{ry}年", y_counts, t1_years)
    else:
        for i, ry in enumerate([2024, 2025]):
            write_pivot_row(7 + i, 1, f"{ry}年", {}, t1_years)
    
    # 行9-20 (1-based): 2025年月份明细 (1-12月, 12行)
    if has_report_date:
        y2025_data = df_stat[df_stat['报备年份'] == 2025]
        for m in range(1, 13):
            m_data = y2025_data[y2025_data['报备月份'] == m]
            m_counts = get_year_counts(m_data)
            write_pivot_row(8 + m, 1, f"{m}月", m_counts, t1_years)
    else:
        for m in range(1, 13):
            write_pivot_row(8 + m, 1, f"{m}月", {}, t1_years)
    
    # 行21-26 (1-based): 2026年月份明细 (1-6月, 6行) 截至6月30，所以只有到6月
    if has_report_date:
        y2026_data = df_stat[df_stat['报备年份'] == 2026]
        for m in range(1, 7):
            m_data = y2026_data[y2026_data['报备月份'] == m]
            m_counts = get_year_counts(m_data)
            write_pivot_row(20 + m, 1, f"{m}月", m_counts, t1_years)
    else:
        for m in range(1, 7):
            write_pivot_row(20 + m, 1, f"{m}月", {}, t1_years)
    
    # 行27 (1-based): 总计行
    total_counts = get_year_counts(df_stat)
    write_pivot_row(27, 1, "总计", total_counts, t1_years)
    
    # =====================================================
    # 表2：异常归档期间-合同归档年度（列15-25）
    # =====================================================
    col2 = 15
    # 行6 (1-based): 表头
    ws.cell(row=6, column=col2, value="行标签")
    for ci, y in enumerate(t1_years):
        ws.cell(row=6, column=col2 + 1 + ci, value=f"{y}年")
    ws.cell(row=6, column=col2 + 1 + len(t1_years), value="总计")
    
    if has_archive_date:
        archived = df_stat[df_stat['归档_dt'].notna()]
        
        # 行7 (1-based): <2025/3/17
        cutoff = pd.Timestamp('2025-03-17')
        pre_data = archived[archived['归档_dt'] < cutoff]
        pre_counts = get_year_counts(pre_data)
        write_pivot_row(7, col2, "<2025/3/17", pre_counts, t1_years)
        
        # 行8 (1-based): 2025年
        a2025 = archived[archived['归档年份'] == 2025]
        a2025_counts = get_year_counts(a2025)
        write_pivot_row(8, col2, "2025年", a2025_counts, t1_years)
        
        # 行9-20 (1-based): 2025年月份明细 1-12月
        for m in range(1, 13):
            m_data = a2025[a2025['归档月份'] == m]
            m_counts = get_year_counts(m_data)
            write_pivot_row(8 + m, col2, f"{m}月", m_counts, t1_years)
        
        # 行21-26 (1-based): 2026年月份明细 1-6月
        a2026 = archived[archived['归档年份'] == 2026]
        for m in range(1, 7):
            m_data = a2026[a2026['归档月份'] == m]
            m_counts = get_year_counts(m_data)
            write_pivot_row(20 + m, col2, f"{m}月", m_counts, t1_years)
        
        # 行27 (1-based): 总计
        arc_total_counts = get_year_counts(archived)
        write_pivot_row(27, col2, "总计", arc_total_counts, t1_years)
    else:
        # 无数据，填充结构
        write_pivot_row(7, col2, "<2025/3/17", {}, t1_years)
        write_pivot_row(8, col2, "2025年", {}, t1_years)
        for m in range(1, 13):
            write_pivot_row(8 + m, col2, f"{m}月", {}, t1_years)
        for m in range(1, 7):
            write_pivot_row(20 + m, col2, f"{m}月", {}, t1_years)
        write_pivot_row(27, col2, "总计", {}, t1_years)
    
    # =====================================================
    # 表3：处理中/异常类别-合同归档年度（列29-35）
    # 筛选条件：年(异常报备日期)=2026年, 年(异常归档日期)=(全部)
    # 数据行：行8标题, 行9表头, 行10-16类别数据+总计
    # =====================================================
    col3 = 29
    # 行9: 标题行
    ws.cell(row=9, column=col3, value="处理中/异常类别-合同归档年度")
    ws.cell(row=9, column=col3 + 1, value="列标签")
    
    # 行10: 表头
    ws.cell(row=10, column=col3, value="行标签")
    for ci, y in enumerate(t3_years):
        ws.cell(row=10, column=col3 + 1 + ci, value=f"{y}年")
    ws.cell(row=10, column=col3 + 1 + len(t3_years), value="总计")
    
    # 获取2026年报备的数据
    # REF 表3 total=27，对应 2026年报备 + 验收类（2：验收 + 3：确收+验收）
    if has_report_date and has_category:
        proc_2026 = df_stat[
            (df_stat['报备年份'] == 2026) & 
            (df_stat['异常影响情况'].isin(['2：验收', '3：确收+验收']))
        ]
    else:
        proc_2026 = df_stat.iloc[0:0]  # empty DataFrame
    
    # 行10-14: 5种异常类别（REF 中有数据的 5 类）
    # 从数据中获取实际存在的类别，按顺序排列
    if has_category and len(proc_2026) > 0:
        actual_cats = [c for c in std_categories if c in proc_2026['异常项目-类别'].values]
    else:
        actual_cats = std_categories[:5]
    
    # 确保至少有5行（用空类别补齐）
    while len(actual_cats) < 5:
        actual_cats.append('')
    
    for ci, cat in enumerate(actual_cats[:5]):
        row_num = 10 + ci
        if has_category and len(proc_2026) > 0 and cat:
            cat_data = proc_2026[proc_2026['异常项目-类别'] == cat]
            cat_counts = get_year_counts(cat_data)
        else:
            cat_counts = {}
        write_pivot_row(row_num, col3, cat if cat else '', cat_counts, t3_years)
    
    # 行16: 总计行
    if len(proc_2026) > 0:
        total_counts = get_year_counts(proc_2026)
    else:
        total_counts = {}
    write_pivot_row(15, col3, "总计", total_counts, t3_years)
    
    # =====================================================
    # 表4：处理中/异常类别-合同归档年度（列40-48）
    # 筛选条件：年(异常报备日期)=(全部), 年(异常归档日期)=<2025/3/17(=未归档)
    # 行标签：异常项目-处置方案（8种 + 总计 = 9行）
    # REF total=72, 对应 验收类（2：验收 + 3：确收+验收）+ 未归档
    # =====================================================
    col4 = 40
    # 行9: 标题行
    ws.cell(row=9, column=col4, value="处理中/异常类别-合同归档年度")
    ws.cell(row=9, column=col4 + 1, value="列标签")
    
    # 行10: 表头
    ws.cell(row=10, column=col4, value="行标签")
    for ci, y in enumerate(t4_years):
        ws.cell(row=10, column=col4 + 1 + ci, value=f"{y}年")
    ws.cell(row=10, column=col4 + 1 + len(t4_years), value="总计")
    
    # 标准处置方案（8种）
    std_plans = [
        '1：坏账处理',
        '2：签署终止协议',
        '3：签署补充协议',
        '4：销售协助交付',
        '5：销售协助验收',
        '6：内部验收',
        '7：待销售反馈',
        '8：建议坏账处理',
    ]
    
    # 获取数据：验收类 + 未归档（异常归档日期为空 = <2025/3/17 筛选效果）
    has_plan_col = '异常项目-处置方案' in df_stat.columns
    if has_plan_col and has_archive_date:
        proc_all = df_stat[
            (df_stat['归档_dt'].isna()) &
            (df_stat['异常影响情况'].isin(['2：验收', '3：确收+验收']))
        ]
    elif has_plan_col:
        # 无归档日期字段，用状态筛选
        proc_all = df_stat[
            (df_stat['状态'] != '已完成') &
            df_stat['异常影响情况'].isin(['2：验收', '3：确收+验收'])
        ]
    else:
        proc_all = df_stat.iloc[0:0]
    
    # 行11-18: 8种处置方案
    for ci, plan in enumerate(std_plans):
        row_num = 11 + ci
        if has_plan_col and len(proc_all) > 0:
            # 处置方案可能包含组合值（如"6：内部验收、7：待销售反馈"）
            plan_prefix = plan.split('：')[0] + '：'
            plan_data = proc_all[proc_all['异常项目-处置方案'].astype(str).str.contains(plan_prefix, na=False)]
            plan_counts = get_year_counts(plan_data)
        else:
            plan_counts = {}
        write_pivot_row(row_num, col4, plan, plan_counts, t4_years)
    
    # 行19: 总计行
    if len(proc_all) > 0:
        total_counts = get_year_counts(proc_all)
    else:
        total_counts = {}
    write_pivot_row(19, col4, "总计", total_counts, t4_years)
    
    # =====================================================
    # 表5：销售事业部-合同归档年度（列53-54）
    # 筛选：年(异常报备日期)=(全部), 年(异常归档日期)=<2025/3/17(=未归档)
    # REF total=76, 对应 确收类（1：确收 + 3：确收+验收）+ 未归档
    # 行9: 表头, 行10-19: 10个事业部, 行20: 总计
    # =====================================================
    col5 = 53
    # 行9: 表头（REF 在第9行）
    ws.cell(row=9, column=col5, value="行标签")
    ws.cell(row=9, column=col5 + 1, value="销售事业部-合同归档年度")
    
    # 使用 事业部（区域）列（从 55 列完整版 CSV）
    dept_col = '事业部（区域）' if '事业部（区域）' in df_stat.columns else None
    if not dept_col and '责任销售所属团队' in df_stat.columns:
        dept_col = '责任销售所属团队'
    
    # 获取数据：确收类 + 未归档
    if dept_col and has_archive_date:
        dept_df = df_stat[
            (df_stat['归档_dt'].isna()) &
            (df_stat['异常影响情况'].isin(['1：确收', '3：确收+验收']))
        ]
    elif dept_col:
        # 回退：用状态筛选 + 确收类
        dept_df = df_stat[
            (df_stat['状态'] != '已完成') &
            (df_stat['异常影响情况'].isin(['1：确收', '3：确收+验收']))
        ]
    else:
        dept_df = df_stat.iloc[0:0]
    
    # 获取事业部列表（REF 中的 10 个事业部，按数量降序）
    std_depts5 = []
    if dept_col and len(dept_df) > 0:
        dept_counts = dept_df.groupby(dept_col)['ID'].nunique().sort_values(ascending=False)
        # 排除 NaN 和 '其他'
        for dept_name in dept_counts.index:
            if pd.notna(dept_name) and str(dept_name) != 'nan' and str(dept_name) != '其他':
                std_depts5.append(dept_name)
            if len(std_depts5) >= 10:
                break
    else:
        std_depts5 = [f'事业部{i}' for i in range(1, 11)]
    
    # 行10-19: 10个事业部数据
    total_count = 0
    for di, dept in enumerate(std_depts5[:10]):
        row_num = 10 + di
        ws.cell(row=row_num, column=col5, value=dept)
        if dept_col and len(dept_df) > 0:
            val = int(dept_df[dept_df[dept_col] == dept]['ID'].nunique()) if dept in dept_df[dept_col].values else 0
        else:
            val = 0
        ws.cell(row=row_num, column=col5 + 1, value=val)
        total_count += val
    
    # 如果不足 10 个，补空行
    for di in range(len(std_depts5), 10):
        row_num = 10 + di
        ws.cell(row=row_num, column=col5, value='')
        ws.cell(row=row_num, column=col5 + 1, value=0)
    
    # 行20: 总计行
    ws.cell(row=20, column=col5, value="总计")
    ws.cell(row=20, column=col5 + 1, value=total_count)
    
    # 确保54列 × 28行 结构完整（填充空单元格确保维度）
    for r in range(1, 29):
        for c in range(1, 55):
            # 访问单元格以确保存在
            _ = ws.cell(row=r, column=c)


def _fill_abnormal_empty(ws):
    """无数据时填充异常统计Sheet结构（28行×54列）"""
    # 简化版：只写入结构，全部填0
    t1_years = [2016, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]
    t3_years = [2022, 2023, 2024, 2025, 2026]
    t4_years = [2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]
    std_categories = [f'类别{i}' for i in range(1, 9)]
    std_depts = [f'事业部{i}' for i in range(1, 11)]
    
    # 筛选器
    for col_start in [1, 15, 29, 40, 53]:
        ws.cell(row=1, column=col_start, value="异常影响情况")
        ws.cell(row=1, column=col_start + 1, value="(多项)")
        ws.cell(row=2, column=col_start, value="状态")
        ws.cell(row=2, column=col_start + 1, value="(全部)")
        ws.cell(row=3, column=col_start, value="项目经理团队")
        ws.cell(row=3, column=col_start + 1, value="(全部)")
    
    ws.cell(row=4, column=29, value="年(异常报备日期)")
    ws.cell(row=4, column=30, value="2026年")
    ws.cell(row=4, column=40, value="年(异常报备日期)")
    ws.cell(row=4, column=41, value="(全部)")
    ws.cell(row=4, column=53, value="年(异常报备日期)")
    ws.cell(row=4, column=54, value="(全部)")
    
    ws.cell(row=5, column=1, value="异常报备期间-合同归档年度")
    ws.cell(row=5, column=2, value="列标签")
    ws.cell(row=5, column=15, value="异常归档期间-合同归档年度")
    ws.cell(row=5, column=16, value="列标签")
    ws.cell(row=5, column=29, value="月(异常报备日期)")
    ws.cell(row=5, column=30, value="(全部)")
    ws.cell(row=5, column=40, value="月(异常报备日期)")
    ws.cell(row=5, column=41, value="(全部)")
    ws.cell(row=5, column=53, value="月(异常报备日期)")
    ws.cell(row=5, column=54, value="(全部)")
    
    for r in range(1, 29):
        for c in range(1, 55):
            _ = ws.cell(row=r, column=c)


# ============================================================
# 4. 产品-授权&维保统计
# ============================================================
def build_product_stats(ws):
    """产品-授权&维保统计：左表年份×产品大类，右表产品大类×客户名称
    精确匹配参考报表的 49行×16列 格式"""
    _load_data()
    df = _sign_df_full.copy()
    
    df['立项年份'] = pd.to_datetime(df['立项日期'].astype(str).str[:10], errors='coerce').dt.year
    
    # 筛选：产品大类只包含 "产品"、"维保+服务"两类
    df_prod = df[df['所属产线'].isin(['产品', '维保+服务'])].copy()
    
    # === 筛选器区域：行1-4 ===
    filter_labels = ["立项年份", "项目类型(概览)", "责任销售所属团队", "产品大类"]
    for row in range(1, 5):
        ws.cell(row=row, column=1, value=filter_labels[row-1])
        ws.cell(row=row, column=2, value="(全部)")
        ws.cell(row=row, column=11, value=filter_labels[row-1])
        ws.cell(row=row, column=12, value="(全部)")
    
    # === 左表：立项年份 × 产品大类 ===
    # 行6: 表头
    ws.cell(row=6, column=1, value="行标签")
    ws.cell(row=6, column=2, value="计数项:ID")
    
    years = sorted([int(y) for y in df_prod['立项年份'].dropna().unique() if pd.notna(y)])
    if 2019 not in years:
        years.insert(0, 2019)
    
    for i, year in enumerate(years):
        row = 7 + i
        ws.cell(row=row, column=1, value=f"{year}年")
        cnt = int(df_prod[df_prod['立项年份'] == year]['ID'].nunique())
        ws.cell(row=row, column=2, value=cnt)
    
    # 左表总计行
    total_row = 7 + len(years)
    ws.cell(row=total_row, column=1, value="总计")
    ws.cell(row=total_row, column=2, value=int(df_prod['ID'].nunique()))
    
    # === 右表：产品大类 × 客户名称 交叉表 ===
    # 行6: 表头
    ws.cell(row=6, column=11, value="行标签")
    ws.cell(row=6, column=12, value="计数项:ID")
    
    prod_lines = sorted(df_prod['所属产线'].dropna().unique())
    for i, line in enumerate(prod_lines):
        row = 7 + i
        ws.cell(row=row, column=11, value=str(line))
        cnt = int(df_prod[df_prod['所属产线'] == line]['ID'].nunique())
        ws.cell(row=row, column=12, value=cnt)
    
    # 右表总计行
    total_row_r = 7 + len(prod_lines)
    ws.cell(row=total_row_r, column=11, value="总计")
    ws.cell(row=total_row_r, column=12, value=int(df_prod['ID'].nunique()))


# ============================================================
# 5. 提前实施分事业部统计
# ============================================================
def build_poc_dept_stats(ws):
    """提前实施分事业部统计：左表未关联/已关联/各部门，右表周期分布 × 部门
    精确匹配参考报表的 31行×20列 格式"""
    _load_data()
    df = _poc_df.copy()
    
    df['立项年份'] = pd.to_datetime(df['立项日期'].astype(str).str[:10], errors='coerce').dt.year
    
    early_df = df[df['项目类型(概览)'] == '提前实施'].copy()
    _, duration_stat, is_linked = _compute_poc_duration(early_df, _sign_df)
    early_df['持续周期-统计'] = duration_stat.values
    early_df['是否关联合同'] = is_linked.values
    
    # === 筛选器行 ===
    filter_labels = ["立项年份", "项目经理所属部门", "持续周期"]
    for row in range(1, 4):
        ws.cell(row=row, column=1, value=filter_labels[row-1])
        ws.cell(row=row, column=2, value="(全部)")
        ws.cell(row=row, column=11, value=filter_labels[row-1])
        ws.cell(row=row, column=12, value="(全部)")
    
    # === 左表：未关联+已关联+各部门汇总 ===
    # 行6: 表头
    ws.cell(row=6, column=1, value="行标签")
    ws.cell(row=6, column=2, value="计数项:ID")
    
    # 未关联汇总
    row_num = 7
    unlinked = early_df[early_df['是否关联合同'] == '未关联']
    ws.cell(row=row_num, column=1, value="未关联")
    ws.cell(row=row_num, column=2, value=len(unlinked))
    row_num += 1
    
    # 未关联各部门（排序，不含分公司）
    unlinked_depts = sorted([d for d in unlinked['责任销售所属团队'].dropna().unique() if '分公司' not in str(d)])
    for dept in unlinked_depts:
        cnt = len(unlinked[unlinked['责任销售所属团队'] == dept])
        ws.cell(row=row_num, column=1, value=str(dept))
        ws.cell(row=row_num, column=2, value=cnt)
        row_num += 1
    
    # 已关联汇总
    linked = early_df[early_df['是否关联合同'] == '已关联']
    ws.cell(row=row_num, column=1, value="已关联")
    ws.cell(row=row_num, column=2, value=len(linked))
    row_num += 1
    
    # 已关联各部门（排序，不含分公司）
    linked_depts = sorted([d for d in linked['责任销售所属团队'].dropna().unique() if '分公司' not in str(d)])
    for dept in linked_depts:
        cnt = len(linked[linked['责任销售所属团队'] == dept])
        ws.cell(row=row_num, column=1, value=str(dept))
        ws.cell(row=row_num, column=2, value=cnt)
        row_num += 1
    
    # 总计行
    ws.cell(row=row_num, column=1, value="总计")
    ws.cell(row=row_num, column=2, value=len(early_df))
    row_num += 1
    
    # === 右表：持续周期 × 部门 交叉表 ===
    # 行6: 表头
    dur_order = ['超过1年', '1个月内', '3个月内', '6个月内', '1年内', '总计']
    ws.cell(row=6, column=11, value="行标签")
    for ci, dur in enumerate(dur_order):
        ws.cell(row=6, column=12 + ci, value=dur)
    
    depts = sorted([d for d in early_df['责任销售所属团队'].dropna().unique() if '分公司' not in str(d)])
    
    row_num = 7
    for dept in depts:
        dept_data = early_df[early_df['责任销售所属团队'] == dept]
        dpiv = dept_data.groupby('持续周期-统计')['ID'].nunique()
        ws.cell(row=row_num, column=11, value=str(dept))
        row_total = 0
        for ci, dur in enumerate(dur_order[:-1]):
            val = int(dpiv.get(dur, 0))
            ws.cell(row=row_num, column=12 + ci, value=val)
            row_total += val
        ws.cell(row=row_num, column=12 + 5, value=row_total)
        row_num += 1
    
    # 总计行
    all_piv = early_df.groupby('持续周期-统计')['ID'].nunique()
    ws.cell(row=row_num, column=11, value="总计")
    row_total = 0
    for ci, dur in enumerate(dur_order[:-1]):
        val = int(all_piv.get(dur, 0))
        ws.cell(row=row_num, column=12 + ci, value=val)
        row_total += val
    ws.cell(row=row_num, column=12 + 5, value=row_total)
    row_num += 1


# ============================================================
# 6. 交付异常分事业部统计
# ============================================================
def build_abnormal_dept_stats(ws):
    """交付异常分事业部统计：未处理、处理中、已完成 三个透视表横向排列
    精确匹配参考报表的 27行×28列 格式"""
    _load_data()
    df = _abnormal_df.copy()
    
    if df is None:
        # 填充结构
        for r in range(1, 28):
            for c in range(1, 29):
                _ = ws.cell(row=r, column=c)
        return
    
    has_dept = '事业部（区域）' in df.columns
    if not has_dept:
        has_dept = '责任销售所属团队' in df.columns
        if not has_dept:
            for r in range(1, 28):
                for c in range(1, 29):
                    _ = ws.cell(row=r, column=c)
            return
    
    dept_col = '事业部（区域）' if '事业部（区域）' in df.columns else '责任销售所属团队'
    df['合同归档年份'] = pd.to_datetime(df['合同归档日期'].astype(str).str[:10], errors='coerce').dt.year
    
    # 状态分类：未处理/处理中/已完成
    state_map = {
        '营销处置中': '处理中',
        '交付处置中': '处理中',
        '已完成': '已完成',
        '未开始': '未处理',
        '进行中': '未处理',
    }
    df['状态分类'] = df['状态'].map(lambda s: state_map.get(s, '未处理'))
    
    # === 表头行 ===
    ws.cell(row=1, column=1, value="状态分类")
    ws.cell(row=1, column=2, value="(全部)")
    ws.cell(row=1, column=11, value="状态分类")
    ws.cell(row=1, column=12, value="(全部)")
    ws.cell(row=1, column=21, value="状态分类")
    ws.cell(row=1, column=22, value="(全部)")
    ws.cell(row=2, column=1, value="合同归档年份")
    ws.cell(row=2, column=2, value="行标签")
    ws.cell(row=2, column=11, value="合同归档年份")
    ws.cell(row=2, column=12, value="行标签")
    ws.cell(row=2, column=21, value="合同归档年份")
    ws.cell(row=2, column=22, value="行标签")
    ws.cell(row=3, column=1, value="行标签")
    ws.cell(row=3, column=11, value="行标签")
    ws.cell(row=3, column=21, value="行标签")
    
    # === 年份 ===
    years = [2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]
    for ci, y in enumerate(years):
        ws.cell(row=4, column=1 + 1 + ci, value=y)
        ws.cell(row=4, column=11 + 1 + ci, value=y)
        ws.cell(row=4, column=21 + 1 + ci, value=y)
    
    # 获取事业部列表（降序排列）
    depts = df.groupby(dept_col)['ID'].nunique().sort_values(ascending=False).index
    dept_list = [d for d in depts if pd.notna(d) and str(d) != 'nan' and '分公司' not in str(d)]
    
    # 每个状态一个表：未处理、处理中、已完成
    states = ['未处理', '处理中', '已完成']
    col_starts = [1, 11, 21]
    
    for s_idx, state in enumerate(states):
        col_start = col_starts[s_idx]
        state_data = df[df['状态分类'] == state]
        if len(state_data) == 0:
            continue
        
        # 每个事业部一行
        row_num = 5
        for di, dept in enumerate(dept_list):
            cnt_by_year = state_data[state_data[dept_col] == dept].groupby('合同归档年份')['ID'].nunique()
            ws.cell(row=row_num, column=col_start, value=str(dept))
            total_dept = 0
            for ci, y in enumerate(years):
                val = int(cnt_by_year.get(y, 0))
                ws.cell(row=row_num, column=col_start + 1 + ci, value=val)
                total_dept += val
            total_dept += val
            row_num += 1
        
        # 总计行
        total_by_year = state_data.groupby('合同归档年份')['ID'].nunique()
        ws.cell(row=row_num, column=col_start, value="总计")
        for ci, y in enumerate(years):
            val = int(total_by_year.get(y, 0))
            ws.cell(row=row_num, column=col_start + 1 + ci, value=val)


# ============================================================
# 7. 交接统计
# ============================================================
def build_handover_stats(ws):
    """确收/验收交接统计：确收/验收两张透视表横向排列
    精确匹配参考报表的 18行×20列 格式"""
    _load_data()
    
    # 读取确收/验收原始数据
    rev_df = _try_read_csv(REF_DIR / "202606确收凭证交接-确收.csv", dtype=str)
    acc_df = _try_read_csv(REF_DIR / "202606确收凭证交接-验收.csv", dtype=str)
    
    if rev_df is None and acc_df is None:
        for r in range(1, 19):
            for c in range(1, 21):
                _ = ws.cell(row=r, column=c)
        return
    
    # 改名对齐
    if "深圳分公司-营销" in rev_df.columns:
        rev_df.rename(columns={"深圳分公司-营销": "销售部门"}, inplace=True)
    if "深圳分公司-营销" in acc_df.columns:
        acc_df.rename(columns={"深圳分公司-营销": "销售部门"}, inplace=True)
    
    # 筛选条件行
    filter_labels = ["销售部门", "交接日期", "财务是否接收"]
    for row in range(1, 4):
        ws.cell(row=row, column=1, value=filter_labels[row-1])
        ws.cell(row=row, column=2, value="(全部)")
        ws.cell(row=row, column=11, value=filter_labels[row-1])
        ws.cell(row=row, column=12, value="(全部)")
    
    # 表头行
    ws.cell(row=5, column=1, value="行标签")
    ws.cell(row=5, column=2, value="计数项:ID")
    ws.cell(row=5, column=11, value="行标签")
    ws.cell(row=5, column=12, value="计数项:ID")
    
    # 处理确收
    if rev_df is not None and "交接日期" in rev_df.columns:
        # 处理多种日期格式：提取前六位数字
        rev_df['year_month'] = rev_df['交接日期'].astype(str).str.replace(r'[^0-9]', '', regex=True).str[:6]
        rev_df['year_month'] = rev_df['year_month'].replace('', np.nan)
        ym_counts = rev_df.groupby('year_month')['ID'].nunique()
        sorted_ym = sorted([str(int(ym)) for ym in ym_counts.index if ym and not pd.isna(ym)])
        row_num = 6
        for ym in sorted_ym:
            cnt = int(ym_counts[ym])
            ws.cell(row=row_num, column=1, value=ym)
            ws.cell(row=row_num, column=2, value=cnt)
            row_num += 1
        # 总计
        ws.cell(row=row_num, column=1, value="总计")
        ws.cell(row=row_num, column=2, value=int(rev_df['ID'].nunique()))
    
    # 处理验收
    if acc_df is not None and "交接日期" in acc_df.columns:
        # 处理多种日期格式：提取前六位数字
        acc_df['year_month'] = acc_df['交接日期'].astype(str).str.replace(r'[^0-9]', '', regex=True).str[:6]
        acc_df['year_month'] = acc_df['year_month'].replace('', np.nan)
        ym_counts = acc_df.groupby('year_month')['ID'].nunique()
        sorted_ym = sorted([str(int(ym)) for ym in ym_counts.index if ym and not pd.isna(ym)])
        row_num = 6
        for ym in sorted_ym:
            cnt = int(ym_counts[ym])
            ws.cell(row=row_num, column=11, value=ym)
            ws.cell(row=row_num, column=12, value=cnt)
            row_num += 1
        # 总计
        ws.cell(row=row_num, column=11, value="总计")
        ws.cell(row=row_num, column=12, value=int(acc_df['ID'].nunique()))


# ============================================================
# 8. 异常台账
# ============================================================
def build_abnormal_ledger(ws):
    """异常台账：明细Sheet，按异常报备日期排序，35列
    精确匹配参考报表结构"""
    _load_data()
    df = _abnormal_df.copy()
    
    if df is None:
        return
    
    expected_cols = [
        "销售合同编号", "最终用户名称", "客户名称", "责任销售（履约项）",
        "责任销售所属团队", "负责人", "所属项目", "项目类型(概览)",
        "项目状态", "立项日期", "基线-预估结项日期", "实际结项日期",
        "合同名称", "直签或代理", "合同归档日期", "合同起始日期", "合同结束日期",
        "交付服务开始日期", "交付服务结束日期", "标题",
        "异常报备日期", "预估异常处置完成日期", "异常归档日期", "异常影响情况",
        "异常项目-类别", "异常项目-处置方案", "异常处置方案-影响",
        "交付说明（异常履约项统计类别）",
        "交付说明（履约项交付情况、合同交付条款）",
        "交付中心反馈", "营销中心反馈", "项目异常内容", "预估金额",
        "项目经理团队", "项目验收状态",
    ]
    
    # 按异常报备日期升序排列
    if "异常报备日期" in df.columns:
        df['_sort_dt'] = pd.to_datetime(df['异常报备日期'].astype(str).str[:10], errors='coerce')
        df = df.sort_values('_sort_dt', ascending=True).drop(columns=['_sort_dt'])
    else:
        df = df.sort_values("销售合同编号", ascending=True)
    
    # 只保留预期列，顺序对齐
    actual_cols = [c for c in expected_cols if c in df.columns]
    df = df[actual_cols]
    
    # 写入Sheet
    _write_df_to_sheet(ws, df, start_row=1, start_col=1)


# ============================================================
# 9. 图例
# ============================================================
def build_legend(ws):
    """图例 Sheet：项目经理 → 所属部门映射，34行 × 2列
    从旧版配置文件 `legend_pm_dept.json` 读取"""
    legend_path = Path(__file__).parent.parent / "delivery_center" / "config" / "legend_pm_dept.json"
    if legend_path.exists():
        legend = json.loads(legend_path.read_text(encoding="utf-8"))
    else:
        legend = {}
    
    # 写入表头
    ws.cell(row=1, column=1, value="项目经理")
    ws.cell(row=1, column=2, value="所属区域")
    
    # 写入数据
    row_num = 2
    for pm, region in sorted(legend.items()):
        ws.cell(row=row_num, column=1, value=pm)
        ws.cell(row=row_num, column=2, value=region)
        row_num += 1
