#!/usr/bin/env python3
"""
给已生成的交付月报应用精确格式（列宽、行高、冻结窗格），
所有数值均从手工参考报表 `2026交付月报-20260630.xlsx` 提取。
"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent / "config"))
from sheet_formats import SHEET_FORMAT_MAP

GEN_PATH = "/Users/bangcle/.openclaw/data/reports/交付月报-202606-v2.xlsx"

def apply_format_to_sheet(ws, fmt: dict):
    """给单个 worksheet 应用精确格式"""
    # 1. 列宽
    if fmt.get("widths"):
        for i, width in enumerate(fmt["widths"], 1):
            if i > ws.max_column:
                break
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = width
    
    # 2. 行高
    if fmt.get("row_heights"):
        for row_num, height in fmt["row_heights"].items():
            ws.row_dimensions[row_num].height = height
    
    # 3. 冻结窗格
    if fmt.get("freeze"):
        ws.freeze_panes = fmt["freeze"]


def main():
    import sys
    gen_path = sys.argv[1] if len(sys.argv) > 1 else GEN_PATH
    print(f"📂 打开文件: {gen_path}")
    wb = load_workbook(gen_path)
    
    applied = 0
    for sheet_name, fmt in SHEET_FORMAT_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  跳过 {sheet_name}（不存在）")
            continue
        ws = wb[sheet_name]
        apply_format_to_sheet(ws, fmt)
        applied += 1
        print(f"  ✅ {sheet_name}: {len(fmt['widths'])}列宽 / {len(fmt['row_heights'])}行高 / 冻结={fmt['freeze']}")
    
    wb.save(gen_path)
    wb.close()
    print(f"\n✅ 完成！共应用 {applied} 个 Sheet 的精确格式")
    print(f"📄 输出: {gen_path}")


if __name__ == "__main__":
    main()
