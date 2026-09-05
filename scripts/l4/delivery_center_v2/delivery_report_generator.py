#!/usr/bin/env python3
"""
BDMS 交付月报 V2 — 生成器入口

版本：v2.0 (重新设计)
设计文档：docs/architecture/components/l4-delivery-center-v2/
"""

import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# 业务引擎
sys.path.insert(0, str(Path(__file__).parent / "engines"))
from status_engine import add_status_columns
from scoring_engine import add_scoring_columns
from mapping_engine import (
    add_mapping_columns,
    add_simple_computed_columns,
    add_exception_lookup_columns
)
from exception_engine import build_exception_df
from handover_engine import build_revenue_handover_df, build_acceptance_handover_df
# 统计Sheet生成器
sys.path.insert(0, str(Path(__file__).parent / "generators"))
from build_stat_sheets import (
    build_sign_stats, build_poc_stats, build_abnormal_stats,
    build_product_stats, build_poc_dept_stats, build_abnormal_dept_stats,
    build_handover_stats, build_abnormal_ledger, build_legend
)
# utils
sys.path.insert(0, str(Path(__file__).parent / "utils"))
from poc_hours import get_poc_project_hours
# 精确格式配置（从参考表提取）
sys.path.insert(0, str(Path(__file__).parent / "config"))
from sheet_formats import SHEET_FORMAT_MAP

# 统一样式常量，精确匹配参考表（从手工报表styles.xml提取）
# 表头字体：微软雅黑 10号 加粗 白色
HEADER_FONT = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
# 表头底色：精确 FF2D73BA（蓝色）
HEADER_FILL = PatternFill(start_color="FF2D73BA", end_color="FF2D73BA", fill_type="solid")
# 数据字体：微软雅黑 10号 普通 黑色
DATA_FONT = Font(name="微软雅黑", size=10, color="FF000000")
# 细边框
THIN_BORDER = Border(
    left=Side(style='thin', color="FF000000"),
    right=Side(style='thin', color="FF000000"),
    top=Side(style='thin', color="FF000000"),
    bottom=Side(style='thin', color="FF000000")
)
# 表头对齐：居中
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
# 数据对齐：左对齐
DATA_ALIGN = Alignment(horizontal="left", vertical="center")

# ============================================================
# 常量
# ============================================================

# 报告截止日期（参考模板中第1行的日期）
# 报告截止日期：报告月份最后一天（在 main 中按 report_month 计算）
REPORT_DATE = "2026-06-30"
REPORT_MONTH = "202606"

# 数据源目录
ONES_DIR = Path.home() / ".openclaw" / "data" / "ones_exports"
# HANDOVER_DIR 按 report_month 动态构建（见 load_handover_* 函数）
REF_BASE_DIR = Path("/Users/bangcle/Bangcle Workspace/01. Management/2026/2026团队报告")
REPORT_OUTPUT_DIR = Path.home() / ".openclaw" / "data" / "reports"

# Sheet 名称（按顺序）
SHEET_ORDER = [
    "交付效率统计",
    "签约统计",
    "产品-授权&维保统计",
    "签约",
    "POC&提前实施统计",
    "提前实施分事业部统计",
    "POC&提前实施",
    "异常统计",
    "异常台账",
    "交付异常分事业部统计",
    "异常项目",
    "交接统计",
    "确收交接",
    "验收交接",
    "图例",
]


def load_ones_sign_contracts(report_month: str = None) -> pd.DataFrame:
    """加载 ONES 签约项目统计 CSV
    优先按月份查找 {report_month}周报-签约项目统计.csv，否则用默认签约项目统计.csv
    """
    if report_month:
        monthly_path = ONES_DIR / f"{report_month}周报-签约项目统计.csv"
        if monthly_path.exists():
            csv_path = monthly_path
        else:
            csv_path = ONES_DIR / "签约项目统计.csv"
    else:
        csv_path = ONES_DIR / "签约项目统计.csv"
    df = pd.read_csv(csv_path, dtype=str, encoding="utf-8")
    # 列名对齐：ONES 导出列名和模板列名的差异
    col_rename = {
        "履约项异常/变更备注": "履约项异常/变更类型",  # ONES 叫备注，模板叫类型，同一个东西
        "合同开始日期": "合同起始日期",  # ONES 叫开始日期，模板叫起始日期
    }
    df = df.rename(columns={k: v for k, v in col_rename.items() if k in df.columns})
    print(f"📊 签约项目统计: {len(df)} 行, {len(df.columns)} 列")
    return df


def load_ones_poc(report_month: str = None) -> pd.DataFrame:
    """加载 ONES POC&提前实施统计 CSV
    优先按月份查找 {report_month}周报-POC&提前实施统计.csv，否则用默认 poc_提前实施.csv
    """
    if report_month:
        monthly_path = ONES_DIR / f"{report_month}周报-POC&提前实施统计.csv"
        if monthly_path.exists():
            csv_path = monthly_path
        else:
            csv_path = ONES_DIR / "poc_提前实施.csv"
    else:
        csv_path = ONES_DIR / "poc_提前实施.csv"
    df = pd.read_csv(csv_path, dtype=str, encoding="utf-8")
    col_rename = {
        "履约项异常/变更备注": "履约项异常/变更类型",
        "合同开始日期": "合同起始日期",
    }
    df = df.rename(columns={k: v for k, v in col_rename.items() if k in df.columns})
    print(f"📊 POC&提前实施统计: {len(df)} 行, {len(df.columns)} 列")
    return df


def load_ones_exceptions(report_month: str = None) -> pd.DataFrame:
    """加载 ONES 异常项目处置 CSV（优先用 55 列完整版，含异常详情字段）"""
    # 优先按月份查找完整版
    if report_month:
        full_path = ONES_DIR / f"{report_month}-签约项目异常处置.csv"
    else:
        full_path = ONES_DIR / "202606-签约项目异常处置.csv"
    if full_path.exists():
        df = pd.read_csv(full_path, dtype=str, encoding="utf-8")
        # 如果有异常影响情况等关键字段，说明是完整版
        if "异常影响情况" in df.columns:
            col_rename = {
                "履约项异常/变更备注": "履约项异常/变更类型",
                "合同开始日期": "合同起始日期",
            }
            df = df.rename(columns={k: v for k, v in col_rename.items() if k in df.columns})
            print(f"📊 异常项目处置（完整版）: {len(df)} 行, {len(df.columns)} 列")
            return df
    
    # 回退：标准版（40 列）
    csv_path = ONES_DIR / "异常处置.csv"
    df = pd.read_csv(csv_path, dtype=str, encoding="utf-8")
    col_rename = {
        "履约项异常/变更备注": "履约项异常/变更类型",
        "合同开始日期": "合同起始日期",
    }
    df = df.rename(columns={k: v for k, v in col_rename.items() if k in df.columns})
    print(f"📊 异常项目处置: {len(df)} 行, {len(df.columns)} 列")
    return df


def load_handover_revenue(report_month: str = "202606") -> pd.DataFrame:
    """加载企业微信确收交接 CSV"""
    csv_path = REF_BASE_DIR / report_month / f"{report_month}确收凭证交接-确收.csv"
    df = pd.read_csv(csv_path, dtype=str, encoding="utf-8")
    # 根据 Rex 审核修正验收第 10 列列名
    if "深圳分公司-营销" in df.columns:
        df.rename(columns={"深圳分公司-营销": "销售部门"}, inplace=True)
    
    # 去掉全空列（Rex 要求：忽略 "情况" "0" 等全部为空的列）
    df = df.dropna(axis=1, how="all")
    
    print(f"📊 确收交接: {len(df)} 行 × {len(df.columns)} 列")
    return df


def load_handover_acceptance(report_month: str = "202606") -> pd.DataFrame:
    """加载企业微信验收交接 CSV"""
    csv_path = REF_BASE_DIR / report_month / f"{report_month}确收凭证交接-验收.csv"
    df = pd.read_csv(csv_path, dtype=str, encoding="utf-8")
    # 根据 Rex 审核修正验收第 10 列列名
    if "深圳分公司-营销" in df.columns:
        df.rename(columns={"深圳分公司-营销": "销售部门"}, inplace=True)
    print(f"📊 验收交接: {len(df)} 行 × {len(df.columns)} 列")
    return df


def _write_df_to_sheet(ws, df: pd.DataFrame, sheet_name: str = None, start_row: int = 1, start_col: int = 1):
    """把 DataFrame 写到指定 worksheet
    格式精确匹配参考手工报表（从 SHEET_FORMAT_MAP 读取精确值）：
    - 表头：微软雅黑 10号 加粗 白色字 + FF2D73BA 蓝色底色 + 居中 + 细边框
    - 数据：微软雅黑 10号 黑色字 + 左对齐 + 细边框
    - 列宽、行高、冻结窗格：全部从参考表精确提取
    """
    from openpyxl.utils import get_column_letter
    
    # 获取该 sheet 的精确格式配置
    fmt = SHEET_FORMAT_MAP.get(sheet_name, None) if sheet_name else None
    
    # 列宽：优先用精确配置，否则自动计算
    column_widths = []
    if fmt and fmt.get('widths'):
        for i in range(len(df.columns)):
            if i < len(fmt['widths']):
                column_widths.append(fmt['widths'][i])
            else:
                column_widths.append(10.83203125)
    else:
        for i, col in enumerate(df.columns):
            if len(df) > 0:
                col_str = df.iloc[:, i].astype(str)
                lengths = col_str.map(lambda x: sum(2 if ord(c) > 127 else 1 for c in str(x)) if x is not None and not pd.isna(x) else 3)
                max_len = max(lengths.max(), len(str(col)))
            else:
                max_len = len(str(col))
            width = max_len + 2 if max_len + 2 >= 8 else 8
            column_widths.append(width)
    
    # 写入表头（带样式）
    for col_idx, (col_name, width) in enumerate(zip(df.columns, column_widths), start_col):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # 行高：优先用精确配置
    if fmt and fmt.get('row_heights'):
        for row_num, height in fmt['row_heights'].items():
            ws.row_dimensions[row_num].height = height
    else:
        ws.row_dimensions[start_row].height = 20
    
    # 写入数据（带边框和字体）
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, (value, width) in enumerate(zip(row, column_widths), start_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER
    
    # 冻结窗格：精确位置
    if fmt and fmt.get('freeze'):
        ws.freeze_panes = fmt['freeze']
    else:
        ws.freeze_panes = f"A{start_row + 1}"


def build_sign_sheet_df(df_sign_raw: pd.DataFrame, df_exc: pd.DataFrame) -> pd.DataFrame:
    """构建签约 Sheet 完整 DataFrame（83 列）"""
    df = df_sign_raw.copy()

    # 过滤条件：只保留立项日期 <= 报告月份最后一天（2026-06-30）
    # 差异来源：ONES 导出在 2026-09 包含了 7-9 月新增数据，手工报表是截止 2026-06-30
    df["_立项_date"] = pd.to_datetime(df["立项日期"].astype(str).str[:10], errors="coerce")
    df = df[df["_立项_date"] <= pd.to_datetime(REPORT_DATE)]
    df.drop(columns=["_立项_date"], inplace=True)

    # 列 41-44：简单计算列
    df = add_simple_computed_columns(df)

    # 列 45-67：状态相关列
    df = add_status_columns(df, REPORT_DATE)

    # 列 68-76：考核相关列
    df = add_scoring_columns(df)

    # 列 77-79：映射相关列
    df = add_mapping_columns(df)

    # 列 80-83：异常关联列
    df = add_exception_lookup_columns(df, df_exc)

    # 确保列顺序正确（和参考报表一致）
    expected_cols = [
        # 原始 40 列
        "BI履约ID", "最终用户名称", "客户名称", "责任销售（履约项）", "责任销售所属团队",
        "负责人", "所属项目", "项目类型(概览)", "项目状态", "立项日期",
        "基线-预估结项日期", "实际结项日期", "销售合同编号", "合同名称", "直签或代理",
        "合同归档日期", "合同起始日期", "合同结束日期", "交付服务开始日期", "交付服务结束日期",
        "合同验收条款", "验收时点", "验收方式", "标题", "标准产品/服务序号",
        "履约类型", "所属产线", "状态", "履约项异常/变更类型", "履约项优先级",
        "预估交付完成日期", "预算-预估交付完成日期", "交付邮件发送日期",
        "实际服务/授权开始日期", "实际服务/授权结束日期",
        "预估验收完成日期", "预算-预估验收完成日期", "备注", "PMO备注", "ID",
        # 简单计算列 (41-44)
        "项目编号", "统计项目编号", "统计合同编号", "合同归档年度",
        # 状态标记列 (45-54)
        "实施未开始", "义务已拆分", "实施进行中", "实施已完成",
        "交付邮件交接中", "交付邮件已归档", "验收文件交接中", "验收文件已归档",
        "履约项合计", "校验",
        # 统计状态列 (55-67)
        "项目统计状态", "履约项统计状态（即，财报-交付/确收状态）",
        "1：正常交付", "2：应交未交", "3：交付异常", "4：正常验收",
        "5：应验未验", "6：验收异常", "7：正常服务", "8：应结未结", "9：已结项",
        "统计校验", "项目验收状态（即，财报-验收状态）",
        # 考核列 (68-76)
        "基线-预估结项日期（计算后反填）",
        "交付计划准确率“差异”", "交付计划准确率“提前/延后”",
        "交付计划准确率“是否跨月”", "交付计划准确率-考核扣分",
        "按时交付率“差异”", "按时交付率“提前/延后”",
        "按时交付率“是否跨月”", "按时交付率-考核扣分",
        # 映射列 (77-79)
        "项目经理", "项目经理所属部门", "销售团队-统计",
        # 异常关联列 (80-83)
        "异常项目对比", "异常处置状态", "异常影响情况", "交付说明",
    ]

    # 过滤存在的列（有些列可能因为数据缺失没有）
    actual_cols = [c for c in expected_cols if c in df.columns]
    df = df[actual_cols]

    return df


def add_poc_computed_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    给 POC Sheet 添加几个计算列：
    1. 提前实施项目持续周期（天）：实际结项日期 - 立项日期
    2. 提前实施项目持续周期-统计：分档分类
    3. 提前实施项目是否已关联合同：是否关联合同 → 是/否
    4. 关联合同归档日期：从合同表里关联（TODO）
    5. 统计所属项目：项目编号
    """
    from datetime import datetime

    def _calc_duration(l_date, r_date):
        l_dt = pd.to_datetime(l_date.astype(str).str[:10], errors="coerce")
        r_dt = pd.to_datetime(r_date.astype(str).str[:10], errors="coerce")
        return (r_dt - l_dt).dt.days

    df["提前实施项目持续周期（天）"] = _calc_duration(df["立项日期"], df["实际结项日期"])

    def _duration_bucket(days):
        if pd.isna(days):
            return ""
        if days <= 30:
            return "1个月内"
        elif days <= 90:
            return "3个月内"
        elif days <= 180:
            return "6个月内"
        else:
            return "超过1年"

    df["提前实施项目持续周期-统计"] = df["提前实施项目持续周期（天）"].apply(_duration_bucket)
    df["提前实施项目是否已关联合同"] = df["销售合同编号"].notna().map({True: "是", False: "否"})
    # TODO: 关联合同归档日期需要 OA 表关联
    # df["关联合同归档日期"] = ...
    df["统计所属项目"] = df["项目编号"]

    return df


def build_poc_sheet_df(df_poc_raw: pd.DataFrame, df_exc: pd.DataFrame) -> pd.DataFrame:
    """构建 POC&提前实施 Sheet 完整 DataFrame（84 列）"""
    df = df_poc_raw.copy()

    # 过滤条件：和签约一样，只保留立项日期 <= 报告月份最后一天（2026-06-30）
    df["_立项_date"] = pd.to_datetime(df["立项日期"].astype(str).str[:10], errors="coerce")
    df = df[df["_立项_date"] <= pd.to_datetime(REPORT_DATE)]
    df.drop(columns=["_立项_date"], inplace=True)

    # 列 41-44：简单计算列
    df = add_simple_computed_columns(df)

    # POC 增加计算列
    df = add_poc_computed_columns(df)

    # 列 45-67：状态相关列
    df = add_status_columns(df, REPORT_DATE)

    # 列 68-76：考核相关列
    df = add_scoring_columns(df)

    # 列 77-79：映射相关列 — 注意顺序：这里项目经理已经是计算列，原始列不要保留
    df = add_mapping_columns(df)

    # POC 工时合计：按项目汇总工时
    poc_hours = get_poc_project_hours()
    df = df.merge(poc_hours, left_on="项目编号", right_on="项目编号", how="left")

    # 确保列顺序正确（和参考报表一致）
    expected_cols = [
        # 原始 40 列
        "BI履约ID", "最终用户名称", "客户名称", "责任销售（履约项）", "责任销售所属团队",
        "负责人", "所属项目", "项目类型(概览)", "项目状态", "立项日期",
        "基线-预估结项日期", "实际结项日期", "销售合同编号", "合同名称", "直签或代理",
        "合同归档日期", "合同起始日期", "合同结束日期", "交付服务开始日期", "交付服务结束日期",
        "合同验收条款", "验收时点", "验收方式", "标题", "标准产品/服务序号",
        "履约类型", "所属产线", "状态", "履约项异常/变更类型", "履约项优先级",
        "预估交付完成日期", "预算-预估交付完成日期", "交付邮件发送日期",
        "实际服务/授权开始日期", "实际服务/授权结束日期",
        "预估验收完成日期", "预算-预估验收完成日期", "备注", "PMO备注", "ID",
        # 简单计算列 (41-44)
        "项目编号", "统计项目编号", "统计合同编号", "合同归档年度",
        # 状态标记列 (45-53)
        "实施未开始", "义务已拆分", "实施进行中", "实施已完成",
        "交付邮件交接中", "交付邮件已归档", "验收文件交接中", "验收文件已归档",
        "履约项合计", "校验",
        # 统计状态列 (55-67)
        "项目统计状态", "履约项统计状态（即，财报-交付/确收状态）",
        "1：正常交付", "2：应交未交", "3：交付异常", "4：正常验收",
        "5：应验未验", "6：验收异常", "7：正常服务", "8：应结未结", "9：已结项",
        "统计校验", "项目验收状态（即，财报-验收状态）",
        # 考核列 (68-76)
        "基线-预估结项日期（计算后反填）",
        "交付计划准确率“差异”", "交付计划准确率“提前/延后”",
        "交付计划准确率“是否跨月”", "交付计划准确率-考核扣分",
        "按时交付率“差异”", "按时交付率“提前/延后”",
        "按时交付率“是否跨月”", "按时交付率-考核扣分",
        # 计算列 (77-81)
        "提前实施项目持续周期（天）",
        "提前实施项目持续周期-统计",
        "提前实施项目是否已关联合同",
        "关联合同归档日期",
        "统计所属项目",
        # 映射列 (82-84)
        "项目经理", "项目经理所属部门", "销售团队-统计",
        # POC 工时
        "POC项目工时合计（小时）",
    ]

    # 过滤存在的列（有些列可能因为数据缺失没有）
    actual_cols = [c for c in expected_cols if c in df.columns]
    df = df[actual_cols]

    return df


def build_exception_df(df_exc_raw: pd.DataFrame, df_sign: pd.DataFrame) -> pd.DataFrame:
    """
    构建异常项目完整 DataFrame（38 列）
    
    列顺序对齐参考表异常项目 Sheet：
      1-36: 异常项目基础信息 + 异常详情（从完整版 CSV 取）
      37: 项目经理团队（从签约表关联）
      38: 项目验收状态（从签约表关联）
    """
    df = df_exc_raw.copy()
    
    # 过滤：异常报备日期 <= 报告截止日（异常项目没有立项日期，用报备日期）
    if "异常报备日期" in df.columns:
        df["_报备_date"] = pd.to_datetime(df["异常报备日期"].astype(str).str[:10], errors="coerce")
        df = df[df["_报备_date"] <= pd.to_datetime(REPORT_DATE)]
        df.drop(columns=["_报备_date"], inplace=True)
    
    # 从签约表关联项目经理团队、项目验收状态
    if "销售合同编号" in df.columns and "销售合同编号" in df_sign.columns:
        # 项目经理团队 = 签约表的项目经理所属部门
        dept_lookup = df_sign.set_index("销售合同编号")["项目经理所属部门"].to_dict()
        df["项目经理团队"] = df["销售合同编号"].map(
            lambda x: dept_lookup.get(x, "") if pd.notna(x) else ""
        )
        # 项目验收状态 = 签约表的项目验收状态（财报-验收状态）
        acc_status_col = "项目验收状态（即，财报-验收状态）"
        if acc_status_col in df_sign.columns:
            acc_lookup = df_sign.set_index("销售合同编号")[acc_status_col].to_dict()
            df["项目验收状态"] = df["销售合同编号"].map(
                lambda x: acc_lookup.get(x, "") if pd.notna(x) else ""
            )
    
    # 确保列顺序正确（对齐参考表异常项目 Sheet，38 列）
    expected_cols = [
        # 列 1-16：合同+项目基础信息
        "销售合同编号", "合同归档日期", "最终用户名称", "客户名称",
        "责任销售（履约项）", "责任销售所属团队", "负责人", "所属项目",
        "项目类型(概览)", "合同起始日期", "合同结束日期",
        "交付服务开始日期", "交付服务结束日期", "合同验收条款",
        "验收时点", "验收方式",
        # 列 17-21：异常项标识
        "标题", "状态", "备注", "PMO备注", "ID",
        # 列 22-36：异常详情
        "事业部（区域）", "事业部负责人", "异常报备日期",
        "预估异常处置完成日期", "异常归档日期", "异常影响情况",
        "异常项目-类别", "异常项目-处置方案", "异常处置方案-影响",
        "交付说明（异常履约项统计类别）",
        "交付说明（履约项交付情况、合同交付条款）",
        "交付中心反馈", "营销中心反馈", "项目异常内容", "预估金额",
        # 列 37-38：关联签约表
        "项目经理团队", "项目验收状态",
    ]
    
    # 过滤存在的列
    actual_cols = [c for c in expected_cols if c in df.columns]
    missing = [c for c in expected_cols if c not in df.columns]
    if missing:
        print(f"  ⚠️  异常项目缺列: {missing}")
    df = df[actual_cols]
    
    return df


def generate_delivery_report(period: str = "202606") -> Path:
    """生成完整交付月报"""
    global REPORT_DATE, REPORT_MONTH
    report_month = period
    REPORT_MONTH = report_month
    
    # 计算报告月份最后一天作为截止日期
    year = int(report_month[:4])
    month = int(report_month[4:])
    if month == 12:
        next_month = pd.Timestamp(year=year+1, month=1, day=1)
    else:
        next_month = pd.Timestamp(year=year, month=month+1, day=1)
    REPORT_DATE = (next_month - pd.Timedelta(days=1)).strftime("%Y-%m-%d")
    
    print(f"\n🚀 开始生成交付月报: {report_month} (截止 {REPORT_DATE})")
    print("=" * 60)

    # 1. 加载原始数据
    print("\n[1/5] 加载原始数据...")
    df_sign = load_ones_sign_contracts(report_month)
    df_poc = load_ones_poc(report_month)
    df_exc = load_ones_exceptions(report_month)
    df_rev = load_handover_revenue(report_month)
    df_acc = load_handover_acceptance(report_month)

    # 2. 创建工作簿
    print("\n[2/5] 创建工作簿...")
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认 sheet

    # 3. 生成明细 Sheet（带完整计算列）
    print("\n[3/5] 生成明细 Sheet (5个)...")

    # --- 签约（83 列）---
    print("  📝 签约: 构建完整 83 列...")
    df_sign_full = build_sign_sheet_df(df_sign, df_exc)
    ws_sign = wb.create_sheet("签约")
    ws_sign.cell(row=1, column=1, value=REPORT_DATE)
    _write_df_to_sheet(ws_sign, df_sign_full, sheet_name="签约", start_row=2)
    print(f"  ✅ 签约: {len(df_sign_full)} 行 × {len(df_sign_full.columns)} 列")

    # --- POC&提前实施（84 列）---
    print("  📝 POC&提前实施: 构建完整 84 列...")
    df_poc_full = build_poc_sheet_df(df_poc, df_exc)
    ws_poc = wb.create_sheet("POC&提前实施")
    ws_poc.cell(row=1, column=1, value=REPORT_DATE)
    _write_df_to_sheet(ws_poc, df_poc_full, sheet_name="POC&提前实施", start_row=2)
    print(f"  ✅ POC&提前实施: {len(df_poc_full)} 行 × {len(df_poc_full.columns)} 列")

    # --- 异常项目 ---
    print("  📝 异常项目: 构建完整 38 列...")
    df_exc_full = build_exception_df(df_exc, df_sign_full)
    ws_exc = wb.create_sheet("异常项目")
    _write_df_to_sheet(ws_exc, df_exc_full, sheet_name="异常项目", start_row=1)
    print(f"  ✅ 异常项目: {len(df_exc_full)} 行 × {len(df_exc_full.columns)} 列")

    # --- 确收交接 ---
    print("  📝 确收交接: 转换 CSV → 23 列...")
    df_rev_full = build_revenue_handover_df(df_rev, period)
    ws_rev = wb.create_sheet("确收交接")
    _write_df_to_sheet(ws_rev, df_rev_full, sheet_name="确收交接", start_row=1)
    print(f"  ✅ 确收交接: {len(df_rev_full)} 行 × {len(df_rev_full.columns)} 列")

    # --- 验收交接 ---
    print("  📝 验收交接: 转换 CSV → 27 列...")
    df_acc_full = build_acceptance_handover_df(df_acc, period)
    ws_acc = wb.create_sheet("验收交接")
    _write_df_to_sheet(ws_acc, df_acc_full, sheet_name="验收交接", start_row=1)
    # 第22列（索引从1开始）列名原来是 "截至目前全部/部分验收.1"，要改成和第16列一样的 "截至目前全部/部分验收"
    # 参考表里两列同名
    ws_acc.cell(row=1, column=22, value="截至目前全部/部分验收")
    print(f"  ✅ 验收交接: {len(df_acc_full)} 行 × {len(df_acc_full.columns)} 列")

    # 4. 统计 Sheet 九个
    print("\n[4/5] 生成统计 Sheet (9个)...")
    
    # 九个统计 Sheet，按参考报表顺序创建
    # 参考报表 Sheet 顺序：
    # 1.交付效率统计  2.签约统计  3.产品-授权&维保统计  4.签约  5.POC&提前实施统计
    # 6.提前实施分事业部统计  7.POC&提前实施  8.异常统计  9.异常台账  10.交付异常分事业部统计
    # 11.异常项目  12.交接统计  13.确收交接  14.验收交接  15.图例
    stat_builders = [
        ("交付效率统计", build_sign_stats),
        ("签约统计", build_sign_stats),
        ("产品-授权&维保统计", build_product_stats),
        ("签约", None),  # 明细已经生成，跳过
        ("POC&提前实施统计", build_poc_stats),
        ("提前实施分事业部统计", build_poc_dept_stats),
        ("POC&提前实施", None),
        ("异常统计", build_abnormal_stats),
        ("异常台账", build_abnormal_ledger),
        ("交付异常分事业部统计", build_abnormal_dept_stats),
        ("异常项目", None),
        ("交接统计", build_handover_stats),
        ("确收交接", None),
        ("验收交接", None),
        ("图例", build_legend),
    ]
    
    for sheet_name, builder in stat_builders:
        if builder is None:
            continue  # 跳过已经生成的明细
        print(f"  📝 生成统计 Sheet: {sheet_name}")
        ws = wb.create_sheet(sheet_name)
        builder(ws)
        print(f"  ✅ {sheet_name}: 已生成")

    # 重排 Sheet 顺序，对齐参考报表
    ref_order = [
        "交付效率统计",
        "签约统计",
        "产品-授权&维保统计",
        "签约",
        "POC&提前实施统计",
        "提前实施分事业部统计",
        "POC&提前实施",
        "异常统计",
        "异常台账",
        "交付异常分事业部统计",
        "异常项目",
        "交接统计",
        "确收交接",
        "验收交接",
        "图例",
    ]
    wb._sheets.sort(key=lambda ws: ref_order.index(ws.title))

    # 5. 保存
    print("\n[5/5] 保存文件...")
    REPORT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = REPORT_OUTPUT_DIR / f"交付月报-{period}-v2.xlsx"
    wb.save(output_path)
    print(f"✅ 已保存到: {output_path}")

    return output_path


if __name__ == "__main__":
    period = sys.argv[1] if len(sys.argv) > 1 else "202606"
    generate_delivery_report(period)
