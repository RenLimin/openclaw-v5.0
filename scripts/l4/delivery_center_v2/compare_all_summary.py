#!/usr/bin/env python3
import pandas as pd
import openpyxl
from pathlib import Path

# 参考和生成路径
REF_PATH = Path("/Users/bangcle/Bangcle Workspace/01. Management/2026/2026团队报告/202606/2026交付月报-20260630.xlsx")
GEN_PATH = Path("/Users/bangcle/.openclaw/data/reports/交付月报-202606-v2.xlsx")

def compare_sheet_summary(sheet_name):
    # 读取参考和生成
    wb_ref = openpyxl.load_workbook(REF_PATH, read_only=True)
    wb_gen = openpyxl.load_workbook(GEN_PATH, read_only=True)
    
    if sheet_name not in wb_ref.sheetnames or sheet_name not in wb_gen.sheetnames:
        return {
            "sheet": sheet_name,
            "status": "missing",
            "ref_rows": 0,
            "ref_cols": 0,
            "gen_rows": 0,
            "gen_cols": 0,
            "row_diff": 0,
            "col_diff": 0,
        }
    
    ws_ref = wb_ref[sheet_name]
    ws_gen = wb_gen[sheet_name]
    
    ref_rows = ws_ref.max_row
    ref_cols = ws_ref.max_column
    gen_rows = ws_gen.max_row
    gen_cols = ws_gen.max_column
    
    row_diff = gen_rows - ref_rows
    col_diff = gen_cols - ref_cols
    
    wb_ref.close()
    wb_gen.close()
    
    return {
        "sheet": sheet_name,
        "ref_rows": ref_rows,
        "ref_cols": ref_cols,
        "gen_rows": gen_rows,
        "gen_cols": gen_cols,
        "row_diff": row_diff,
        "col_diff": col_diff,
    }

def main():
    sheets = [
        "交付效率统计",
        "签约统计",
        "产品-授权&维保统计",
        "签约",
        "POC&提前实施统计",
        "提前实施分事业部统计",
        "POC&提前实施",
        "异常统计",
        "异常台账",
        "交付异常分事业部统计",
        "异常项目",
        "交接统计",
        "确收交接",
        "验收交接",
        "图例",
    ]
    
    results = []
    for sheet in sheets:
        res = compare_sheet_summary(sheet)
        results.append(res)
    
    # 打印表格
    print(f"| Sheet | 参考行数 | 生成行数 | 行数差 | 参考列数 | 生成列数 | 列数差 |")
    print(f"|-------|----------|----------|--------|----------|----------|--------|")
    for r in results:
        print(f"| {r['sheet']} | {r['ref_rows']} | {r['gen_rows']} | {r['row_diff']:+} | {r['ref_cols']} | {r['gen_cols']} | {r['col_diff']:+} |")

if __name__ == "__main__":
    main()
