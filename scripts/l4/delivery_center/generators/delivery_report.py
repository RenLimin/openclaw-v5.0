"""交付月报生成器

生成 12 Sheet 的交付月报 Excel 文件。
基于 openpyxl 库，支持公式、样式、多 Sheet。

Sheet 列表（基于 2026 交付中心报告模版 v1.0）：
  1. 签约 (79列)
  2. POC&提前实施 (84列)
  3. 异常项目 (37列)
  4. 确收交接 (17列)
  5. 验收交接 (14列)
  6. 交付效率统计 (18列)
  7. 签约统计 (15列)
  8. POC&提前实施统计 (28列)
  9. 异常统计 (30列)
  10. 异常台账 (17列)
  11. 交接统计 (21列)
  12. 图例 (29列)

已验证 2026-09-01。
"""

import pandas as pd
from pathlib import Path
from typing import Optional
from datetime import datetime

OUTPUT_DIR = Path.home() / ".openclaw" / "data" / "reports"


def _write_df_to_sheet(ws, df: pd.DataFrame, start_row: int = 1):
    """将 DataFrame 写入 openpyxl Worksheet"""
    if df.empty:
        ws.cell(row=1, column=1, value="无数据")
        return

    # 写入表头
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=col_idx, value=str(col_name))

    # 写入数据
    for row_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            if pd.isna(value):
                ws.cell(row=row_idx, column=col_idx, value="")
            elif isinstance(value, (datetime, pd.Timestamp)):
                ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)


def _fill_contract_sheet(ws, df: pd.DataFrame):
    """填充签约 Sheet"""
    _write_df_to_sheet(ws, df)


def _fill_poc_sheet(ws, df: pd.DataFrame):
    """填充 POC&提前实施 Sheet"""
    _write_df_to_sheet(ws, df)


def _fill_exception_sheet(ws, df: pd.DataFrame):
    """填充异常项目 Sheet"""
    _write_df_to_sheet(ws, df)


def _fill_revenue_sheet(ws, df: pd.DataFrame):
    """填充确收交接 Sheet"""
    _write_df_to_sheet(ws, df)


def _fill_acceptance_sheet(ws, df: pd.DataFrame):
    """填充验收交接 Sheet"""
    _write_df_to_sheet(ws, df)


def _fill_efficiency_sheet(ws, df: pd.DataFrame):
    """填充交付效率统计 Sheet"""
    if df.empty:
        ws.cell(row=1, column=1, value="无数据")
        return

    # 表头
    headers = ["部门", "项目经理", "交付计划扣分", "按时交付扣分", "总扣分", "项目数"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    # 数据
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=str(row.get("部门", "")))
        ws.cell(row=row_idx, column=2, value=str(row.get("项目经理", "")))
        ws.cell(row=row_idx, column=3, value=float(row.get("交付计划扣分", 0)))
        ws.cell(row=row_idx, column=4, value=float(row.get("按时交付扣分", 0)))
        ws.cell(row=row_idx, column=5, value=float(row.get("总扣分", 0)))
        ws.cell(row=row_idx, column=6, value=int(row.get("项目数", 0)))


def _fill_contract_stats_sheet(ws, df: pd.DataFrame):
    """填充签约统计 Sheet（按部门统计）

    支持两种数据源：
    1. ONES 数据：项目经理所属部门 + 项目编号
    2. OA 合同数据：项目经理 + htbh（合同编号）+ legend_pm_dept 映射
    """
    if df.empty:
        ws.cell(row=1, column=1, value="无数据")
        return

    from pathlib import Path
    import json

    config_dir = Path(__file__).parent.parent / "config"

    # ONES 数据源
    if "项目经理所属部门" in df.columns and "项目编号" in df.columns:
        stats = df.groupby("项目经理所属部门").agg({
            "项目编号": "nunique",
        }).reset_index()
        stats.columns = ["部门", "项目数"]
        _write_df_to_sheet(ws, stats)
        return

    # OA 合同数据源：用项目经理→部门映射
    pm_col = None
    for candidate in ["项目经理", "项目管理部负责人"]:
        if candidate in df.columns and df[candidate].notna().sum() > 10:
            pm_col = candidate
            break
    if pm_col is None:
        for candidate in ["项目经理", "项目管理部负责人"]:
            if candidate in df.columns:
                pm_col = candidate
                break

    if not pm_col or "htbh" not in df.columns:
        ws.cell(row=1, column=1, value="无部门数据")
        return

    # 加载项目经理→部门映射
    legend_path = config_dir / "legend_pm_dept.json"
    if legend_path.exists():
        legend = json.loads(legend_path.read_text(encoding="utf-8"))
    else:
        legend = {}

    # 映射项目经理到部门
    df["_dept"] = df[pm_col].map(lambda x: legend.get(str(x).strip(), "未知") if pd.notna(x) else "未知")

    stats = df.groupby("_dept").agg({
        "htbh": "nunique",
    }).reset_index()
    stats.columns = ["部门", "项目数"]
    stats = stats.sort_values("项目数", ascending=False)
    _write_df_to_sheet(ws, stats)


def _fill_poc_stats_sheet(ws, df: pd.DataFrame):
    """填充 POC 统计 Sheet"""
    _write_df_to_sheet(ws, df)


def _fill_exception_stats_sheet(ws, df: pd.DataFrame):
    """填充异常统计 Sheet"""
    _write_df_to_sheet(ws, df)


def _fill_exception_ledger_sheet(ws, df: pd.DataFrame):
    """填充异常台账 Sheet"""
    _write_df_to_sheet(ws, df)


def _fill_handover_stats_sheet(ws, df: pd.DataFrame):
    """填充交接统计 Sheet（按部门汇总确收/验收）"""
    if df.empty:
        ws.cell(row=1, column=1, value="无数据")
        return

    # 按部门汇总
    dept_col = None
    for candidate in ["部门", "销售部门", "所属分部"]:
        if candidate in df.columns:
            dept_col = candidate
            break

    if dept_col:
        stats = df.groupby(dept_col).size().reset_index(name="数量")
        stats.columns = ["部门", "数量"]
        _write_df_to_sheet(ws, stats)
    else:
        _write_df_to_sheet(ws, df)


def _fill_legend_sheet(ws):
    """填充图例 Sheet"""
    import json
    config_dir = Path(__file__).parent.parent / "config"

    # 项目经理-部门映射
    legend_pm = json.loads((config_dir / "legend_pm_dept.json").read_text(encoding="utf-8"))
    ws.cell(row=1, column=1, value="项目经理")
    ws.cell(row=1, column=2, value="部门")
    for row_idx, (pm, dept) in enumerate(legend_pm.items(), 2):
        ws.cell(row=row_idx, column=1, value=pm)
        ws.cell(row=row_idx, column=2, value=dept)


def generate_delivery_report(
    month: str,
    contract_df: pd.DataFrame = None,
    poc_df: pd.DataFrame = None,
    exception_df: pd.DataFrame = None,
    revenue_df: pd.DataFrame = None,
    acceptance_df: pd.DataFrame = None,
    efficiency_df: pd.DataFrame = None,
    output_dir: Optional[str] = None,
) -> str:
    """生成交付月报

    Args:
        month: 报告月份（YYYYMM）
        contract_df: 签约项目数据
        poc_df: POC 项目数据
        exception_df: 异常项目数据
        revenue_df: 确收凭证数据
        acceptance_df: 验收凭证数据
        efficiency_df: 交付效率统计数据
        output_dir: 输出目录

    Returns:
        生成的 Excel 文件路径
    """
    from openpyxl import Workbook

    out_dir = Path(output_dir) if output_dir else OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 签约
    ws1 = wb.create_sheet("签约")
    _fill_contract_sheet(ws1, contract_df if contract_df is not None else pd.DataFrame())

    # Sheet 2: POC&提前实施
    ws2 = wb.create_sheet("POC&提前实施")
    _fill_poc_sheet(ws2, poc_df if poc_df is not None else pd.DataFrame())

    # Sheet 3: 异常项目
    ws3 = wb.create_sheet("异常项目")
    _fill_exception_sheet(ws3, exception_df if exception_df is not None else pd.DataFrame())

    # Sheet 4: 确收交接
    ws4 = wb.create_sheet("确收交接")
    _fill_revenue_sheet(ws4, revenue_df if revenue_df is not None else pd.DataFrame())

    # Sheet 5: 验收交接
    ws5 = wb.create_sheet("验收交接")
    _fill_acceptance_sheet(ws5, acceptance_df if acceptance_df is not None else pd.DataFrame())

    # Sheet 6: 交付效率统计
    ws6 = wb.create_sheet("交付效率统计")
    _fill_efficiency_sheet(ws6, efficiency_df if efficiency_df is not None else pd.DataFrame())

    # Sheet 7: 签约统计
    ws7 = wb.create_sheet("签约统计")
    _fill_contract_stats_sheet(ws7, contract_df if contract_df is not None else pd.DataFrame())

    # Sheet 8: POC&提前实施统计
    ws8 = wb.create_sheet("POC&提前实施统计")
    _fill_poc_stats_sheet(ws8, poc_df if poc_df is not None else pd.DataFrame())

    # Sheet 9: 异常统计
    ws9 = wb.create_sheet("异常统计")
    _fill_exception_stats_sheet(ws9, exception_df if exception_df is not None else pd.DataFrame())

    # Sheet 10: 异常台账
    ws10 = wb.create_sheet("异常台账")
    _fill_exception_ledger_sheet(ws10, exception_df if exception_df is not None else pd.DataFrame())

    # Sheet 11: 交接统计
    ws11 = wb.create_sheet("交接统计")
    _fill_handover_stats_sheet(ws11, revenue_df if revenue_df is not None else pd.DataFrame())

    # Sheet 12: 图例
    ws12 = wb.create_sheet("图例")
    _fill_legend_sheet(ws12)

    # 保存
    output_path = out_dir / f"交付月报-{month}.xlsx"
    wb.save(output_path)
    print(f"✅ 交付月报已生成: {output_path}")
    return str(output_path)


if __name__ == "__main__":
    print("=== 交付月报生成器测试 ===\n")

    # 用测试数据生成报告
    test_revenue = pd.DataFrame([
        {"合同编号": "XSZS001", "客户名称": "杭州银行", "项目经理": "梁杨", "交接日期": "2026-06-09"},
        {"合同编号": "XSZS002", "客户名称": "恒丰银行", "项目经理": "孙康", "交接日期": "2026-06-12"},
    ])

    test_acceptance = pd.DataFrame([
        {"合同编号": "XSZS003", "客户名称": "国鸿信息", "项目经理": "孙康", "交接日期": "2026-06-12"},
    ])

    test_efficiency = pd.DataFrame([
        {"部门": "北区", "项目经理": "张三", "交付计划扣分": 1.0, "按时交付扣分": 0.5, "总扣分": 1.5, "项目数": 3},
        {"部门": "南区", "项目经理": "李四", "交付计划扣分": 0.0, "按时交付扣分": 0.5, "总扣分": 0.5, "项目数": 2},
    ])

    output = generate_delivery_report(
        month="202606",
        revenue_df=test_revenue,
        acceptance_df=test_acceptance,
        efficiency_df=test_efficiency,
    )
    print(f"\n输出: {output}")
