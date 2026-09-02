"""
202606 交付月报统计计算 → 存储到 BDMS → 生成 Excel
按照既定方案：通过公式计算后存储最终结果至本地数据库
"""
import sqlite3, csv, json
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# === 配置 ===
ONES_DIR = Path.home() / ".openclaw" / "data" / "ones_exports"
BDMS_DB = Path.home() / ".openclaw" / "data" / "bdms.db"
OUTPUT_DIR = Path.home() / ".openclaw" / "data" / "reports"
CONFIG_DIR = Path(__file__).parent.parent / "config"

REPORT_MONTH = "202606"
REPORT_DATE = "2026-06-30"

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=10, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def load_csv(path):
    with open(path, 'r', encoding='utf-8-sig') as f:
        return pd.DataFrame(csv.DictReader(f))


def load_bdms_table(table):
    conn = sqlite3.connect(BDMS_DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(f"SELECT * FROM {table}")
    cols = [d[0] for d in c.description]
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return cols, rows


def init_stats_table(conn):
    """创建统计结果表"""
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS report_statistics (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        report_month TEXT NOT NULL,
        sheet_name TEXT NOT NULL,
        stat_type TEXT,
        row_key TEXT,
        col_key TEXT,
        value_num REAL,
        value_text TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(report_month, sheet_name, stat_type, row_key, col_key)
    )''')
    conn.commit()


def store_stat(conn, month, sheet, stat_type, row_key, col_key, value):
    """存储统计结果"""
    c = conn.cursor()
    if isinstance(value, (int, float)) and not pd.isna(value):
        val_num = float(value)
        val_text = None
    else:
        val_num = None
        val_text = str(value) if value else None
    
    c.execute('''INSERT OR REPLACE INTO report_statistics 
        (report_month, sheet_name, stat_type, row_key, col_key, value_num, value_text)
        VALUES (?, ?, ?, ?, ?, ?, ?)''',
        (month, sheet, stat_type, row_key, col_key, val_num, val_text))


def compute_all_stats(conn, sign_df, poc_df, abnormal_df, rev_rows, acc_rows):
    """计算所有统计指标并存储"""
    print("  计算统计指标...")
    
    # === 1. 签约统计 ===
    sign_df['立项年份'] = pd.to_datetime(sign_df['立项日期'], errors='coerce').dt.year
    
    # 1a. 按签约年份统计
    sign_by_year = sign_df.groupby('立项年份')['ID'].nunique()
    for year, count in sign_by_year.items():
        if pd.notna(year):
            store_stat(conn, REPORT_MONTH, '签约统计', 'by_year', str(int(year)), '项目数', count)
    
    # 1b. 按项目状态×年份交叉表
    status_year = sign_df.groupby(['项目状态', '立项年份'])['ID'].nunique().reset_index()
    for _, row in status_year.iterrows():
        year = int(row['立项年份']) if pd.notna(row['立项年份']) else 0
        store_stat(conn, REPORT_MONTH, '签约统计', 'status_year', 
                   str(row['项目状态']), str(year), row['ID'])
    
    # 1c. 按部门统计
    if '责任销售所属团队' in sign_df.columns:
        dept_stats = sign_df.groupby('责任销售所属团队')['ID'].nunique()
        for dept, count in dept_stats.items():
            store_stat(conn, REPORT_MONTH, '签约统计', 'by_dept', str(dept), '项目数', count)
    
    print("    ✅ 签约统计")
    
    # === 2. POC&提前实施统计 ===
    poc_df['立项年份'] = pd.to_datetime(poc_df['立项日期'], errors='coerce').dt.year
    
    # 2a. 按年份×项目类型统计
    poc_by_year_type = poc_df.groupby(['立项年份', '项目类型(概览)'])['ID'].nunique().reset_index()
    for _, row in poc_by_year_type.iterrows():
        year = int(row['立项年份']) if pd.notna(row['立项年份']) else 0
        store_stat(conn, REPORT_MONTH, 'POC&提前实施统计', 'year_type',
                   str(year), str(row['项目类型(概览)']), row['ID'])
    
    # 2b. 按部门统计
    if '责任销售所属团队' in poc_df.columns:
        poc_dept = poc_df.groupby('责任销售所属团队')['ID'].nunique()
        for dept, count in poc_dept.items():
            store_stat(conn, REPORT_MONTH, 'POC&提前实施统计', 'by_dept', str(dept), '项目数', count)
    
    print("    ✅ POC&提前实施统计")
    
    # === 3. 异常统计 ===
    abnormal_df['归档年份'] = pd.to_datetime(abnormal_df['合同归档日期'], errors='coerce').dt.year
    
    # 3a. 按归档年份统计
    abn_by_year = abnormal_df.groupby('归档年份')['ID'].nunique()
    for year, count in abn_by_year.items():
        if pd.notna(year):
            store_stat(conn, REPORT_MONTH, '异常统计', 'by_year', str(int(year)), '异常数', count)
    
    # 3b. 按部门统计
    if '责任销售所属团队' in abnormal_df.columns:
        abn_dept = abnormal_df.groupby('责任销售所属团队')['ID'].nunique()
        for dept, count in abn_dept.items():
            store_stat(conn, REPORT_MONTH, '异常统计', 'by_dept', str(dept), '异常数', count)
    
    print("    ✅ 异常统计")
    
    # === 4. 异常台账 ===
    for year in sorted(abnormal_df['归档年份'].dropna().astype(int).unique()):
        subset = abnormal_df[abnormal_df['归档年份'] == year]
        store_stat(conn, REPORT_MONTH, '异常台账', 'by_year', str(year), '合计', len(subset))
    
    print("    ✅ 异常台账")
    
    # === 5. 交接统计 ===
    rev_df = pd.DataFrame(rev_rows)
    acc_df = pd.DataFrame(acc_rows)
    
    store_stat(conn, REPORT_MONTH, '交接统计', 'summary', '确收总数', '值', len(rev_df))
    store_stat(conn, REPORT_MONTH, '交接统计', 'summary', '验收总数', '值', len(acc_df))
    
    if '是否接收' in rev_df.columns and len(rev_df) > 0:
        qualified = rev_df[rev_df['是否接收'].astype(str).str.contains('是', na=False)]
        store_stat(conn, REPORT_MONTH, '交接统计', '确收', '合格数', '值', len(qualified))
        rate = len(qualified) / len(rev_df) if len(rev_df) > 0 else 0
        store_stat(conn, REPORT_MONTH, '交接统计', '确收', '合格率', '值', rate)
    
    if '财务是否接收' in acc_df.columns and len(acc_df) > 0:
        qualified = acc_df[acc_df['财务是否接收'].astype(str).str.contains('是', na=False)]
        store_stat(conn, REPORT_MONTH, '交接统计', '验收', '合格数', '值', len(qualified))
        rate = len(qualified) / len(acc_df) if len(acc_df) > 0 else 0
        store_stat(conn, REPORT_MONTH, '交接统计', '验收', '合格率', '值', rate)
    
    print("    ✅ 交接统计")
    
    # === 6. 产品-授权&维保统计 ===
    if '标准产品/服务序号' in sign_df.columns:
        product_stats = sign_df.groupby('标准产品/服务序号')['ID'].nunique()
        for prod, count in product_stats.items():
            if prod and str(prod).strip():
                store_stat(conn, REPORT_MONTH, '产品-授权&维保统计', 'by_product', str(prod), '项目数', count)
    
    print("    ✅ 产品-授权&维保统计")
    
    # === 7. 提前实施分事业部统计 ===
    if '项目类型(概览)' in poc_df.columns:
        early_df = poc_df[poc_df['项目类型(概览)'].str.contains('提前实施', na=False)]
        if '责任销售所属团队' in early_df.columns:
            early_dept = early_df.groupby('责任销售所属团队')['ID'].nunique()
            for dept, count in early_dept.items():
                store_stat(conn, REPORT_MONTH, '提前实施分事业部统计', 'by_dept', str(dept), '项目数', count)
    
    print("    ✅ 提前实施分事业部统计")
    
    # === 8. 交付异常分事业部统计 ===
    if '责任销售所属团队' in abnormal_df.columns:
        abn_dept2 = abnormal_df.groupby('责任销售所属团队')['ID'].nunique()
        for dept, count in abn_dept2.items():
            store_stat(conn, REPORT_MONTH, '交付异常分事业部统计', 'by_dept', str(dept), '异常数', count)
    
    print("    ✅ 交付异常分事业部统计")
    
    # === 9. 交付效率统计 ===
    # 按项目经理统计项目数
    if '负责人' in sign_df.columns:
        pm_stats = sign_df.groupby('负责人')['ID'].nunique()
        for pm, count in pm_stats.items():
            store_stat(conn, REPORT_MONTH, '交付效率统计', 'by_pm', str(pm), '项目数', count)
    
    # 按部门统计
    if '责任销售所属团队' in sign_df.columns:
        dept_stats2 = sign_df.groupby('责任销售所属团队')['ID'].nunique()
        for dept, count in dept_stats2.items():
            store_stat(conn, REPORT_MONTH, '交付效率统计', 'by_dept', str(dept), '项目数', count)
    
    print("    ✅ 交付效率统计")
    
    conn.commit()
    print("  ✅ 所有统计已存储到 BDMS")


def read_stats(conn, month, sheet_name, stat_type):
    """从 BDMS 读取统计结果"""
    c = conn.cursor()
    c.execute('''SELECT row_key, col_key, value_num, value_text 
        FROM report_statistics 
        WHERE report_month=? AND sheet_name=? AND stat_type=?
        ORDER BY row_key, col_key''',
        (month, sheet_name, stat_type))
    return c.fetchall()


def write_df(ws, df, start_row=1):
    """写入 DataFrame 到 worksheet"""
    if df.empty:
        ws.cell(row=1, column=1, value="无数据")
        return
    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=str(h))
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    for ri, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for ci, val in enumerate(row, 1):
            v = "" if pd.isna(val) else val
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = THIN_BORDER


def write_df_with_summary(ws, df, date_str=REPORT_DATE):
    """写入 DataFrame 到 worksheet，首行加汇总信息（日期+排序占位）
    
    参考报表结构：
    - 行0: 日期(0) + 排序(40) + 统计数字(52+)
    - 行1: 列名
    - 行2+: 数据
    """
    if df.empty:
        ws.cell(row=1, column=1, value="无数据")
        return
    
    # 行0: 汇总行 - 日期在第1列，"排序"在第41列
    ws.cell(row=1, column=1, value=date_str)
    # 第41列（index 40）写"排序"
    sort_col = 41
    if sort_col <= len(df.columns):
        ws.cell(row=1, column=sort_col, value="排序")
    
    # 行2+: 列名 + 数据
    write_df(ws, df, start_row=2)


def reorder_abnormal_columns(abnormal_df):
    """重新排列异常项目 Sheet 的列顺序，匹配参考报表
    
    参考报表前几列顺序：
    销售合同编号(0)、合同归档日期(1)、最终用户名称(2)、客户名称(3)、
    责任销售（履约项）(4)、责任销售所属团队(5)、负责人(6)、所属项目(7)、
    项目类型(概览)(8)、合同起始日期(9)、合同结束日期(10)...
    
    ONES CSV 原始顺序从 BI履约ID 开始，销售合同编号在第12列。
    """
    # 参考报表的列顺序（与 ONES CSV 中存在的列对应）
    ref_order = [
        '销售合同编号',
        '合同归档日期',
        '最终用户名称',
        '客户名称',
        '责任销售（履约项）',
        '责任销售所属团队',
        '负责人',
        '所属项目',
        '项目类型(概览)',
        '合同开始日期',    # 参考叫"合同起始日期"，ONES 叫"合同开始日期"
        '合同结束日期',
        '交付服务开始日期',
        '交付服务结束日期',
        '合同验收条款',
        '验收时点',
        '验收方式',
        '标题',
        '状态',
        '备注',
        'PMO备注',
        'ID',
    ]
    
    # 获取 DataFrame 中实际存在的列
    existing_cols = list(abnormal_df.columns)
    
    # 按参考顺序排列存在的列
    ordered_cols = [c for c in ref_order if c in existing_cols]
    
    # 剩余列保持原始顺序，追加到后面
    remaining_cols = [c for c in existing_cols if c not in ordered_cols]
    final_cols = ordered_cols + remaining_cols
    
    # 重命名"合同开始日期"为"合同起始日期"以匹配参考报表
    result_df = abnormal_df[final_cols].copy()
    if '合同开始日期' in result_df.columns:
        result_df = result_df.rename(columns={'合同开始日期': '合同起始日期'})
    
    return result_df


def map_revenue_columns(rev_df):
    """映射确收交接 Sheet 列名，匹配参考报表
    
    BDMS 确收表 → 参考报表列名映射：
    - id → 不显示
    - voucher_id → ID
    - bi_id → BI履约ID
    - 合同编号 → 合同编号1
    - 合同名称 → 标题
    - 客户名称 → 客户名称
    - 销售部门 → 销售部门
    - 项目经理 → 项目经理
    - 交接日期 → 交接日期
    - 财务 → 财务接收人
    - 是否接收 → 财务是否接收
    
    参考报表完整列顺序：
    月份、标题、ID、BI履约ID、合同编号1、邮件编号、合同编号、客户名称、
    销售部门、项目经理、备注、交接日期、财务接收人、财务是否接收、
    财务反馈、交付邮件是否跨月、PMO提交人、PMO反馈、是否修改ONES状态、
    项目经理所属区域、跨月交接、跨月交接原因、是否合格
    """
    column_mapping = {
        'voucher_id': 'ID',
        'bi_id': 'BI履约ID',
        '合同编号': '合同编号1',
        '合同名称': '标题',
        '客户名称': '客户名称',
        '销售部门': '销售部门',
        '项目经理': '项目经理',
        '交接日期': '交接日期',
        '财务': '财务接收人',
        '是否接收': '财务是否接收',
    }
    
    # 参考报表列顺序（严格按照此顺序输出）
    ref_order = [
        '月份', '标题', 'ID', 'BI履约ID', '合同编号1', '邮件编号', '合同编号',
        '客户名称', '销售部门', '项目经理', '备注', '交接日期', '财务接收人',
        '财务是否接收', '财务反馈', '交付邮件是否跨月', 'PMO提交人', 'PMO反馈',
        '是否修改ONES状态', '项目经理所属区域', '跨月交接', '跨月交接原因', '是否合格'
    ]
    
    # 重命名原始数据列
    result_df = rev_df.copy()
    rename_map = {k: v for k, v in column_mapping.items() if k in result_df.columns}
    result_df = result_df.rename(columns=rename_map)
    
    # 按参考报表顺序构建列：有数据的用数据，没数据的填空列
    final_cols = []
    for col in ref_order:
        if col not in result_df.columns:
            result_df[col] = ''
        final_cols.append(col)
    
    result_df = result_df[final_cols]
    return result_df


def map_acceptance_columns(acc_df):
    """映射验收交接 Sheet 列名，匹配参考报表
    
    BDMS 验收表 → 参考报表列名映射：
    - id → 不显示
    - voucher_id → ID
    - bi_id → BI履约ID
    - 合同编号 → 合同编号1
    - 合同名称 → 合同名称
    - 客户名称 → 客户名称
    - 项目经理 → 项目经理
    - 验收单编号 → 验收单编号
    - 交接日期 → 交接日期
    - 验收方式 → 验收方式
    - 全部或部分 → 截至目前全部/部分验收（第1个）
    - 财务 → 财务接收人
    - 财务是否接收 → 是否接收
    
    参考报表完整列顺序（含重复列名）：
    月份、合同名称、标题、ID、BI履约ID、验收单编号-财务端、合同编号1、
    验收单编号、合同编号、客户名称、销售部门、项目经理、备注、交接日期、
    验收方式、截至目前全部/部分验收、是否为渠道、财务接收人、是否接收、
    实际验收方式、财务反馈、截至目前全部/部分验收、PMO提交人、PMO反馈、
    是否修改ones及OA状态、项目经理所属区域、是否合格
    """
    column_mapping = {
        'voucher_id': 'ID',
        'bi_id': 'BI履约ID',
        '合同编号': '合同编号1',
        '合同名称': '合同名称',
        '客户名称': '客户名称',
        '项目经理': '项目经理',
        '验收单编号': '验收单编号',
        '交接日期': '交接日期',
        '验收方式': '验收方式',
        '全部或部分': '截至目前全部/部分验收',
        '财务': '财务接收人',
        '财务是否接收': '是否接收',
    }
    
    # 参考报表列顺序（严格按照此顺序输出，包含重复列名）
    ref_order = [
        '月份', '合同名称', '标题', 'ID', 'BI履约ID', '验收单编号-财务端',
        '合同编号1', '验收单编号', '合同编号', '客户名称', '销售部门',
        '项目经理', '备注', '交接日期', '验收方式', '截至目前全部/部分验收',
        '是否为渠道', '财务接收人', '是否接收', '实际验收方式', '财务反馈',
        '截至目前全部/部分验收', 'PMO提交人', 'PMO反馈',
        '是否修改ones及OA状态', '项目经理所属区域', '是否合格'
    ]
    
    # 重命名原始数据列
    result_df = acc_df.copy()
    rename_map = {k: v for k, v in column_mapping.items() if k in result_df.columns}
    result_df = result_df.rename(columns=rename_map)
    
    # 按参考报表顺序构建：处理重复列名的情况
    # 第一个"截至目前全部/部分验收"用数据列，第二个留空
    seen_counts = {}
    final_data = {}
    for i, col in enumerate(ref_order):
        if col not in seen_counts:
            seen_counts[col] = 0
        seen_counts[col] += 1
        
        if col in result_df.columns and seen_counts[col] == 1:
            # 第一次出现且有数据，用实际数据
            final_data[i] = result_df[col].values
        else:
            # 重复出现或无数据，填空
            final_data[i] = [''] * len(result_df)
    
    # 构建新的 DataFrame，使用整数列索引避免重名冲突
    out_df = pd.DataFrame({i: final_data[i] for i in range(len(ref_order))})
    out_df.columns = ref_order
    return out_df


def generate_excel_from_stats(conn, sign_df, poc_df, abnormal_df, rev_rows, acc_rows):
    """从 BDMS 统计结果生成 Excel"""
    print("  生成 Excel...")
    wb = Workbook()
    wb.remove(wb.active)
    
    # Sheet 1: 签约（带汇总行 + 计算列）
    from compute_columns import compute_sign_columns, reorder_sign_columns, compute_poc_columns, reorder_poc_columns
    ws = wb.create_sheet("签约")
    sign_full = reorder_sign_columns(compute_sign_columns(sign_df))
    write_df_with_summary(ws, sign_full)
    
    # Sheet 2: POC&提前实施（带汇总行 + 计算列）
    ws = wb.create_sheet("POC&提前实施")
    poc_full = reorder_poc_columns(compute_poc_columns(poc_df))
    write_df_with_summary(ws, poc_full)
    
    # Sheet 3: 异常项目（重排列顺序）
    ws = wb.create_sheet("异常项目")
    abnormal_reordered = reorder_abnormal_columns(abnormal_df)
    write_df(ws, abnormal_reordered)
    
    # Sheet 4: 确收交接（中文列名映射）
    ws = wb.create_sheet("确收交接")
    rev_df = pd.DataFrame(rev_rows)
    rev_mapped = map_revenue_columns(rev_df)
    write_df(ws, rev_mapped)
    
    # Sheet 5: 验收交接（中文列名映射）
    ws = wb.create_sheet("验收交接")
    acc_df = pd.DataFrame(acc_rows)
    acc_mapped = map_acceptance_columns(acc_df)
    write_df(ws, acc_mapped)
    
    # Sheet 6-15: 从 BDMS 统计结果生成（透视表格式）
    from build_stat_sheets import build_all_stat_sheets
    build_all_stat_sheets(wb, conn)
    
    # 图例 Sheet
    ws = wb.create_sheet("图例")
    legend_path = CONFIG_DIR / "legend_pm_dept.json"
    if legend_path.exists():
        legend = json.loads(legend_path.read_text(encoding="utf-8"))
        ws.cell(row=1, column=1, value="项目经理")
        ws.cell(row=1, column=2, value="部门")
        ws.cell(row=1, column=1).font = HEADER_FONT_WHITE
        ws.cell(row=1, column=1).fill = HEADER_FILL
        ws.cell(row=1, column=2).font = HEADER_FONT_WHITE
        ws.cell(row=1, column=2).fill = HEADER_FILL
        for ri, (pm, dept) in enumerate(legend.items(), 2):
            ws.cell(row=ri, column=1, value=pm)
            ws.cell(row=ri, column=2, value=dept)
    
    # 保存
    output_path = OUTPUT_DIR / f"交付月报-{REPORT_MONTH}.xlsx"
    wb.save(output_path)
    print(f"  ✅ Excel 已生成: {output_path}")
    return str(output_path)


def main():
    print("=== 202606 交付月报：统计计算 → 入库 → 生成 ===\n")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # 1. 加载数据
    print("1. 加载数据...")
    sign_df = load_csv(ONES_DIR / "签约项目统计.csv")
    poc_df = load_csv(ONES_DIR / "poc_提前实施.csv")
    abnormal_df = load_csv(ONES_DIR / "异常处置.csv")
    rev_cols, rev_rows = load_bdms_table("revenue_vouchers")
    acc_cols, acc_rows = load_bdms_table("acceptance_vouchers")
    print(f"   签约: {len(sign_df)}, POC: {len(poc_df)}, 异常: {len(abnormal_df)}")
    print(f"   确收: {len(rev_rows)}, 验收: {len(acc_rows)}")
    
    # 2. 初始化统计表
    print("\n2. 初始化统计表...")
    conn = sqlite3.connect(BDMS_DB)
    init_stats_table(conn)
    
    # 3. 清空旧统计（幂等）
    c = conn.cursor()
    c.execute("DELETE FROM report_statistics WHERE report_month=?", (REPORT_MONTH,))
    conn.commit()
    
    # 4. 计算并存储统计
    print("\n3. 计算统计指标并存储...")
    compute_all_stats(conn, sign_df, poc_df, abnormal_df, rev_rows, acc_rows)
    
    # 5. 验证存储
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM report_statistics WHERE report_month=?", (REPORT_MONTH,))
    stat_count = c.fetchone()[0]
    print(f"\n4. 验证: {stat_count} 条统计记录已存储")
    
    c.execute("SELECT sheet_name, COUNT(*) FROM report_statistics WHERE report_month=? GROUP BY sheet_name", 
              (REPORT_MONTH,))
    for row in c.fetchall():
        print(f"   {row[0]}: {row[1]} 条")
    
    # 6. 生成 Excel
    print("\n5. 生成 Excel...")
    output = generate_excel_from_stats(conn, sign_df, poc_df, abnormal_df, rev_rows, acc_rows)
    
    conn.close()
    print(f"\n✅ 完成: {output}")


if __name__ == "__main__":
    main()
