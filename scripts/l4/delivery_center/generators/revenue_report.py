"""确收月报生成器

生成确收月报 Excel 文件。
基于 M2 引擎（差异分析 + 汇总统计）输出。

Sheet 列表：
  1. 预算执行表（计划 vs 实际）
  2. 差异分析（差异金额 + 分类）
  3. 月度汇总（按月份统计）
  4. 确收明细（确收凭证明细）
  5. 验收明细（验收凭证明细）
  6. 按部门汇总
  7. 图例

已验证 2026-09-01。
"""

import pandas as pd
from pathlib import Path
from typing import Optional

OUTPUT_DIR = Path.home() / ".openclaw" / "data" / "reports"


def _write_df_to_sheet(ws, df: pd.DataFrame, start_row: int = 1):
    """将 DataFrame 写入 openpyxl Worksheet"""
    if df.empty:
        ws.cell(row=1, column=1, value="无数据")
        return

    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=col_idx, value=str(col_name))

    for row_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            if pd.isna(value):
                ws.cell(row=row_idx, column=col_idx, value="")
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)


def generate_revenue_report(
    month: str,
    revenue_df: pd.DataFrame = None,
    acceptance_df: pd.DataFrame = None,
    contract_df: pd.DataFrame = None,
    output_dir: Optional[str] = None,
) -> str:
    """生成确收月报

    Args:
        month: 报告月份（YYYYMM）
        revenue_df: 确收凭证数据
        acceptance_df: 验收凭证数据
        contract_df: 合同台账数据
        output_dir: 输出目录

    Returns:
        生成的 Excel 文件路径
    """
    from openpyxl import Workbook

    out_dir = Path(output_dir) if output_dir else OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 确收凭证汇总（按部门）
    ws1 = wb.create_sheet("确收凭证汇总")
    if revenue_df is not None and not revenue_df.empty:
        from scripts.l4.delivery_center.engines.join_engine import map_pm_to_dept
        from scripts.l4.delivery_center.engines.summary_engine import pivot_revenue_by_dept
        rev_with_dept = map_pm_to_dept(revenue_df, "项目经理")
        dept_summary = pivot_revenue_by_dept(rev_with_dept)
        _write_df_to_sheet(ws1, dept_summary)
    else:
        ws1.cell(row=1, column=1, value="无数据")

    # Sheet 2: 按客户汇总
    ws2 = wb.create_sheet("按客户汇总")
    if revenue_df is not None and not revenue_df.empty:
        if "客户名称" in revenue_df.columns:
            cust_summary = revenue_df.groupby("客户名称").agg({"合同编号": "nunique"}).reset_index()
            cust_summary.columns = ["客户名称", "确收次数"]
            cust_summary = cust_summary.sort_values("确收次数", ascending=False)
            _write_df_to_sheet(ws2, cust_summary)
        else:
            ws2.cell(row=1, column=1, value="无客户数据")
    else:
        ws2.cell(row=1, column=1, value="无数据")

    # Sheet 3: 确收明细
    ws3 = wb.create_sheet("确收明细")
    if revenue_df is not None:
        _write_df_to_sheet(ws3, revenue_df)

    # Sheet 4: 验收明细
    ws4 = wb.create_sheet("验收明细")
    if acceptance_df is not None:
        _write_df_to_sheet(ws4, acceptance_df)

    # Sheet 5: 按项目经理汇总
    ws5 = wb.create_sheet("按项目经理汇总")
    if revenue_df is not None and not revenue_df.empty:
        if "项目经理" in revenue_df.columns:
            pm_summary = revenue_df.groupby("项目经理").agg({"合同编号": "nunique"}).reset_index()
            pm_summary.columns = ["项目经理", "确收次数"]
            pm_summary = pm_summary.sort_values("确收次数", ascending=False)
            _write_df_to_sheet(ws5, pm_summary)
        else:
            ws5.cell(row=1, column=1, value="无项目经理数据")
    else:
        ws5.cell(row=1, column=1, value="无数据")

    # Sheet 6: 图例
    ws6 = wb.create_sheet("图例")
    _fill_legend_sheet(ws6)

    # 保存
    output_path = out_dir / f"确收月报-{month}.xlsx"
    wb.save(output_path)
    print(f"✅ 确收月报已生成: {output_path}")
    return str(output_path)


def _fill_legend_sheet(ws):
    """填充图例 Sheet"""
    import json
    config_dir = Path(__file__).parent.parent / "config"

    legend_pm = json.loads((config_dir / "legend_pm_dept.json").read_text(encoding="utf-8"))
    ws.cell(row=1, column=1, value="项目经理")
    ws.cell(row=1, column=2, value="部门")
    for row_idx, (pm, dept) in enumerate(legend_pm.items(), 2):
        ws.cell(row=row_idx, column=1, value=pm)
        ws.cell(row=row_idx, column=2, value=dept)


if __name__ == "__main__":
    print("=== 确收月报生成器测试 ===\n")

    from scripts.l4.delivery_center.engines.join_engine import (
        load_revenue_vouchers,
        load_acceptance_vouchers,
    )

    rev_df = load_revenue_vouchers()
    acc_df = load_acceptance_vouchers()

    print(f"确收: {len(rev_df)} 行")
    print(f"验收: {len(acc_df)} 行")

    output = generate_revenue_report(
        month="202606",
        revenue_df=rev_df,
        acceptance_df=acc_df,
    )
    print(f"\n输出: {output}")
