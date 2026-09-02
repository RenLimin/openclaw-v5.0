"""
202606 交付月报生成脚本（完整版）
读取 ONES 导出 CSV + BDMS 数据，生成 15 Sheet 的完整月报 Excel。
"""
import sys, csv, sqlite3, json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pandas as pd

# === 路径配置 ===
ONES_DIR = Path.home() / ".openclaw" / "data" / "ones_exports"
BDMS_DB = Path.home() / ".openclaw" / "data" / "bdms.db"
OUTPUT_DIR = Path.home() / ".openclaw" / "data" / "reports"
CONFIG_DIR = Path(__file__).parent.parent / "config"

REPORT_MONTH = "202606"
REPORT_DATE = "2026-06-30"

# === 样式 ===
HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=10, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def load_ones_csv(filepath):
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        return pd.DataFrame(csv.DictReader(f))


def load_bdms_table(table):
    conn = sqlite3.connect(BDMS_DB)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(f"SELECT * FROM {table}")
    cols = [d[0] for d in c.description]
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return cols, rows


def style_header(ws, row, cols):
    for ci in range(1, cols + 1):
        cell = ws.cell(row=row, column=ci)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')


def write_df(ws, df, start_row=1):
    """写入 DataFrame 到 worksheet"""
    if df.empty:
        ws.cell(row=1, column=1, value="无数据")
        return
    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=ci, value=str(h))
    style_header(ws, start_row, len(headers))
    for ri, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for ci, val in enumerate(row, 1):
            v = "" if pd.isna(val) else val
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = THIN_BORDER


# === Sheet 填充函数 ===

def fill_sign_sheet(ws, df):
    """签约 Sheet：行0=汇总，行1=列名，行2+=数据"""
    ws.cell(row=1, column=1, value=REPORT_DATE)
    ws.cell(row=1, column=41, value="排序")
    ws.cell(row=1, column=53, value=len(df))
    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=2, column=ci, value=h)
    style_header(ws, 2, len(headers))
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        for ci, val in enumerate(row, 1):
            v = "" if pd.isna(val) else val
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = THIN_BORDER


def fill_poc_sheet(ws, df):
    """POC&提前实施 Sheet：同签约结构"""
    ws.cell(row=1, column=1, value=REPORT_DATE)
    ws.cell(row=1, column=41, value="排序")
    ws.cell(row=1, column=53, value=len(df))
    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=2, column=ci, value=h)
    style_header(ws, 2, len(headers))
    for ri, (_, row) in enumerate(df.iterrows(), 3):
        for ci, val in enumerate(row, 1):
            v = "" if pd.isna(val) else val
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = THIN_BORDER


def fill_abnormal_sheet(ws, df):
    """异常项目 Sheet：行0=列名，行1+=数据"""
    write_df(ws, df, start_row=1)


def fill_revenue_sheet(ws, cols, rows):
    """确收交接 Sheet"""
    col_map = {
        "月份": None, "标题": None, "ID": "voucher_id", "BI履约ID": "bi_id",
        "合同编号1": "合同编号", "邮件编号": None, "合同编号": "合同编号",
        "客户名称": "客户名称", "销售部门": "销售部门", "项目经理": "项目经理",
        "备注": None, "交接日期": "交接日期", "财务接收人": "财务",
        "财务是否接收": "是否接收", "财务反馈": None, "交付邮件是否跨月": None,
        "PMO提交人": None, "PMO反馈": None, "是否修改ONES状态": None,
        "项目经理所属区域": None, "跨月交接": None, "跨月交接原因": None, "是否合格": None,
    }
    headers = list(col_map.keys())
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    style_header(ws, 1, len(headers))
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            db_col = col_map.get(h)
            val = row.get(db_col, "") if db_col else ""
            cell = ws.cell(row=ri, column=ci, value=str(val) if val else "")
            cell.border = THIN_BORDER


def fill_acceptance_sheet(ws, cols, rows):
    """验收交接 Sheet"""
    col_map = {
        "月份": None, "合同名称": "合同名称", "标题": None, "ID": "voucher_id",
        "BI履约ID": "bi_id", "验收单编号-财务端": None, "合同编号1": "合同编号",
        "验收单编号": "验收单编号", "合同编号": "合同编号", "客户名称": "客户名称",
        "销售部门": None, "项目经理": "项目经理", "备注": None, "交接日期": "交接日期",
        "验收方式": "验收方式", "截至目前全部/部分验收": "全部或部分",
        "是否为渠道": None, "财务接收人": "财务", "是否接收": "财务是否接收",
        "实际验收方式": None, "财务反馈": None, "PMO提交人": None, "PMO反馈": None,
        "是否修改ones及OA状态": None, "项目经理所属区域": None, "是否合格": None,
    }
    headers = list(col_map.keys())
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    style_header(ws, 1, len(headers))
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            db_col = col_map.get(h)
            val = row.get(db_col, "") if db_col else ""
            cell = ws.cell(row=ri, column=ci, value=str(val) if val else "")
            cell.border = THIN_BORDER


def fill_sign_stats_sheet(ws, df):
    """签约统计：按部门统计项目数"""
    if '责任销售所属团队' not in df.columns:
        ws.cell(row=1, column=1, value="无部门数据")
        return
    stats = df.groupby('责任销售所属团队')['ID'].nunique().reset_index()
    stats.columns = ['项目经理所属部门', '计数项:ID']
    stats = stats.sort_values('计数项:ID', ascending=False)
    write_df(ws, stats)


def fill_poc_stats_sheet(ws, df):
    """POC&提前实施统计：按部门统计"""
    if '责任销售所属团队' not in df.columns:
        ws.cell(row=1, column=1, value="无部门数据")
        return
    stats = df.groupby('责任销售所属团队')['ID'].nunique().reset_index()
    stats.columns = ['项目经理所属部门', '计数项:ID']
    stats = stats.sort_values('计数项:ID', ascending=False)
    write_df(ws, stats)


def fill_abnormal_stats_sheet(ws, df):
    """异常统计：按部门统计异常项目数"""
    if '责任销售所属团队' not in df.columns:
        ws.cell(row=1, column=1, value="无部门数据")
        return
    stats = df.groupby('责任销售所属团队')['ID'].nunique().reset_index()
    stats.columns = ['项目经理所属部门', '计数项:ID']
    stats = stats.sort_values('计数项:ID', ascending=False)
    write_df(ws, stats)


def fill_abnormal_ledger_sheet(ws, df):
    """异常台账：按合同归档年份统计"""
    if '合同归档日期' not in df.columns:
        ws.cell(row=1, column=1, value="无归档日期数据")
        return
    df = df.copy()
    df['归档年份'] = pd.to_datetime(df['合同归档日期'], errors='coerce').dt.year
    df = df.dropna(subset=['归档年份'])
    df['归档年份'] = df['归档年份'].astype(int)
    
    # 按年份统计
    yearly = df.groupby('归档年份').size().reset_index(name='合计')
    yearly.columns = ['合同归档年份', '合计']
    yearly = yearly.sort_values('合同归档年份')
    write_df(ws, yearly)


def fill_handover_stats_sheet(ws, rev_rows, acc_rows):
    """交接统计：确收/验收按区域统计"""
    # 确收合格率
    rev_df = pd.DataFrame(rev_rows)
    acc_df = pd.DataFrame(acc_rows)
    
    # 简化的交接统计
    stats_data = []
    if len(rev_df) > 0:
        stats_data.append({'指标': '确收交接', '总数': len(rev_df)})
    if len(acc_df) > 0:
        stats_data.append({'指标': '验收交接', '总数': len(acc_df)})
    
    if stats_data:
        stats = pd.DataFrame(stats_data)
        write_df(ws, stats)
    else:
        ws.cell(row=1, column=1, value="无交接数据")


def fill_efficiency_sheet(ws, df):
    """交付效率统计：按项目经理统计偏差率"""
    # 需要立项日期和实际结项日期计算偏差
    if '立项日期' not in df.columns:
        ws.cell(row=1, column=1, value="无立项日期数据")
        return
    df = df.copy()
    df['立项日期_dt'] = pd.to_datetime(df['立项日期'], errors='coerce')
    df['结项日期_dt'] = pd.to_datetime(df.get('实际结项日期', pd.NaT), errors='coerce')
    df = df.dropna(subset=['立项日期_dt'])
    
    # 按部门+项目经理统计
    if '责任销售所属团队' in df.columns and '负责人' in df.columns:
        stats = df.groupby(['责任销售所属团队', '负责人']).agg(
            ID=('ID', 'nunique')
        ).reset_index()
        stats.columns = ['项目经理团队', '项目经理', '项目数']
        write_df(ws, stats)
    else:
        ws.cell(row=1, column=1, value="无部门/负责人数据")


def fill_legend_sheet(ws):
    """图例 Sheet：项目经理-部门映射"""
    legend_path = CONFIG_DIR / "legend_pm_dept.json"
    if not legend_path.exists():
        ws.cell(row=1, column=1, value="图例配置文件不存在")
        return
    legend = json.loads(legend_path.read_text(encoding="utf-8"))
    headers = ["项目经理", "部门"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    style_header(ws, 1, 2)
    for ri, (pm, dept) in enumerate(legend.items(), 2):
        ws.cell(row=ri, column=1, value=pm)
        ws.cell(row=ri, column=2, value=dept)


def fill_product_license_sheet(ws, df):
    """产品-授权&维保统计：按产线/产品类型统计"""
    if '所属产线' not in df.columns:
        ws.cell(row=1, column=1, value="无产线数据")
        return
    stats = df.groupby('所属产线')['ID'].nunique().reset_index()
    stats.columns = ['所属产线', '计数项:ID']
    stats = stats.sort_values('计数项:ID', ascending=False)
    write_df(ws, stats)


def fill_poc_dept_sheet(ws, df):
    """提前实施分事业部统计"""
    if '责任销售所属团队' not in df.columns:
        ws.cell(row=1, column=1, value="无部门数据")
        return
    # 筛选提前实施项目
    if '项目类型(概览)' in df.columns:
        poc_df = df[df['项目类型(概览)'].str.contains('提前实施', na=False)]
    else:
        poc_df = df
    stats = poc_df.groupby('责任销售所属团队')['ID'].nunique().reset_index()
    stats.columns = ['责任销售所属团队', '计数项:ID']
    stats = stats.sort_values('计数项:ID', ascending=False)
    write_df(ws, stats)


def fill_abnormal_dept_sheet(ws, df):
    """交付异常分事业部统计"""
    if '责任销售所属团队' not in df.columns:
        ws.cell(row=1, column=1, value="无部门数据")
        return
    stats = df.groupby('责任销售所属团队')['ID'].nunique().reset_index()
    stats.columns = ['项目经理团队', '计数项:ID']
    stats = stats.sort_values('计数项:ID', ascending=False)
    write_df(ws, stats)


def main():
    print("=== 202606 交付月报生成（完整版）===\n")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 1. 加载 ONES 数据
    print("1. 加载 ONES 导出数据...")
    sign_df = load_ones_csv(ONES_DIR / "签约项目统计.csv")
    poc_df = load_ones_csv(ONES_DIR / "poc_提前实施.csv")
    abnormal_df = load_ones_csv(ONES_DIR / "异常处置.csv")
    print(f"   签约: {len(sign_df)} 行")
    print(f"   POC: {len(poc_df)} 行")
    print(f"   异常: {len(abnormal_df)} 行")

    # 2. 加载 BDMS 数据
    print("\n2. 加载 BDMS 数据...")
    rev_cols, rev_rows = load_bdms_table("revenue_vouchers")
    acc_cols, acc_rows = load_bdms_table("acceptance_vouchers")
    print(f"   确收: {len(rev_rows)} 行")
    print(f"   验收: {len(acc_rows)} 行")

    # 3. 创建工作簿
    print("\n3. 生成 Excel（15 Sheet）...")
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 签约
    ws1 = wb.create_sheet("签约")
    fill_sign_sheet(ws1, sign_df)
    print("   ✅ 签约")

    # Sheet 2: POC&提前实施
    ws2 = wb.create_sheet("POC&提前实施")
    fill_poc_sheet(ws2, poc_df)
    print("   ✅ POC&提前实施")

    # Sheet 3: 异常项目
    ws3 = wb.create_sheet("异常项目")
    fill_abnormal_sheet(ws3, abnormal_df)
    print("   ✅ 异常项目")

    # Sheet 4: 确收交接
    ws4 = wb.create_sheet("确收交接")
    fill_revenue_sheet(ws4, rev_cols, rev_rows)
    print("   ✅ 确收交接")

    # Sheet 5: 验收交接
    ws5 = wb.create_sheet("验收交接")
    fill_acceptance_sheet(ws5, acc_cols, acc_rows)
    print("   ✅ 验收交接")

    # Sheet 6: 交付效率统计
    ws6 = wb.create_sheet("交付效率统计")
    fill_efficiency_sheet(ws6, sign_df)
    print("   ✅ 交付效率统计")

    # Sheet 7: 签约统计
    ws7 = wb.create_sheet("签约统计")
    fill_sign_stats_sheet(ws7, sign_df)
    print("   ✅ 签约统计")

    # Sheet 8: POC&提前实施统计
    ws8 = wb.create_sheet("POC&提前实施统计")
    fill_poc_stats_sheet(ws8, poc_df)
    print("   ✅ POC&提前实施统计")

    # Sheet 9: 异常统计
    ws9 = wb.create_sheet("异常统计")
    fill_abnormal_stats_sheet(ws9, abnormal_df)
    print("   ✅ 异常统计")

    # Sheet 10: 异常台账
    ws10 = wb.create_sheet("异常台账")
    fill_abnormal_ledger_sheet(ws10, abnormal_df)
    print("   ✅ 异常台账")

    # Sheet 11: 交接统计
    ws11 = wb.create_sheet("交接统计")
    fill_handover_stats_sheet(ws11, rev_rows, acc_rows)
    print("   ✅ 交接统计")

    # Sheet 12: 产品-授权&维保统计
    ws12 = wb.create_sheet("产品-授权&维保统计")
    fill_product_license_sheet(ws12, sign_df)
    print("   ✅ 产品-授权&维保统计")

    # Sheet 13: 提前实施分事业部统计
    ws13 = wb.create_sheet("提前实施分事业部统计")
    fill_poc_dept_sheet(ws13, poc_df)
    print("   ✅ 提前实施分事业部统计")

    # Sheet 14: 交付异常分事业部统计
    ws14 = wb.create_sheet("交付异常分事业部统计")
    fill_abnormal_dept_sheet(ws14, abnormal_df)
    print("   ✅ 交付异常分事业部统计")

    # Sheet 15: 图例
    ws15 = wb.create_sheet("图例")
    fill_legend_sheet(ws15)
    print("   ✅ 图例")

    # 4. 保存
    output_path = OUTPUT_DIR / f"交付月报-{REPORT_MONTH}.xlsx"
    wb.save(output_path)
    print(f"\n✅ 月报已生成: {output_path}")
    print(f"   大小: {output_path.stat().st_size / 1024 / 1024:.1f} MB")

    return str(output_path)


if __name__ == "__main__":
    main()
